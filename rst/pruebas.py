# -*- coding: utf-8 -*-
"""Regresión del motor del SIMPLE contra el caso modelo hecho a mano.

Debe decir «TODAS LAS PRUEBAS PASAN» después de cualquier cambio al lector, a
los parámetros o a la liquidación.

    python -m rst.pruebas
"""
import os
import sys

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(BASE))

from rst import calculos
from rst import generar
from rst import lector
from rst import libro
from rst import parametros as P
from rst import plazos

# El caso de referencia —archivo fuente y cifras verificadas a mano— vive en
# `rst\caso_referencia.py`, que NO se versiona: son datos de un contribuyente
# real, sujetos a la reserva del art. 583 E.T. Sin ese archivo la suite corre
# igual y solo omite el bloque del caso real.
try:
    from rst import caso_referencia as CASO
except ImportError:
    CASO = None

fallos = []


def check(nombre, obtenido, esperado, tol=0.02):
    if abs(obtenido - esperado) > tol:
        fallos.append('%s: esperado %s, obtenido %s' % (nombre, esperado, obtenido))


def probar_parametros():
    check('UVT 2026', P.uvt(2026), 52374, 0)
    check('tarifa grupo 3 tramo 1', P.tarifa(3, 884.6), 0.059, 0)
    check('tarifa grupo 3 tramo 2', P.tarifa(3, 1500), 0.073, 0)
    check('tarifa grupo 3 tramo 4', P.tarifa(3, 20000), 0.145, 0)
    check('tarifa grupo 2 tramo 1', P.tarifa(2, 884.6), 0.016, 0)
    # Por encima del último tramo se mantiene la tarifa mayor, nunca cero: una
    # base de 18.000 UVT con tarifa 0% daba un anticipo de $0. El libro tiene
    # que hacer lo mismo que esto (ver `probar_libro`).
    for g in P.TARIFAS:
        check('tarifa grupo %d sobre el tope' % g, P.tarifa(g, 99999),
              P.TARIFAS[g][-1][2], 0)
        check('tarifa grupo %d con base cero' % g, P.tarifa(g, 0),
              P.TARIFAS[g][0][2], 0)
    if P.nombre_bimestre(3) != 'Mayo y Junio':
        fallos.append('nombre del bimestre 3 incorrecto')
    if P.meses_bimestre(6) != (11, 12):
        fallos.append('meses del bimestre 6 incorrectos')
    try:
        P.uvt(2099)
        fallos.append('uvt() debería fallar con un año sin resolución cargada')
    except KeyError:
        pass
    try:
        P.tarifa(9, 100)
        fallos.append('tarifa() debería fallar con un grupo inexistente')
    except KeyError:
        pass


def probar_caso_real():
    """El caso de referencia, si está disponible en este equipo."""
    if CASO is None:
        print('  (sin rst/caso_referencia.py: se omite el caso real)')
        return None
    if not os.path.exists(CASO.CONSOLIDADO):
        fallos.append('No se encuentra el consolidado de referencia: %s' % CASO.CONSOLIDADO)
        return None
    ficha = generar.cargar_ficha(CASO.FICHA)
    fuente = lector.leer(CASO.CONSOLIDADO)

    for clave, valor in CASO.CONTEOS.items():
        check(clave, len(fuente[clave]), valor, 0)

    liq = calculos.liquidar(ficha, fuente)
    for clave, valor in CASO.ESPERADO.items():
        check(clave, liq[clave], valor, 0.02 if clave != 'base_uvt' else 0.01)

    for v in liq['validaciones']:
        if not v['ok']:
            fallos.append('validación en rojo que debería cuadrar: %s' % v['nombre'])
    if len(liq['validaciones']) < 10:
        fallos.append('se esperaban al menos 10 validaciones, hay %d' % len(liq['validaciones']))
    if liq['semaforo'] != 'AMARILLO':
        fallos.append('semáforo esperado AMARILLO (hay alertas ALTA), obtenido %s'
                      % liq['semaforo'])
    temas = {a['tema'] for a in liq['alertas']}
    for esperado in ('Grupo de actividad', 'Diferencias en certificados de ReteIVA',
                     'Completitud de las planillas de pensión', 'DV y Dirección Seccional'):
        if esperado not in temas:
            fallos.append('falta la alerta «%s»' % esperado)
    return liq, ficha


def probar_plazos():
    """El calendario del SIMPLE va por el ÚLTIMO dígito del NIT, uno solo.

    En renta son los dos últimos: confundirlos daría fechas equivocadas, y una
    fecha de vencimiento mal puesta cuesta sanción por extemporaneidad.
    """
    import datetime
    check('bim 3, NIT terminado en 2', 
          plazos.vencimiento(2026, 3, '901555552').toordinal(),
          datetime.date(2026, 7, 10).toordinal(), 0)
    check('bim 1, NIT terminado en 0',
          plazos.vencimiento(2026, 1, '900000000').toordinal(),
          datetime.date(2026, 5, 26).toordinal(), 0)
    # El sexto bimestre se paga en enero del año SIGUIENTE.
    f = plazos.vencimiento(2026, 6, '900000001')
    check('bim 6 cae en enero del año siguiente', f.toordinal(),
          datetime.date(2027, 1, 13).toordinal(), 0)
    # El dígito de verificación no cuenta para el plazo.
    if plazos.vencimiento(2026, 3, '901555552-7') != plazos.vencimiento(2026, 3, '901555552'):
        fallos.append('el DV no debe alterar el plazo')
    if plazos.vencimiento(2099, 1, '900000000') is not None:
        fallos.append('un año sin calendario cargado debe devolver None, no inventar fecha')
    if 'No hay calendario' not in plazos.texto(2099, 1, '900000000'):
        fallos.append('sin calendario, el texto debe decirlo')


def probar_modo_crudo():
    """Lo importante: el archivo crudo de la DIAN basta.

    Las hojas F.VENTA/F.COMPRA/RETEIVA son trabajo manual redundante; el motor
    deriva el desglose de ingresos, las compras y la conciliación a partir de
    Rp_Doc_… y Rp_Docpras, y tiene que dar exactamente lo mismo.
    """
    if CASO is None or not getattr(CASO, 'CRUDO', None):
        return
    if not os.path.exists(CASO.CRUDO):
        fallos.append('No se encuentra el archivo crudo de referencia: %s' % CASO.CRUDO)
        return
    ficha = generar.cargar_ficha(CASO.FICHA)
    ficha['aporte_pension_total'] = CASO.APORTE_PENSION
    fuente = lector.leer(CASO.CRUDO)
    if fuente['origen'] != 'crudo':
        fallos.append('el lector no reconoció el archivo como crudo')
    liq = calculos.liquidar(ficha, fuente)
    for clave, valor in CASO.ESPERADO.items():
        if clave == 'reteiva_certificados':
            continue            # el certificado no está en ningún archivo de la DIAN
        check('crudo · ' + clave, liq[clave], valor,
              0.02 if clave != 'base_uvt' else 0.01)
    for v in liq['validaciones']:
        if not v['ok']:
            fallos.append('crudo · validación en rojo: %s' % v['nombre'])
    # Una razón derivada no puede fallar: no debe contar como comprobación crítica.
    derivadas = [v for v in liq['validaciones']
                 if v['nombre'].startswith(('IVA generado', 'ReteIVA ÷'))]
    if any(v.get('critica') for v in derivadas):
        fallos.append('en modo crudo las razones derivadas no pueden marcarse críticas: '
                      'no pueden fallar y darían un verde falso')


def probar_libro(liq, ficha):
    wb = libro.construir(liq, ficha)
    if wb.sheetnames != libro.ORDEN:
        fallos.append('el orden de las hojas cambió: %s' % wb.sheetnames)
    datos = libro.a_bytes(wb)
    if len(datos) < 10000:
        fallos.append('el libro generado pesa %d bytes: parece vacío' % len(datos))
    ws = wb['2.INGRESOS']
    # el detalle debe traer una fila por factura más encabezado y totales
    if sum(1 for f in ws.iter_rows() if f[0].value == 'TOTALES') != 1:
        fallos.append('la hoja 2 no tiene exactamente una fila de TOTALES')

    # La fórmula de la tarifa tiene que acotar la base al rango de la tabla. Sin
    # esa cota, una base por encima del último tramo no casaba con ninguna fila
    # del SUMIFS: Excel devolvía 0%, el anticipo salía $0 y el libro —que es el
    # papel de trabajo definitivo— contradecía a la liquidación de Python.
    tarifa = wb['1.LIQUIDACIÓN 2593']['F24'].value or ''
    if 'MIN(MAX(' not in tarifa:
        fallos.append('la fórmula de la tarifa (F24) no acota la base al rango de la '
                      'tabla: por encima del último tramo daría 0%')
    # El descuento de pensión no puede operar sobre un componente nacional
    # negativo: el MIN se quedaría con el negativo y la resta lo volvería suma.
    pension = wb['1.LIQUIDACIÓN 2593']['F30'].value or ''
    if 'MAX(0,F29)' not in pension:
        fallos.append('la fórmula del descuento de pensión (F30) no protege el caso de '
                      'componente nacional negativo')


def _liq_minima(**cambios):
    liq = {'ingresos_gravados': 100, 'iva_generado': 19, 'reteiva': 2.85,
           'reteiva_certificados': 2.85, 'facturas': [], 'ingresos_no_gravados': 0,
           'iva_descontable': 0, 'validaciones': [], 'alertas': [],
           'facturas_fuera_periodo': [], 'periodos_del_archivo': [],
           'ano': 2026, 'bimestre': 3, 'nombre_bimestre': 'Mayo y Junio'}
    liq.update(cambios)
    return liq


def probar_semaforo_rojo():
    """Si las razones no cuadran, el semáforo tiene que ponerse en ROJO."""
    liq = _liq_minima(iva_generado=5, reteiva=0.75, reteiva_certificados=0.75)
    liq['validaciones'] = calculos.validar(liq, {}, {'reteiva': [], 'totales_declarados': {}})
    if calculos.semaforo(liq) != 'ROJO':
        fallos.append('una razón IVA del 5%% debería dar ROJO, dio %s'
                      % calculos.semaforo(liq))


def probar_periodo_vacio_es_rojo():
    """El fallo que motivó todo esto: un archivo de mayo-junio liquidado como
    bimestre 1 salía en ceros con semáforo AMARILLO, como si fuera válido.

    Un período sin un solo documento, teniendo el archivo documentos de otros
    períodos, es un período mal escogido. Va en ROJO.
    """
    fuera = [{'fecha': None} for _ in range(8270)]
    liq = _liq_minima(bimestre=1, nombre_bimestre='Enero y Febrero',
                      ingresos_gravados=0, iva_generado=0, reteiva=0,
                      reteiva_certificados=0, facturas_fuera_periodo=fuera,
                      periodos_del_archivo=[{'ano': 2026, 'bimestre': 3,
                                             'documentos': 8270, 'total': 1.0}])
    liq['validaciones'] = calculos.validar(liq, {}, {'reteiva': [], 'totales_declarados': {}})
    if calculos.semaforo(liq) != 'ROJO':
        fallos.append('un bimestre sin documentos teniendo el archivo otros períodos '
                      'debe dar ROJO, dio %s' % calculos.semaforo(liq))

    # Pero un bimestre realmente sin ventas SÍ se puede declarar en ceros: el
    # SIMPLE obliga a presentar el anticipo aunque no haya habido ingresos.
    vacio = _liq_minima(ingresos_gravados=0, iva_generado=0, reteiva=0,
                        reteiva_certificados=0)
    vacio['validaciones'] = calculos.validar(vacio, {}, {'reteiva': [],
                                                         'totales_declarados': {}})
    cobertura = [v for v in vacio['validaciones'] if v['nombre'].startswith('Documentos del')]
    if not cobertura or not cobertura[0]['ok']:
        fallos.append('un bimestre sin ninguna venta en todo el archivo debe poder '
                      'declararse en ceros, no marcarse en rojo')


def probar_clasificacion_documentos():
    """Las notas crédito tienen que RESTAR, y los documentos soporte no son ingreso.

    El literal «nota crédito» con tilde nunca coincidía contra el tipo ya
    normalizado sin tildes, así que las 29 notas crédito de un caso real
    sumaban al ingreso en vez de restarlo: el doble de su valor, de más.
    """
    casos = [
        ('Nota de crédito electrónica', 'emitidos', 'venta', -1),
        ('Nota Crédito electrónica', 'emitidos', 'venta', -1),
        ('NOTA CREDITO', 'emitidos', 'venta', -1),
        ('Nota de débito electrónica', 'emitidos', 'venta', 1),
        ('Factura electrónica', 'emitidos', 'venta', 1),
        # Documento soporte y sus notas de ajuste son COMPRAS a no obligados,
        # aunque el contribuyente sea quien los emite.
        ('Documento soporte con no obligados', 'emitidos', 'compra', 1),
        ('Nota de ajuste del documento soporte', 'emitidos', 'compra', 1),
        ('Application response', 'emitidos', None, 1),
        ('Nómina electrónica', 'emitidos', None, 1),
        # Lo recibido es compra, y ahí el documento soporte también lo es.
        ('Factura electrónica', 'recibidos', 'compra', 1),
        ('Documento soporte con no obligados', 'recibidos', 'compra', 1),
        ('Nota de crédito electrónica', 'recibidos', 'compra', -1),
        ('Application response', 'recibidos', None, 1),
    ]
    for tipo, lado, destino_esp, signo_esp in casos:
        destino, signo = lector.clasificar(tipo, lado)
        if destino != destino_esp or (destino and signo != signo_esp):
            fallos.append('clasificar(%r, %r) dio (%r, %s), se esperaba (%r, %s)'
                          % (tipo, lado, destino, signo, destino_esp, signo_esp))


def _archivo_dian(filas, hoja='Rp_Doc_20260801_1708'):
    """Un consolidado crudo de la DIAN en memoria, con su maquetación real."""
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(['Tipo de documento', 'CUFE/CUDE', 'Folio', 'Prefijo', 'Divisa',
               'Forma de Pago', 'Medio de Pago', 'Fecha Emisión', 'Fecha Recepción',
               'NIT Emisor', 'Nombre Emisor', 'NIT Receptor', 'Nombre Receptor',
               'IVA', 'ICA', 'IC', 'INC', 'Timbre', 'INC Bolsas', 'IN Carbono',
               'IN Combustibles', 'IC Datos', 'ICL', 'INPP', 'IBUA', 'ICUI',
               'Rete IVA', 'Rete Renta', 'Rete ICA', 'Total', 'Estado', 'Grupo'])
    for tipo, ne, nome, nr, nomr, iva, total, grupo in filas:
        f = [tipo, 'CUFE', '1', 'L', 'COP', '1', '47', '15-06-2026', '15-06-2026',
             ne, nome, nr, nomr, iva] + [0] * 15 + [total, 'Aprobado', grupo]
        ws.append(f)
    datos = io.BytesIO()
    wb.save(datos)
    datos.seek(0)
    return datos


def probar_reparto_ventas_compras():
    """Ventas y compras se separan por el NIT, no por la hoja.

    La exportación de la DIAN mete emitidos y recibidos en la MISMA hoja
    `Rp_Doc_…`. Tomando la hoja entera como ventas, las facturas que el
    contribuyente RECIBIÓ se sumaban al ingreso —inflando la base del
    anticipo— y de paso su IVA no llegaba al descontable. El error corría en
    los dos sentidos y siempre en contra del contribuyente.
    """
    YO, NOMBRE = '900711617', 'LUMASEG CONSULTORES EN SEGUROS LTDA'
    filas = [
        # tipo, NIT emisor, nombre emisor, NIT receptor, nombre receptor, IVA, total, grupo
        ('Factura electrónica', YO, NOMBRE, '890903790', 'SURAMERICANA',
         19000, 119000, 'Emitido'),
        ('Documento soporte con no obligados', YO, NOMBRE, '91280239', 'JAIME MARTINEZ',
         0, 500000, 'Emitido'),
        ('Nomina Individual', YO, NOMBRE, '1098780744', 'LINEY MEDINA',
         0, 2000000, 'Emitido'),
        ('Factura electrónica', '830122566', 'COLOMBIA TELECOMUNICACIONES', YO, NOMBRE,
         1900, 11900, 'Recibido'),
        ('Nota de crédito electrónica', '830122566', 'COLOMBIA TELECOMUNICACIONES',
         YO, NOMBRE, 190, 1190, 'Recibido'),
    ]
    f = lector.leer(_archivo_dian(filas), YO)

    check('reparto · ventas', len(f['ventas']), 1, 0)
    check('reparto · compras', len(f['compras']), 3, 0)
    # La factura recibida NO puede engordar el ingreso: 100.000, no 110.000.
    check('reparto · ingreso gravado', sum(v['gravado'] for v in f['ventas']), 100000, 0.02)
    # …y su IVA sí tiene que llegar al descontable, con la nota crédito restando.
    check('reparto · IVA descontable', sum(c['iva'] for c in f['compras']), 1710, 0.02)
    check('reparto · base de compras', sum(c['base'] for c in f['compras']), 9000, 0.02)

    if f['ventas'][0]['tercero'] != 'SURAMERICANA':
        fallos.append('en una venta el tercero es el RECEPTOR, dio %r'
                      % f['ventas'][0]['tercero'])
    proveedores = {c['proveedor'] for c in f['compras']}
    if 'COLOMBIA TELECOMUNICACIONES' not in proveedores:
        fallos.append('en una compra recibida el proveedor es el EMISOR, dio %r' % proveedores)
    # El documento soporte lo emite el comprador: es compra, y el proveedor es
    # el no obligado que figura como receptor del documento.
    if 'JAIME MARTINEZ' not in proveedores:
        fallos.append('el documento soporte con no obligados es una COMPRA a la persona '
                      'natural, dio %r' % proveedores)

    # Sin NIT el reparto se sostiene con la columna «Grupo» de la DIAN.
    g = lector.leer(_archivo_dian(filas))
    check('reparto sin NIT · ventas', len(g['ventas']), 1, 0)
    check('reparto sin NIT · compras', len(g['compras']), 3, 0)

    # Un consolidado de otro contribuyente tiene que avisarse, no procesarse en
    # silencio como si fuera del cliente que se está declarando.
    otro = lector.leer(_archivo_dian(filas), '901555552')
    if not any('no aparece en ninguna fila' in a for a in otro['avisos']):
        fallos.append('subir el consolidado de otro NIT debe avisarse: %r' % otro['avisos'])


def probar_deteccion_periodo():
    """El bimestre sale de las fechas del archivo: no hay que acertarlo."""
    import datetime

    def venta(y, m, total=100.0):
        return {'fecha': datetime.date(y, m, 15), 'total': total}

    ventas = [venta(2026, 5), venta(2026, 6), venta(2026, 6), venta(2026, 2)]
    p = lector.periodos(ventas)
    if not p or (p[0]['ano'], p[0]['bimestre'], p[0]['documentos']) != (2026, 3, 3):
        fallos.append('el período dominante de un archivo de mayo-junio debe ser el '
                      'bimestre 3 con 3 documentos, dio %r' % (p[:1],))
    if len(p) != 2 or p[1]['bimestre'] != 1:
        fallos.append('los períodos secundarios deben quedar listados, dio %r' % (p,))

    fuente = {'periodos': p}
    ficha = generar.resolver_periodo({'ano': 2026, 'bimestre': None}, fuente)
    if (ficha['ano'], ficha['bimestre']) != (2026, 3):
        fallos.append('sin bimestre en la ficha debe tomarse el dominante, dio %r'
                      % ((ficha['ano'], ficha['bimestre']),))
    if not ficha.get('_periodo_auto'):
        fallos.append('la detección automática debe quedar anotada en la ficha')

    # Un bimestre pedido a mano que el archivo no tiene: aborta con mensaje.
    try:
        generar.resolver_periodo({'ano': 2026, 'bimestre': 5}, fuente)
        fallos.append('pedir un bimestre que el archivo no trae debe abortar, no '
                      'devolver un libro en ceros')
    except ValueError as ex:
        if 'bimestre 3' not in str(ex):
            fallos.append('el error del período equivocado debe nombrar el correcto: %s' % ex)


def main():
    probar_parametros()
    probar_clasificacion_documentos()
    probar_reparto_ventas_compras()
    probar_deteccion_periodo()
    resultado = probar_caso_real()
    if resultado:
        probar_libro(*resultado)
    probar_plazos()
    probar_modo_crudo()
    probar_semaforo_rojo()
    probar_periodo_vacio_es_rojo()

    if fallos:
        print('FALLARON %d COMPROBACIONES:' % len(fallos))
        for f in fallos:
            print('  · %s' % f)
        return 1
    print('TODAS LAS PRUEBAS PASAN')
    return 0


if __name__ == '__main__':
    sys.exit(main())
