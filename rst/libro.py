# -*- coding: utf-8 -*-
"""Arma el libro de trabajo del SIMPLE: 9 hojas, en memoria.

Mismo criterio del libro de renta: las celdas que el contador puede tocar van
en crema/amarillo y todo lo que dependa de ellas es FÓRMULA, no un número
pegado. Si el contador corrige la tarifa de ICA o agrega una planilla, la
liquidación entera se recalcula sola dentro de Excel.
"""
import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import parametros as P
from . import plazos

# Paleta del libro del SIMPLE (azul), distinta de la del libro de renta (verde):
# son dos entregables diferentes y conviene que se distingan de un vistazo.
AZUL = '1F3864'          # títulos de hoja
AZUL_MEDIO = '2E5C8A'    # encabezados de sección y de tabla
AZUL_CLARO = 'D9E2F3'    # subtotales
VERDE_CLARO = 'E2EFDA'   # resultados
CREMA = 'FFF2CC'         # EDITABLE
ROSA = 'FCE4E4'          # diferencia detectada
CEBRA = 'F2F2F2'

MONEDA = '#,##0;[Red]-#,##0'
MONEDA2 = '#,##0.00;[Red]-#,##0.00'
PORCENTAJE = '0.00%'
TARIFA_ICA_FMT = '0.0000'
UVT_FMT = '#,##0.00'
ENTERO = '#,##0'

_fino = Side(style='thin', color='BFBFBF')
BORDE = Border(left=_fino, right=_fino, top=_fino, bottom=_fino)


# ------------------------------------------------------------------ utilidades

def _c(ws, coord, valor, *, negrita=False, tam=10, color=None, fondo=None,
       fmt=None, alin=None, vert=None, ajuste=None, borde=False):
    celda = ws[coord]
    celda.value = valor
    celda.font = Font(bold=negrita, size=tam, color=color)
    if fondo:
        celda.fill = PatternFill('solid', fgColor=fondo)
    if fmt:
        celda.number_format = fmt
    if alin or vert or ajuste:
        celda.alignment = Alignment(horizontal=alin, vertical=vert, wrap_text=ajuste)
    if borde:
        celda.border = BORDE
    return celda


def _titulo(ws, fila, texto, ancho, tam=13):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho)
    _c(ws, 'A%d' % fila, texto, negrita=True, tam=tam, color='FFFFFF', fondo=AZUL,
       vert='center')
    ws.row_dimensions[fila].height = 24


def _seccion(ws, fila, texto, ancho):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho)
    _c(ws, 'A%d' % fila, texto, negrita=True, color='FFFFFF', fondo=AZUL_MEDIO,
       vert='center')
    ws.row_dimensions[fila].height = 18


def _encabezados(ws, fila, textos, desde=1):
    for i, t in enumerate(textos):
        _c(ws, '%s%d' % (get_column_letter(desde + i), fila), t, negrita=True, tam=9,
           color='FFFFFF', fondo=AZUL_MEDIO, alin='center', vert='center',
           ajuste=True, borde=True)
    ws.row_dimensions[fila].height = 30


def _anchos(ws, anchos):
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w


def _hoja(wb, nombre, anchos, congelar=None):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    _anchos(ws, anchos)
    if congelar:
        ws.freeze_panes = congelar
    return ws


def _pesos(x):
    return f'{x:,.0f}'.replace(',', '.')


# ------------------------------------------------------- 2. INGRESOS

def hoja_ingresos(wb, liq, ficha, A):
    ws = _hoja(wb, '2.INGRESOS', {'A': 11, 'B': 8, 'C': 14, 'D': 40, 'E': 16, 'F': 16,
                                  'G': 16, 'H': 15, 'I': 16}, congelar='A5')
    _titulo(ws, 1, 'CLASIFICACIÓN DE LOS INGRESOS — Bimestre %d (%s %d)'
            % (liq['bimestre'], liq['nombre_bimestre'].lower(), liq['ano']), 9)
    _seccion(ws, 3, 'A. RESUMEN POR MES Y POR NATURALEZA FRENTE AL IVA', 9)
    _encabezados(ws, 4, ['Mes', 'N° fact.', 'Ingresos GRAVADOS con IVA',
                         'Ingresos NO GRAVADOS / excluidos de IVA',
                         'INGRESOS BRUTOS (base SIMPLE)', 'IVA generado %s'
                         % ('%g%%' % (P.TARIFA_IVA * 100)),
                         'ReteIVA que le practicaron %g%%' % (P.TARIFA_RETEIVA * 100),
                         'Total facturado', 'Neto recibido'])

    fila = 5
    for m in liq['por_mes']:
        _c(ws, 'A%d' % fila, '%s %d' % (m['nombre'], liq['ano']), borde=True)
        _c(ws, 'B%d' % fila, m['facturas'], alin='center', borde=True)
        _c(ws, 'C%d' % fila, m['gravado'], fmt=MONEDA, borde=True)
        _c(ws, 'D%d' % fila, m['no_gravado'], fmt=MONEDA, borde=True)
        _c(ws, 'E%d' % fila, '=C%d+D%d' % (fila, fila), fmt=MONEDA, borde=True)
        _c(ws, 'F%d' % fila, m['iva'], fmt=MONEDA, borde=True)
        _c(ws, 'G%d' % fila, m['reteiva'], fmt=MONEDA, borde=True)
        _c(ws, 'H%d' % fila, '=E%d+F%d' % (fila, fila), fmt=MONEDA, borde=True)
        _c(ws, 'I%d' % fila, '=H%d-G%d' % (fila, fila), fmt=MONEDA, borde=True)
        fila += 1

    total = fila
    _c(ws, 'A%d' % total, 'TOTAL BIMESTRE', negrita=True, color='FFFFFF', fondo=AZUL,
       borde=True)
    for col in 'BCDEFGHI':
        _c(ws, '%s%d' % (col, total), '=SUM(%s5:%s%d)' % (col, col, total - 1),
           negrita=True, color='FFFFFF', fondo=AZUL, borde=True,
           fmt=ENTERO if col == 'B' else MONEDA)
    A['ing_total'] = total
    A['ing'] = {'gravado': 'C%d' % total, 'no_gravado': 'D%d' % total,
                'brutos': 'E%d' % total, 'iva': 'F%d' % total, 'reteiva': 'G%d' % total}

    # B. Clasificación tributaria de cada concepto
    ini = total + 2
    _seccion(ws, ini, 'B. CLASIFICACIÓN TRIBUTARIA DE CADA CONCEPTO (qué es y a dónde va)', 9)
    ws.merge_cells('A%d:C%d' % (ini + 1, ini + 1))
    _encabezados(ws, ini + 1, ['Concepto'], desde=1)
    for col, txt in (('D', 'Valor'), ('E', '¿Gravado con IVA?'),
                     ('F', '¿Entra a la base del SIMPLE?'), ('G', '¿Gravado con ICA?'),
                     ('H', 'Sustento')):
        _c(ws, '%s%d' % (col, ini + 1), txt, negrita=True, tam=9, color='FFFFFF',
           fondo=AZUL_MEDIO, alin='center', vert='center', ajuste=True, borde=True)

    conceptos = [
        (ficha.get('concepto_gravado', 'Ingresos gravados con IVA'),
         "='2.INGRESOS'!$%s" % A['ing']['gravado'].replace('C', 'C$'),
         'Sí — %g%%' % (P.TARIFA_IVA * 100), 'Sí', 'Sí', 'Servicio gravado, Art. 420 ET', False),
        ('Ingresos no gravados con IVA (excluidos)',
         "='2.INGRESOS'!$%s" % A['ing']['no_gravado'].replace('D', 'D$'),
         'NO', 'Sí', 'Sí', 'Excluido de IVA pero es ingreso bruto ordinario', False),
        ('Ingresos no constitutivos de renta (INCRNGO)', liq['incrngo'],
         'N/A', 'NO', 'NO', 'Se depuran de la base — Art. 910 ET', True),
        ('Ganancias ocasionales', liq['ganancias_ocasionales'],
         'N/A', 'NO', 'NO', 'No hacen parte del anticipo bimestral', True),
        ('Devoluciones, rebajas y descuentos (notas crédito)', liq['devoluciones'],
         '—', 'Restan', 'Restan',
         'Se detectaron y restaron del detalle' if liq['devoluciones']
         else 'No se detectaron notas crédito en el bimestre', True),
        ('Ingresos por fuera del municipio', 0,
         '—', 'Sí', 'NO en %s' % (liq['filas_ica'][0]['nombre'] or 'el municipio'),
         'Se declaran en el municipio de origen', True),
    ]
    f = ini + 2
    for texto, valor, iva, simple, ica, sustento, editable in conceptos:
        ws.merge_cells('A%d:C%d' % (f, f))
        _c(ws, 'A%d' % f, texto, borde=True)
        _c(ws, 'D%d' % f, valor, fmt=MONEDA, borde=True,
           fondo=CREMA if editable else None)
        for col, v in (('E', iva), ('F', simple), ('G', ica)):
            _c(ws, '%s%d' % (col, f), v, tam=9, alin='center', borde=True)
        _c(ws, 'H%d' % f, sustento, tam=9, borde=True)
        f += 1
    _c(ws, 'A%d' % f, 'BASE GRAVABLE DEL ANTICIPO SIMPLE', negrita=True, color='FFFFFF',
       fondo=AZUL, borde=True)
    ws.merge_cells('A%d:C%d' % (f, f))
    _c(ws, 'D%d' % f, '=D%d+D%d-D%d-D%d-D%d' % (ini + 2, ini + 3, ini + 4, ini + 5, ini + 6),
       negrita=True, color='FFFFFF', fondo=AZUL, fmt=MONEDA, borde=True)

    # C. Detalle de las facturas
    ini = f + 2
    _seccion(ws, ini, 'C. DETALLE DE LAS %d FACTURAS DE VENTA DEL BIMESTRE'
             % len(liq['facturas']), 9)
    _encabezados(ws, ini + 1, ['Fecha', 'Prefijo', 'Folio', 'Cliente', 'Ingreso gravado',
                               'Ingreso no gravado', 'IVA %g%%' % (P.TARIFA_IVA * 100),
                               'ReteIVA %g%%' % (P.TARIFA_RETEIVA * 100), 'Total factura'])
    f = ini + 2
    facturas = sorted(liq['facturas'], key=lambda v: (v['fecha'], v['folio']))
    for i, v in enumerate(facturas):
        fondo = CEBRA if i % 2 else None
        _c(ws, 'A%d' % f, v['fecha'].strftime('%d-%m-%Y') if v['fecha'] else '',
           tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'B%d' % f, v['prefijo'], tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'C%d' % f, v['folio'], tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'D%d' % f, v['tercero'], tam=9, fondo=fondo, borde=True)
        _c(ws, 'E%d' % f, v['gravado'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'F%d' % f, v['no_gravado'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'G%d' % f, v['iva'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'H%d' % f, v['reteiva'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'I%d' % f, '=E%d+F%d+G%d' % (f, f, f), fmt=MONEDA, fondo=fondo, borde=True)
        f += 1
    ws.merge_cells('A%d:D%d' % (f, f))
    _c(ws, 'A%d' % f, 'TOTALES', negrita=True, color='FFFFFF', fondo=AZUL_MEDIO, borde=True)
    for col in 'EFGHI':
        _c(ws, '%s%d' % (col, f), '=SUM(%s%d:%s%d)' % (col, ini + 2, col, f - 1),
           negrita=True, color='FFFFFF', fondo=AZUL_MEDIO, fmt=MONEDA, borde=True)
    return ws


# ------------------------------------------------------- 3. IVA

def hoja_iva(wb, liq, ficha, A):
    ws = _hoja(wb, '3.IVA', {'A': 9, 'B': 18, 'C': 18, 'D': 18, 'E': 12, 'F': 20,
                             'G': 26, 'H': 26})
    _titulo(ws, 1, 'LIQUIDACIÓN DEL IVA DEL BIMESTRE — se transfiere en el recibo 2593', 8)

    def linea(fila, casilla, concepto, formula, nota, fondo=None):
        ws.merge_cells('B%d:E%d' % (fila, fila))
        ws.merge_cells('G%d:H%d' % (fila, fila))
        _c(ws, 'A%d' % fila, casilla, negrita=True, tam=9, color=AZUL_MEDIO,
           alin='center', fondo=fondo, borde=True)
        _c(ws, 'B%d' % fila, concepto, negrita=fondo in (AZUL_CLARO, VERDE_CLARO),
           fondo=fondo, borde=True)
        _c(ws, 'F%d' % fila, formula, fmt=MONEDA, alin='right',
           negrita=fondo in (AZUL_CLARO, VERDE_CLARO), fondo=fondo, borde=True)
        _c(ws, 'G%d' % fila, nota, tam=9, fondo=fondo, borde=True)

    linea(3, 'Cas.', 'CONCEPTO', 'VALOR', 'Observación', AZUL_CLARO)
    linea(4, 77, 'IVA generado por operaciones gravadas (%g%%)' % (P.TARIFA_IVA * 100),
          "='2.INGRESOS'!$%s$%d" % ('F', A['ing_total']), 'Hoja 2 — facturación del bimestre')
    linea(5, 78, '(−) Impuestos descontables del bimestre',
          "=-'4.COMPRAS'!$F$%d" % A['compras_total'],
          '%d compras gravadas — Art. 485 ET' % len(liq['compras_gravadas']))
    linea(6, 79, '(=) Saldo a pagar por IVA del período', '=F4+F5',
          'Generado menos descontable', AZUL_CLARO)
    linea(7, 80, '(−) Retenciones de IVA que le practicaron',
          "=-'2.INGRESOS'!$G$%d" % A['ing_total'],
          'Art. 437-2 num. 9 y 437-1 ET')
    linea(8, 81, '(−) Saldo a favor del bimestre anterior', liq['saldo_favor_iva_anterior'],
          'Editable si viene saldo a favor del bimestre anterior')
    ws['F8'].fill = PatternFill('solid', fgColor=CREMA)
    linea(9, 74, 'TOTAL IVA A PAGAR EN EL RECIBO 2593', '=MAX(0,F6+F7+F8)',
          'Pasa a la hoja 1.LIQUIDACIÓN', VERDE_CLARO)
    A['iva_total'] = 'F9'

    _seccion(ws, 12, 'VERIFICACIÓN DE CONSISTENCIA', 8)
    checks = [
        ('IVA generado ÷ ingresos gravados (debe dar %s%%)'
         % ('%.2f' % (P.TARIFA_IVA * 100)).replace('.', ','),
         "='2.INGRESOS'!$F$%d/'2.INGRESOS'!$C$%d" % (A['ing_total'], A['ing_total']),
         PORCENTAJE),
        ('ReteIVA ÷ IVA generado (debe dar %s%%)'
         % ('%.2f' % (P.TARIFA_RETEIVA * 100)).replace('.', ','),
         "='2.INGRESOS'!$G$%d/'2.INGRESOS'!$F$%d" % (A['ing_total'], A['ing_total']),
         PORCENTAJE),
        ('ReteIVA según certificados de los agentes retenedores (hoja 5)',
         liq['reteiva_certificados'], MONEDA),
        ('ReteIVA según facturación electrónica',
         "='2.INGRESOS'!$G$%d" % A['ing_total'], MONEDA),
        ('Diferencia certificados vs. facturación', '=F15-F16', MONEDA2),
    ]
    for i, (texto, valor, fmt) in enumerate(checks):
        fila = 13 + i
        ws.merge_cells('B%d:E%d' % (fila, fila))
        _c(ws, 'B%d' % fila, texto, tam=9, borde=True)
        _c(ws, 'F%d' % fila, valor, fmt=fmt, alin='right', borde=True)
    return ws


# ------------------------------------------------------- 4. COMPRAS

def hoja_compras(wb, liq, ficha, A):
    ws = _hoja(wb, '4.COMPRAS', {'A': 11, 'B': 13, 'C': 40, 'D': 9, 'E': 12, 'F': 15,
                                 'G': 14, 'H': 15, 'I': 30})
    _titulo(ws, 1, 'CLASIFICACIÓN DE LAS COMPRAS Y DEL IVA DESCONTABLE', 9)
    _seccion(ws, 3, 'A. RESUMEN POR NATURALEZA', 9)
    _encabezados(ws, 4, ['Clasificación'], desde=1)
    ws.merge_cells('A4:C4')
    for col, txt in (('D', 'N° doc.'), ('E', 'Base gravable'), ('F', 'IVA descontable'),
                     ('G', 'Compras sin IVA'), ('H', 'Total pagado'),
                     ('I', 'Efecto tributario')):
        _c(ws, '%s4' % col, txt, negrita=True, tam=9, color='FFFFFF', fondo=AZUL_MEDIO,
           alin='center', vert='center', ajuste=True, borde=True)

    grav, excl = liq['compras_gravadas'], liq['compras_excluidas']
    filas = [
        ('Compras GRAVADAS (generan IVA descontable)', len(grav),
         sum(c['base'] for c in grav), sum(c['iva'] for c in grav), 0,
         sum(c['total'] for c in grav), 'Descontable del IVA generado — Art. 485 ET'),
        ('Compras EXCLUIDAS / NO GRAVADAS', len(excl), 0, 0,
         sum(c['total'] for c in excl), sum(c['total'] for c in excl),
         'No generan descontable; son costo/gasto'),
    ]
    f = 5
    for texto, n, base, iva, sin_iva, total, efecto in filas:
        ws.merge_cells('A%d:C%d' % (f, f))
        _c(ws, 'A%d' % f, texto, borde=True)
        _c(ws, 'D%d' % f, n, alin='center', borde=True)
        for col, v in (('E', base), ('F', iva), ('G', sin_iva), ('H', total)):
            _c(ws, '%s%d' % (col, f), v, fmt=MONEDA, borde=True)
        _c(ws, 'I%d' % f, efecto, tam=9, borde=True)
        f += 1
    ws.merge_cells('A%d:C%d' % (f, f))
    _c(ws, 'A%d' % f, 'TOTAL COMPRAS DEL BIMESTRE', negrita=True, color='FFFFFF',
       fondo=AZUL, borde=True)
    for col in 'DEFGH':
        _c(ws, '%s%d' % (col, f), '=SUM(%s5:%s%d)' % (col, col, f - 1), negrita=True,
           color='FFFFFF', fondo=AZUL, fmt=ENTERO if col == 'D' else MONEDA, borde=True)
    A['compras_total'] = f

    ini = f + 2
    todas = sorted(grav + excl, key=lambda c: (c['fecha'], c['folio']))
    _seccion(ws, ini, 'B. DETALLE DE LAS %d COMPRAS DEL BIMESTRE' % len(todas), 9)
    _encabezados(ws, ini + 1, ['Fecha', 'NIT', 'Proveedor', 'Prefijo', 'Folio',
                               'Base gravable', 'IVA descontable', 'Total', 'Clasificación'])
    f = ini + 2
    for i, c in enumerate(todas):
        fondo = VERDE_CLARO if c['iva'] else (CEBRA if i % 2 else None)
        _c(ws, 'A%d' % f, c['fecha'].strftime('%d-%m-%Y') if c['fecha'] else '',
           tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'B%d' % f, c['nit'], tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'C%d' % f, c['proveedor'], tam=9, fondo=fondo, borde=True)
        _c(ws, 'D%d' % f, c['prefijo'], tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'E%d' % f, c['folio'], tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'F%d' % f, c['base'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'G%d' % f, c['iva'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'H%d' % f, c['total'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'I%d' % f, c['clasificacion'], tam=9, fondo=fondo, borde=True)
        f += 1
    ws.merge_cells('A%d:E%d' % (f, f))
    _c(ws, 'A%d' % f, 'TOTALES', negrita=True, color='FFFFFF', fondo=AZUL_MEDIO, borde=True)
    for col in 'FGH':
        _c(ws, '%s%d' % (col, f), '=SUM(%s%d:%s%d)' % (col, ini + 2, col, f - 1),
           negrita=True, color='FFFFFF', fondo=AZUL_MEDIO, fmt=MONEDA, borde=True)
    return ws


# ------------------------------------------------------- 5. RETEIVA

def hoja_reteiva(wb, liq, ficha, A):
    ws = _hoja(wb, '5.RETEIVA', {'A': 7, 'B': 18, 'C': 46, 'D': 18, 'E': 18, 'F': 16,
                                 'G': 26})
    _titulo(ws, 1, 'RETENCIÓN DE IVA PRACTICADA POR TERCEROS — conciliación de certificados', 7)
    ws.merge_cells('A3:G3')
    _c(ws, 'A3', 'Quien adquiere bienes o servicios gravados de un contribuyente del SIMPLE '
                 'debe practicarle retención de IVA del %g%% del impuesto (Art. 437-2 num. 9 '
                 'y Art. 437-1 ET). La DIAN cruza contra el certificado del agente retenedor.'
       % (P.TARIFA_RETEIVA * 100), tam=9, fondo=VERDE_CLARO, ajuste=True)
    ws.row_dimensions[3].height = 30

    _encabezados(ws, 5, ['N°', 'NIT', 'Tercero (agente retenedor)', 'Base: IVA facturado',
                         'ReteIVA s/ contabilidad', 'ReteIVA recalculada',
                         'Diferencia vs. certificado'])
    f = 6
    for i, ag in enumerate(liq['agentes_reteiva'], 1):
        fondo = ROSA if abs(ag['diferencia']) > 10 else None
        _c(ws, 'A%d' % f, i, tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'B%d' % f, ag['nit'], tam=9, alin='center', fondo=fondo, borde=True)
        _c(ws, 'C%d' % f, ag['tercero'], tam=9, fondo=fondo, borde=True)
        _c(ws, 'D%d' % f, ag['base'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'E%d' % f, ag['contabilidad'], fmt=MONEDA, fondo=fondo, borde=True)
        _c(ws, 'F%d' % f, "=ROUND(D%d*'8.PARAMETROS'!$C$10,2)" % f, fmt=MONEDA,
           fondo=fondo, borde=True)
        _c(ws, 'G%d' % f, ag['diferencia'], fmt=MONEDA, fondo=fondo, borde=True)
        f += 1
    ws.merge_cells('A%d:C%d' % (f, f))
    _c(ws, 'A%d' % f, 'TOTALES', negrita=True, color='FFFFFF', fondo=AZUL, borde=True)
    for col in 'DEFG':
        _c(ws, '%s%d' % (col, f), '=SUM(%s6:%s%d)' % (col, col, f - 1), negrita=True,
           color='FFFFFF', fondo=AZUL, fmt=MONEDA, borde=True)
    ws.merge_cells('A%d:G%d' % (f + 2, f + 2))
    _c(ws, 'A%d' % (f + 2),
       'Filas en rojo: el certificado del agente retenedor no coincide con lo contabilizado. '
       'Solicita el certificado y ajusta antes de presentar.', tam=9, ajuste=True)
    return ws


# ------------------------------------------------------- 6. ICA MUNICIPIO

def hoja_ica(wb, liq, ficha, A):
    ws = _hoja(wb, '6.ICA MUNICIPIO', {'A': 9, 'B': 24, 'C': 14, 'D': 20, 'E': 20,
                                       'F': 16, 'G': 18, 'H': 28})
    _titulo(ws, 1, 'COMPONENTE TERRITORIAL — ICA CONSOLIDADO POR MUNICIPIO', 8)
    ws.merge_cells('A3:H3')
    _c(ws, 'A3', 'El SIMPLE integra el ICA, avisos y tableros y sobretasa bomberil. Se '
                 'liquida dentro del anticipo y la DIAN lo gira al municipio; por eso se '
                 'separa en la casilla 43 y no puede cubrirse con el descuento por aportes '
                 'a pensión.', tam=9, fondo=AZUL_CLARO, ajuste=True)
    ws.row_dimensions[3].height = 30

    _encabezados(ws, 5, ['Cód. DANE', 'Municipio', 'Depto.', 'Ingresos brutos del municipio',
                         '(−) Ingresos NO gravados con ICA', '(=) Base gravable ICA',
                         'Tarifa consolidada', 'ICA consolidado del bimestre'])
    f = 6
    for i, m in enumerate(liq['filas_ica']):
        _c(ws, 'A%d' % f, m['codigo'], alin='center', borde=True)
        _c(ws, 'B%d' % f, m['nombre'], borde=True)
        _c(ws, 'C%d' % f, m['depto'], borde=True)
        _c(ws, 'D%d' % f, "='2.INGRESOS'!$E$%d" % A['ing_total'] if i == 0 else m['ingresos'],
           fmt=MONEDA, borde=True, fondo=None if i == 0 else CREMA)
        _c(ws, 'E%d' % f, m['no_gravados_ica'], fmt=MONEDA, fondo=CREMA, borde=True)
        _c(ws, 'F%d' % f, '=D%d-E%d' % (f, f), fmt=MONEDA, borde=True)
        _c(ws, 'G%d' % f, "='8.PARAMETROS'!$C$6" if i == 0 else m['tarifa'],
           fmt=TARIFA_ICA_FMT, borde=True, fondo=None if i == 0 else CREMA)
        _c(ws, 'H%d' % f, '=ROUND(F%d*G%d,0)' % (f, f), fmt=MONEDA, borde=True)
        f += 1
    # dos filas libres para municipios adicionales: el SIMPLE se declara por municipio
    for _ in range(2):
        for col in 'ABCDEG':
            _c(ws, '%s%d' % (col, f), None, fondo=CREMA, borde=True,
               fmt=MONEDA if col == 'D' else (TARIFA_ICA_FMT if col == 'G' else None))
        _c(ws, 'F%d' % f, '=D%d-E%d' % (f, f), fmt=MONEDA, borde=True)
        _c(ws, 'H%d' % f, '=IF(D%d="",0,ROUND(F%d*G%d,0))' % (f, f, f), fmt=MONEDA, borde=True)
        f += 1
    ws.merge_cells('A%d:C%d' % (f, f))
    _c(ws, 'A%d' % f, 'TOTAL NACIONAL', negrita=True, color='FFFFFF', fondo=AZUL, borde=True)
    for col in 'DEFH':
        _c(ws, '%s%d' % (col, f), '=SUM(%s6:%s%d)' % (col, col, f - 1), negrita=True,
           color='FFFFFF', fondo=AZUL, fmt=MONEDA, borde=True)
    A['ica_total'] = 'H%d' % f

    ini = f + 2
    _seccion(ws, ini, '¿QUÉ VA EN «INGRESOS NO GRAVADOS CON ICA»? (celda amarilla)', 8)
    notas = [
        '• Exportaciones de servicios y ventas al exterior (Art. 39 Ley 14 de 1983).',
        '• Venta de activos fijos.',
        '• Subsidios y aportes de la Nación.',
        '• Ingresos provenientes de actividades ejercidas FUERA del municipio (van en su '
        'propia fila).',
        '• Devoluciones, rebajas y descuentos (notas crédito) del período.',
        '• Ingresos por actividades expresamente no sujetas o exentas en el acuerdo municipal.',
    ]
    for i, n in enumerate(notas):
        ws.merge_cells('B%d:H%d' % (ini + 1 + i, ini + 1 + i))
        _c(ws, 'B%d' % (ini + 1 + i), n, tam=9)
    return ws


# ------------------------------------------------------- 7. SEG.SOCIAL

def hoja_seg_social(wb, liq, ficha, A):
    ws = _hoja(wb, '7.SEG.SOCIAL', {'A': 9, 'B': 22, 'C': 16, 'D': 16, 'E': 14, 'F': 18,
                                    'G': 18, 'H': 30})
    _titulo(ws, 1, 'APORTES A PENSIÓN — DESCUENTO DEL ANTICIPO (Art. 903 Par. 4 ET)', 8)
    ws.merge_cells('A3:H3')
    _c(ws, 'A3', 'Solo es descontable el aporte al Sistema General de Pensiones A CARGO DEL '
                 'EMPLEADOR (%g%% del IBC), PAGADO dentro del bimestre. El %g%% del trabajador y '
                 'los intereses de mora NO son descontables. El descuento no puede exceder el '
                 'anticipo ni cubrir el componente de ICA consolidado.'
       % (P.PENSION_EMPLEADOR * 100, (P.PENSION_TOTAL - P.PENSION_EMPLEADOR) * 100),
       tam=9, fondo=CREMA, ajuste=True)
    ws.row_dimensions[3].height = 40

    _encabezados(ws, 5, ['Período', 'Administradora', 'Fecha de pago', 'Planilla',
                         'N° trab.', 'Aporte pensión total (%g%%)' % (P.PENSION_TOTAL * 100),
                         'Parte del EMPLEADOR (%g%%)' % (P.PENSION_EMPLEADOR * 100),
                         'Intereses de mora (no descontable)'])
    f = 6
    formula_patronal = ("=IF(F{f}=0,0,ROUND(F{f}/'8.PARAMETROS'!$C$8*'8.PARAMETROS'!$C$7,0))")
    for pl in liq['planillas']:
        _c(ws, 'A%d' % f, pl['periodo'], tam=9, borde=True)
        _c(ws, 'B%d' % f, pl['administradora'], tam=9, borde=True)
        _c(ws, 'C%d' % f, pl['fecha_pago'].strftime('%d/%m/%Y') if pl['fecha_pago'] else '',
           tam=9, alin='center', borde=True)
        _c(ws, 'D%d' % f, pl['planilla'], tam=9, alin='center', borde=True)
        _c(ws, 'E%d' % f, pl['trabajadores'], alin='center', borde=True)
        _c(ws, 'F%d' % f, pl['aporte'], fmt=MONEDA, borde=True)
        _c(ws, 'G%d' % f, formula_patronal.format(f=f), fmt=MONEDA, borde=True)
        _c(ws, 'H%d' % f, pl['intereses'], fmt=MONEDA, borde=True)
        f += 1
    # filas libres: la planilla del mes anterior pagada dentro del bimestre también
    # es descontable y casi nunca viene en el archivo.
    for _ in range(2):
        for col in 'ABCDEFH':
            _c(ws, '%s%d' % (col, f), None, fondo=CREMA, borde=True,
               fmt=MONEDA if col in 'FH' else None)
        _c(ws, 'G%d' % f, formula_patronal.format(f=f), fmt=MONEDA, fondo=CREMA, borde=True)
        f += 1
    ws.merge_cells('A%d:E%d' % (f, f))
    _c(ws, 'A%d' % f, 'TOTAL PAGADO EN EL BIMESTRE', negrita=True, color='FFFFFF',
       fondo=AZUL, borde=True)
    for col in 'FGH':
        _c(ws, '%s%d' % (col, f), '=SUM(%s6:%s%d)' % (col, col, f - 1), negrita=True,
           color='FFFFFF', fondo=AZUL, fmt=MONEDA, borde=True)
    A['pension_total'] = 'G%d' % f

    ws.merge_cells('B%d:E%d' % (f + 2, f + 2))
    _c(ws, 'B%d' % (f + 2), 'Descuento aplicable al anticipo →', negrita=True)
    _c(ws, 'F%d' % (f + 2), "='7.SEG.SOCIAL'!$G$%d" % f, fmt=MONEDA, negrita=True,
       fondo=VERDE_CLARO, borde=True)

    ini = f + 4
    _seccion(ws, ini, 'NOTA SOBRE LA COMPLETITUD DEL DATO', 8)
    periodos = ', '.join(sorted({p['periodo'] for p in liq['planillas']})) or '(ninguna)'
    notas = [
        '• El archivo fuente contiene las planillas de: %s.' % periodos,
        '• La regla del Art. 903 Par. 4 ET es «pagado en el bimestre»: también sería '
        'descontable la planilla del período anterior si se pagó dentro de este bimestre.',
        '• Si esa planilla existe, agrégala en las dos filas amarillas y toda la '
        'liquidación se recalcula sola.',
    ]
    mora = [p for p in liq['planillas'] if p.get('dias_mora')]
    if mora:
        notas.append('• Se registran pagos con mora (%s días): los intereses no son '
                     'descontables y ya están separados en la columna H.'
                     % ', '.join(str(p['dias_mora']) for p in mora))
    for i, n in enumerate(notas):
        ws.merge_cells('B%d:H%d' % (ini + 1 + i, ini + 1 + i))
        _c(ws, 'B%d' % (ini + 1 + i), n, tam=9, ajuste=True)
    return ws


# ------------------------------------------------------- 8. PARAMETROS

def hoja_parametros(wb, liq, ficha, A):
    ws = _hoja(wb, '8.PARAMETROS', {'A': 10, 'B': 34, 'C': 16, 'D': 16, 'E': 14, 'F': 16,
                                    'G': 16, 'H': 30})
    _titulo(ws, 1, 'PARÁMETROS DE LIQUIDACIÓN  ·  celdas amarillas = editables', 8)
    _seccion(ws, 3, 'A. DATOS BÁSICOS', 8)
    basicos = [
        ('UVT año gravable %d' % liq['ano'], liq['uvt'], ENTERO,
         'Resolución de la DIAN del año gravable'),
        ('Grupo de actividad SIMPLE (Art. 908 ET)', liq['grupo'], '0',
         '%d = %s' % (liq['grupo'], P.GRUPOS[liq['grupo']])),
        ('Tarifa consolidada ICA del municipio', liq['filas_ica'][0]['tarifa'],
         TARIFA_ICA_FMT, '%g por mil — dato de la ficha del contribuyente'
         % (liq['filas_ica'][0]['tarifa'] * 1000)),
        ('%% aporte pensión a cargo del EMPLEADOR', P.PENSION_EMPLEADOR, PORCENTAJE,
         'Art. 33 Ley 100/93'),
        ('%% aporte pensión total (empleador+trabajador)', P.PENSION_TOTAL, PORCENTAJE,
         '%g%% patronal + %g%% trabajador' % (P.PENSION_EMPLEADOR * 100,
                                              (P.PENSION_TOTAL - P.PENSION_EMPLEADOR) * 100)),
        ('Tarifa general IVA', P.TARIFA_IVA, PORCENTAJE, 'Art. 468 ET'),
        ('Tarifa retención de IVA', P.TARIFA_RETEIVA, PORCENTAJE, 'Art. 437-1 ET'),
    ]
    for i, (texto, valor, fmt, nota) in enumerate(basicos):
        f = 4 + i
        _c(ws, 'B%d' % f, texto, borde=True)
        _c(ws, 'C%d' % f, valor, fmt=fmt, fondo=CREMA, alin='right', borde=True)
        ws.merge_cells('D%d:H%d' % (f, f))
        _c(ws, 'D%d' % f, nota, tam=9, borde=True)
    # C4 UVT, C5 grupo, C6 tarifa ICA, C7 pensión patronal, C8 pensión total,
    # C9 IVA, C10 ReteIVA — las fórmulas de las demás hojas apuntan a estas celdas.

    _seccion(ws, 12, 'B. TABLA DE TARIFAS SIMPLE — BASE BIMESTRAL (Art. 908 ET) · vigente '
                     'tras la Sentencia C-540 de 2023', 8)
    _encabezados(ws, 13, ['Grupo de actividad', 'Desde UVT', 'Hasta UVT', 'Tarifa',
                          'Equivalente anual desde', 'Equivalente anual hasta',
                          'Base bimestral en $ (hasta)'], desde=2)
    _c(ws, 'A13', '', fondo=AZUL_MEDIO, borde=True)
    f = 14
    for grupo in sorted(P.TARIFAS):
        for j, (desde, hasta, tar) in enumerate(P.TARIFAS[grupo]):
            resalte = VERDE_CLARO if grupo == liq['grupo'] else None
            _c(ws, 'A%d' % f, grupo, alin='center', fondo=resalte, borde=True)
            if j == 0:
                _c(ws, 'B%d' % f, '%d · %s' % (grupo, P.GRUPOS[grupo]), tam=9,
                   fondo=resalte, borde=True)
            else:
                _c(ws, 'B%d' % f, None, fondo=resalte, borde=True)
            _c(ws, 'C%d' % f, desde, fmt=ENTERO, fondo=resalte, borde=True)
            _c(ws, 'D%d' % f, hasta, fmt=ENTERO, fondo=resalte, borde=True)
            _c(ws, 'E%d' % f, tar, fmt=PORCENTAJE, fondo=resalte, borde=True)
            _c(ws, 'F%d' % f, '=C%d*6' % f, fmt=ENTERO, fondo=resalte, borde=True)
            _c(ws, 'G%d' % f, '=D%d*6' % f, fmt=ENTERO, fondo=resalte, borde=True)
            _c(ws, 'H%d' % f, "=D%d*'8.PARAMETROS'!$C$4" % f, fmt=MONEDA,
               fondo=resalte, borde=True)
            f += 1
    A['tabla_tarifas'] = (14, f - 1)
    ws.merge_cells('B%d:H%d' % (f + 1, f + 1))
    _c(ws, 'B%d' % (f + 1), '▲ Resaltado en verde: grupo aplicable a este contribuyente.',
       tam=9)

    ini = f + 3
    _seccion(ws, ini, 'C. FUNDAMENTO NORMATIVO', 8)
    for i, (norma, texto) in enumerate(P.NORMAS):
        fila = ini + 1 + i
        _c(ws, 'B%d' % fila, norma, negrita=True, tam=9, vert='top', borde=True)
        ws.merge_cells('C%d:H%d' % (fila, fila))
        _c(ws, 'C%d' % fila, texto, tam=9, ajuste=True, vert='top', borde=True)
        ws.row_dimensions[fila].height = 30
    return ws


# ------------------------------------------------------- 9. ALERTAS

def hoja_alertas(wb, liq, ficha, A):
    ws = _hoja(wb, '9.ALERTAS', {'A': 6, 'B': 34, 'C': 96, 'D': 12})
    _titulo(ws, 1, 'PUNTOS QUE DEBES VERIFICAR ANTES DE PRESENTAR', 4)
    _encabezados(ws, 3, ['#', 'Tema', 'Qué revisar y por qué importa', 'Prioridad'])

    orden = {'ALTA': 0, 'MEDIA': 1, 'BAJA': 2}
    alertas = sorted(liq['alertas'], key=lambda a: orden.get(a['prioridad'], 3))
    f = 4
    for i, al in enumerate(alertas, 1):
        fondo = CEBRA if i % 2 == 0 else None
        _c(ws, 'A%d' % f, i, alin='center', vert='top', fondo=fondo, borde=True)
        _c(ws, 'B%d' % f, al['tema'], negrita=True, vert='top', ajuste=True,
           fondo=fondo, borde=True)
        _c(ws, 'C%d' % f, al['texto'], tam=9, vert='top', ajuste=True, fondo=fondo,
           borde=True)
        _c(ws, 'D%d' % f, al['prioridad'], tam=9, negrita=True, alin='center', vert='top',
           fondo=ROSA if al['prioridad'] == 'ALTA' else fondo, borde=True)
        ws.row_dimensions[f].height = max(30, 15 * (len(al['texto']) // 95 + 1))
        f += 1

    ini = f + 1
    _seccion(ws, ini, 'VALIDACIONES AUTOMÁTICAS DEL MOTOR — semáforo: %s' % liq['semaforo'], 4)
    _encabezados(ws, ini + 1, ['', 'Comprobación', 'Resultado', 'Estado'])
    f = ini + 2
    for v in liq['validaciones']:
        _c(ws, 'A%d' % f, '', borde=True)
        _c(ws, 'B%d' % f, v['nombre'], tam=9, vert='top', ajuste=True, borde=True)
        _c(ws, 'C%d' % f, v['nota'], tam=9, vert='top', ajuste=True, borde=True)
        _c(ws, 'D%d' % f, ('OK' if v['ok'] else 'REVISAR'), tam=9, negrita=True,
           alin='center', fondo=None if v['ok'] else ROSA, borde=True)
        ws.row_dimensions[f].height = 30
        f += 1
    return ws


# ------------------------------------------------------- 1. LIQUIDACIÓN 2593

def hoja_liquidacion(wb, liq, ficha, A):
    ws = _hoja(wb, '1.LIQUIDACIÓN 2593', {'A': 8, 'B': 20, 'C': 16, 'D': 16, 'E': 14,
                                          'F': 20, 'G': 22, 'H': 26})
    _titulo(ws, 1, 'RECIBO ELECTRÓNICO DEL SIMPLE — FORMULARIO 2593', 8, tam=15)
    ws.merge_cells('A2:H2')
    _c(ws, 'A2', 'Anticipo bimestral del impuesto unificado bajo el Régimen Simple de '
                 'Tributación (Art. 910 ET)', color='FFFFFF', fondo=AZUL_MEDIO, vert='center')

    _seccion(ws, 4, 'SECCIÓN A — IDENTIFICACIÓN DEL DECLARANTE', 8)
    ident = [
        ('1', 'Año gravable', liq['ano']),
        ('3', 'Período', '%d — %s' % (liq['bimestre'], liq['nombre_bimestre'])),
        ('5', 'Número de Identificación Tributaria (NIT)', ficha['nit']),
        ('6', 'DV', ficha.get('dv') or '(verificar en RUT)'),
        ('7', 'Razón social', ficha['nombre']),
        ('12', 'Código Dirección Seccional',
         ficha.get('direccion_seccional') or '(verificar en RUT)'),
        ('24', 'Grupo de actividad SIMPLE',
         'Grupo %d — %s' % (liq['grupo'], P.GRUPOS[liq['grupo']])),
        ('—', 'Actividad económica (CIIU)', ficha.get('ciiu', '')),
        ('—', 'Municipio de la actividad', '%s (%s) — %s'
         % (liq['filas_ica'][0]['nombre'], liq['filas_ica'][0]['codigo'],
            liq['filas_ica'][0]['depto'])),
        ('—', 'Responsabilidad de IVA',
         'Sí — responsable, declaración anual Form. 300 (Art. 915 ET)'
         if ficha.get('responsable_iva') else 'No responsable de IVA'),
    ]
    f = 5
    for casilla, etiqueta, valor in ident:
        _c(ws, 'A%d' % f, casilla, negrita=True, tam=9, color=AZUL_MEDIO, alin='center')
        ws.merge_cells('B%d:D%d' % (f, f))
        _c(ws, 'B%d' % f, etiqueta)
        ws.merge_cells('E%d:H%d' % (f, f))
        _c(ws, 'E%d' % f, valor, negrita=True)
        f += 1

    def linea(fila, casilla, concepto, formula, nota='', fondo=None, fmt=MONEDA):
        ws.merge_cells('B%d:E%d' % (fila, fila))
        ws.merge_cells('G%d:H%d' % (fila, fila))
        resaltado = fondo in (AZUL_CLARO, VERDE_CLARO, CREMA)
        _c(ws, 'A%d' % fila, casilla, negrita=True, tam=9, color=AZUL_MEDIO, alin='center',
           fondo=fondo, borde=True)
        _c(ws, 'B%d' % fila, concepto, negrita=resaltado, fondo=fondo, borde=True)
        _c(ws, 'F%d' % fila, formula, fmt=fmt, alin='right', negrita=resaltado,
           fondo=fondo, borde=True)
        _c(ws, 'G%d' % fila, nota, tam=9, fondo=fondo, borde=True)

    ing = A['ing_total']
    _seccion(ws, 16, 'SECCIÓN B — INGRESOS DEL BIMESTRE Y DETERMINACIÓN DE LA TARIFA', 8)
    linea(17, 'Cas.', 'CONCEPTO', 'VALOR', 'Observación', AZUL_CLARO)
    linea(18, 25, 'Ingresos brutos gravados con IVA',
          "='2.INGRESOS'!$C$%d" % ing, ficha.get('concepto_gravado', 'Hoja 2'))
    linea(19, 26, 'Ingresos brutos no gravados / excluidos de IVA',
          "='2.INGRESOS'!$D$%d" % ing, 'Igualmente parte de la base del SIMPLE')
    linea(20, 27, '(−) Ingresos no constitutivos de renta ni ganancia ocasional',
          liq['incrngo'], 'Editable', CREMA)
    linea(21, 28, '(−) Ganancias ocasionales', liq['ganancias_ocasionales'],
          'No hacen parte del anticipo', CREMA)
    linea(22, 29, '(=) INGRESOS BRUTOS BIMESTRALES (base del anticipo)',
          "='2.INGRESOS'!$E$%d-F20-F21" % ing, 'Art. 910 ET', AZUL_CLARO)
    linea(23, '—', 'Base expresada en UVT', "=F22/'8.PARAMETROS'!$C$4",
          'UVT %d = $%s' % (liq['ano'], _pesos(liq['uvt'])), fmt=UVT_FMT)
    ini_t, fin_t = A['tabla_tarifas']
    linea(24, '—', 'Tarifa SIMPLE consolidada aplicable',
          "=SUMIFS('8.PARAMETROS'!$E${i}:$E${f},'8.PARAMETROS'!$A${i}:$A${f},"
          "'8.PARAMETROS'!$C$5,'8.PARAMETROS'!$C${i}:$C${f},\"<\"&F23,"
          "'8.PARAMETROS'!$D${i}:$D${f},\">=\"&F23)".format(i=ini_t, f=fin_t),
          'Grupo %d, tramo por la base en UVT' % liq['grupo'], CREMA, fmt=PORCENTAJE)
    linea(25, 30, '(=) ANTICIPO SIMPLE CONSOLIDADO', '=ROUND(F22*F24,0)',
          'Incluye el componente de ICA', AZUL_CLARO)

    _seccion(ws, 27, 'SECCIÓN C — SEPARACIÓN DEL COMPONENTE TERRITORIAL Y DESCUENTOS', 8)
    linea(28, 43, '(−) Componente de ICA consolidado (%s, %g x mil)'
          % (liq['filas_ica'][0]['nombre'], liq['filas_ica'][0]['tarifa'] * 1000),
          "=-'6.ICA MUNICIPIO'!$%s" % A['ica_total'].replace('H', 'H$'),
          'Se gira al municipio — hoja 6')
    linea(29, '—', '(=) Componente nacional del anticipo', '=F25+F28',
          'Única parte sobre la que opera el descuento', AZUL_CLARO)
    linea(30, 44, '(−) Descuento: aportes a pensión a cargo del empleador',
          "=-MIN('7.SEG.SOCIAL'!$%s,F29)" % A['pension_total'].replace('G', 'G$'),
          '%g%% patronal pagado en el bimestre — hoja 7' % (P.PENSION_EMPLEADOR * 100))
    linea(31, 45, '(−) Retenciones y autorretenciones practicadas antes de ingresar al SIMPLE',
          liq['retenciones_previas'], 'Solo procede en el 1.er bimestre', CREMA)
    linea(32, 49, '(=) ANTICIPO NETO A PAGAR (componente nacional)',
          '=MAX(0,F29+F30-F31)', 'No puede ser negativo', VERDE_CLARO)

    _seccion(ws, 34, 'SECCIÓN D — IMPUESTO SOBRE LAS VENTAS (se transfiere en este recibo)', 8)
    linea(35, 77, 'IVA generado por operaciones gravadas (%g%%)' % (P.TARIFA_IVA * 100),
          "='2.INGRESOS'!$F$%d" % ing, 'Hoja 3')
    linea(36, 78, '(−) Impuestos descontables',
          "=-'4.COMPRAS'!$F$%d" % A['compras_total'], 'Hoja 4 — Art. 485 ET')
    linea(37, 80, '(−) Retenciones de IVA que le practicaron (%g%%)' % (P.TARIFA_RETEIVA * 100),
          "=-'2.INGRESOS'!$G$%d" % ing, 'Hoja 5 — Art. 437-2 num. 9 ET')
    linea(38, 74, '(=) TOTAL IVA A PAGAR', "='3.IVA'!$%s" % A['iva_total'].replace('F', 'F$'),
          'Declaración anual: Form. 300', VERDE_CLARO)

    _seccion(ws, 40, 'SECCIÓN E — IMPUESTO NACIONAL AL CONSUMO', 8)
    linea(41, '—', 'INC de bares y restaurantes (8%)', float(ficha.get('inc', 0) or 0),
          'Solo aplica al grupo 4 (expendio de comidas y bebidas)'
          if liq['grupo'] != 4 else 'Editable — 8% sobre el consumo facturado', CREMA)

    _seccion(ws, 43, 'SECCIÓN F — TOTAL A PAGAR', 8)
    linea(44, 49, 'Anticipo neto del SIMPLE (componente nacional)', '=F32')
    linea(45, 43, 'ICA consolidado (se gira a %s)' % liq['filas_ica'][0]['nombre'],
          "='6.ICA MUNICIPIO'!$%s" % A['ica_total'].replace('H', 'H$'))
    linea(46, 74, 'IVA del bimestre', '=F38')
    linea(47, 80, 'Sanciones', liq['sanciones'], 'Editable — extemporaneidad, si aplica', CREMA)
    linea(48, 81, 'TOTAL A PAGAR', '=SUM(F44:F47)+F41', 'Antes de aproximación', CREMA)
    linea(49, '—', 'TOTAL A PAGAR APROXIMADO AL MÚLTIPLO DE MIL', '=ROUND(F48/1000,0)*1000',
          'Art. 577 ET — se paga con Form. 490', VERDE_CLARO)
    ws.row_dimensions[49].height = 26

    _seccion(ws, 52, 'SEMÁFORO DE REVISIÓN Y PLAZO', 8)
    ws.merge_cells('B53:H53')
    _c(ws, 'B53', 'Semáforo del motor: %s. %d punto(s) por verificar en la hoja 9.'
       % (liq['semaforo'], len(liq['alertas'])), tam=9, ajuste=True, vert='center')
    ws.row_dimensions[53].height = 18
    # La fecha exacta, no «el mes siguiente»: sale del calendario oficial por el
    # último dígito del NIT. Al lado, los días que faltan, con TODAY().
    vence = plazos.vencimiento(liq['ano'], liq['bimestre'], ficha['nit'])
    ws.merge_cells('B54:E54')
    _c(ws, 'B54', 'Plazo de declaración y pago', negrita=True, borde=True)
    _c(ws, 'F54', vence if vence else '(sin calendario cargado)',
       fmt='dd/mm/yyyy' if vence else None, negrita=True, alin='right',
       fondo=CREMA, borde=True)
    ws.merge_cells('G54:H54')
    if vence:
        _c(ws, 'G54', '=IF(F54-TODAY()<0,"Vencido hace "&TODAY()-F54&" día(s)",'
                      '"Faltan "&F54-TODAY()&" día(s)")', tam=9, borde=True)
    ws.merge_cells('B55:H55')
    _c(ws, 'B55', plazos.texto(liq['ano'], liq['bimestre'], ficha['nit'])
       + ' Se paga con el Formulario 490.', tam=9, ajuste=True)
    return ws


# --------------------------------------------------------------------- fachada

ORDEN = ['1.LIQUIDACIÓN 2593', '2.INGRESOS', '3.IVA', '4.COMPRAS', '5.RETEIVA',
         '6.ICA MUNICIPIO', '7.SEG.SOCIAL', '8.PARAMETROS', '9.ALERTAS']


def construir(liq, ficha):
    """Devuelve el workbook completo. No toca el disco: el mismo motor sirve
    para el escritorio y para las funciones sin estado de la web."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    A = {}
    # El orden de construcción no es el de presentación: cada hoja necesita las
    # anclas (filas de total) de las que ya se armaron.
    hoja_ingresos(wb, liq, ficha, A)
    hoja_compras(wb, liq, ficha, A)
    hoja_iva(wb, liq, ficha, A)
    hoja_reteiva(wb, liq, ficha, A)
    hoja_ica(wb, liq, ficha, A)
    hoja_seg_social(wb, liq, ficha, A)
    hoja_parametros(wb, liq, ficha, A)
    hoja_alertas(wb, liq, ficha, A)
    hoja_liquidacion(wb, liq, ficha, A)
    wb._sheets = [wb[t] for t in ORDEN]
    wb.active = 0
    return wb


def a_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
