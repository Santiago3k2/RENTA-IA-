# -*- coding: utf-8 -*-
"""Lee el consolidado de facturación electrónica de la DIAN y devuelve datos
limpios para la liquidación del SIMPLE.

Dos cuidados que no son opcionales:

* El archivo se abre SIEMPRE con ``read_only=True`` y se corta por la primera
  fila vacía, nunca por ``max_row``. El consolidado declara 1.048.576 filas
  (el máximo de Excel) por formato fantasma, pero solo trae ~300 con datos:
  leído así, un archivo de 63 MB se abre en centésimas de segundo.
* Las columnas se localizan por NOMBRE de encabezado, no por posición. En
  ``F.VENTA`` conviven ``Rete IVA`` (viene en cero) y ``RETEIVA`` (la buena):
  distinguirlas mal deja la declaración sin retenciones.
"""
import datetime
import re

import openpyxl

from . import parametros as P


class ErrorLectura(Exception):
    """El archivo no es el consolidado esperado."""


# ------------------------------------------------------- tipos de documento

def clasificar(tipo, lado='emitidos'):
    """«Tipo de documento» de la DIAN → (entra al agregado, signo).

    El signo es −1 para las notas crédito, que RESTAN del período, y +1 para
    todo lo demás, notas débito incluidas.

    La clasificación se hace sobre el nombre normalizado —sin tildes— y por
    palabras sueltas, no por la frase completa: la DIAN escribe «Nota de
    crédito electrónica», así que buscar el literal «nota crédito» no encuentra
    nada y las notas crédito terminan SUMANDO al ingreso en vez de restarlo.

    En el reporte de EMITIDOS viajan documentos que no son ingreso: el
    documento soporte con no obligados a facturar (prefijo DSE) y sus notas de
    ajuste son COMPRAS a personas naturales, la nómina electrónica es un gasto
    y los «application response» son acuses de recibo. En el de RECIBIDOS el
    documento soporte sí es una compra legítima.
    """
    tn = _norm(tipo)
    if not tn:
        return False, 1
    if 'application response' in tn or 'nomina' in tn:
        return False, 1
    if lado == 'emitidos' and 'documento soporte' in tn:
        return False, 1
    if 'nota' in tn and 'credito' in tn:
        return True, -1
    return True, 1


def _contar(conteo, clave):
    conteo[clave] = conteo.get(clave, 0) + 1


# ---------------------------------------------------------------- utilidades

def _norm(txt):
    """Encabezado normalizado para comparar: sin tildes, sin espacios dobles."""
    if txt is None:
        return ''
    t = str(txt).strip().lower()
    for a, b in (('á', 'a'), ('é', 'e'), ('í', 'i'), ('ó', 'o'), ('ú', 'u'), ('ñ', 'n')):
        t = t.replace(a, b)
    return re.sub(r'\s+', ' ', t)


def _num(v):
    """Número tolerante: acepta 1.234,56 (colombiano), $1,234 y flotantes.

    El consolidado trae residuos de coma flotante como -5.82e-11 en columnas
    que deberían ser cero; se limpian aquí para que no contaminen los totales.
    """
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        x = float(v)
        return 0.0 if abs(x) < 1e-6 else x
    t = str(v).strip().replace('$', '').replace(' ', '')
    if t in ('', '-', '.', ','):
        return 0.0
    if ',' in t and '.' in t:
        # el separador decimal es el último que aparezca
        t = t.replace('.', '') if t.rfind(',') > t.rfind('.') else t.replace(',', '')
        t = t.replace(',', '.')
    elif ',' in t:
        ent, _, dec = t.partition(',')
        t = ent.replace('.', '') + ('.' + dec if len(dec) in (1, 2) else dec)
    else:
        # solo puntos: separador de miles si no deja 1 o 2 decimales
        ent, _, dec = t.rpartition('.')
        if ent and len(dec) not in (1, 2):
            t = ent.replace('.', '') + dec
    try:
        x = float(t)
    except ValueError:
        return 0.0
    return 0.0 if abs(x) < 1e-6 else x


def _fecha(v):
    """dd-mm-aaaa, dd/mm/aaaa o datetime → date. None si no se entiende."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if v is None:
        return None
    t = str(v).strip()[:10]
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(t, fmt).date()
        except ValueError:
            pass
    return None


def _indices(encabezado):
    """{nombre normalizado: índice}. Conserva la primera aparición."""
    idx = {}
    for i, c in enumerate(encabezado):
        n = _norm(c)
        if n and n not in idx:
            idx[n] = i
    return idx


def _filas(ws, minimo=2):
    """Filas con datos, cortando en la primera vacía. Nunca usa max_row."""
    out = []
    for fila in ws.iter_rows(min_row=minimo, values_only=True):
        if all(c is None or str(c).strip() == '' for c in fila):
            break
        out.append(fila)
    return out


def _hoja(wb, nombre):
    for n in wb.sheetnames:
        if _norm(n) == _norm(nombre):
            return wb[n]
    return None


def _hoja_prefijo(wb, prefijo):
    """Las hojas Rp_Doc_<fecha> llevan la fecha de descarga en el nombre."""
    p = _norm(prefijo)
    for n in wb.sheetnames:
        if _norm(n).startswith(p):
            return wb[n]
    return None


# ------------------------------------------------------------------ lectores

def _leer_ventas(ws, avisos, conteo):
    enc = next(ws.iter_rows(max_row=1, values_only=True))
    idx = _indices(enc)
    faltan = [c for c in ('ing gravados', 'ing.no gravado', 'iva', 'total') if c not in idx]
    if faltan:
        raise ErrorLectura('La hoja F.VENTA no trae las columnas %s' % ', '.join(faltan))
    # «RETEIVA» es la buena; «Rete IVA» (con espacio) viene en cero.
    i_rete = None
    for i, c in enumerate(enc):
        if str(c or '').strip().upper().replace(' ', '') == 'RETEIVA' and str(c).strip() != 'Rete IVA':
            i_rete = i
    if i_rete is None:
        i_rete = idx.get('reteiva')
    if i_rete is None:
        avisos.append('F.VENTA no trae columna RETEIVA: las retenciones quedan en cero.')

    ventas, descartadas, cola = [], 0, []
    for f in _filas(ws):
        tipo = str(f[idx.get('tipo de documento', 0)] or '')
        conteo['filas'] = conteo.get('filas', 0) + 1
        if not _norm(tipo):
            cola.append(f)          # filas de totales al pie del reporte
            _contar(conteo, 'pie')
            continue
        entra, signo = clasificar(tipo, 'emitidos')
        if not entra:
            descartadas += 1
            _contar(conteo, 'excluidos')
            continue
        _contar(conteo, 'notas_credito' if signo < 0 else 'ingresos')
        ventas.append({
            'tipo': tipo,
            'nota_credito': signo < 0,
            'fecha': _fecha(f[idx['fecha emision']] if 'fecha emision' in idx else None),
            'prefijo': str(f[idx.get('prefijo', 3)] or '').strip(),
            'folio': str(f[idx.get('folio', 2)] or '').strip(),
            'nit': str(f[idx.get('nit receptor', 11)] or '').strip(),
            'tercero': str(f[idx.get('nombre receptor', 12)] or '').strip(),
            'gravado': signo * _num(f[idx['ing gravados']]),
            'no_gravado': signo * _num(f[idx['ing.no gravado']]),
            'iva': signo * _num(f[idx['iva']]),
            'reteiva': signo * _num(f[i_rete]) if i_rete is not None else 0.0,
            'total': signo * _num(f[idx['total']]),
            'estado': str(f[idx.get('estado', -2)] or '').strip(),
        })
    if descartadas:
        avisos.append('F.VENTA: se excluyeron %d documentos que no son ingreso '
                      '(documentos soporte con no obligados, acuses o nómina).' % descartadas)
    totales = _totales_al_pie(cola, {
        'no_gravado': idx['ing.no gravado'], 'gravado': idx['ing gravados'],
        'iva': idx['iva'], 'reteiva': i_rete, 'total': idx['total'],
    })
    return ventas, totales


def _totales_al_pie(cola, columnas):
    """Las filas sin tipo de documento al pie del reporte traen los totales ya
    calculados por quien armó el consolidado. Son la señal de autoverificación
    del módulo: si nuestra suma no los reproduce, algo se leyó de más o de
    menos. La primera fila numérica es la de totales por columna; un número
    suelto en una fila posterior es el gran total de la base.
    """
    tot = {}
    suelto = None
    for f in cola:
        valores = {k: _num(f[i]) for k, i in columnas.items()
                   if i is not None and i < len(f) and isinstance(f[i], (int, float))}
        if not valores:
            continue
        if not tot:
            tot = valores
        elif len(valores) == 1 and suelto is None:
            suelto = list(valores.values())[0]
    if suelto is not None:
        tot['base'] = suelto
    return tot


def _leer_compras(ws, avisos):
    enc = list(next(ws.iter_rows(max_row=1, values_only=True)))
    idx = _indices(enc)
    if 'iva' not in idx or 'total' not in idx:
        raise ErrorLectura('La hoja F.COMPRA no trae las columnas IVA y Total.')
    # La base gravable viene en la columna sin encabezado que precede a «Total».
    i_total = idx['total']
    i_base = i_total - 1 if i_total > 0 and _norm(enc[i_total - 1]) == '' else None
    if i_base is None:
        avisos.append('F.COMPRA no trae la columna de base gravable: se deduce '
                      'dividiendo el IVA entre la tarifa general.')

    compras, cola = [], []
    for f in _filas(ws):
        tipo = str(f[idx.get('tipo de documento', 0)] or '')
        if not _norm(tipo):
            cola.append(f)
            continue
        entra, signo = clasificar(tipo, 'recibidos')
        if not entra:
            continue
        iva = signo * _num(f[idx['iva']])
        base = signo * _num(f[i_base]) if i_base is not None else 0.0
        compras.append({
            'fecha': _fecha(f[idx.get('fecha emision')] if 'fecha emision' in idx else None),
            'nit': str(f[idx.get('nit emisor', 9)] or '').strip(),
            'proveedor': str(f[idx.get('nombre emisor', 10)] or '').strip(),
            'prefijo': str(f[idx.get('prefijo', 3)] or '').strip(),
            'folio': str(f[idx.get('folio', 2)] or '').strip(),
            'base': base,
            'iva': iva,
            'total': signo * _num(f[idx['total']]),
        })
    cols = {'iva': idx['iva'], 'total': i_total}
    if i_base is not None:
        cols['base'] = i_base
    return compras, _totales_al_pie(cola, cols)


def _leer_reteiva(ws, avisos):
    """Certificados de los agentes retenedores. Maquetación libre: la tabla
    empieza donde aparece el encabezado con «Nit» y «Base»."""
    filas = list(ws.iter_rows(values_only=True))
    inicio = None
    for i, f in enumerate(filas):
        nn = [_norm(c) for c in f]
        if 'nit' in nn and 'base' in nn:
            inicio = i
            idx = _indices(f)
            break
    if inicio is None:
        avisos.append('La hoja RETEIVA no tiene un encabezado reconocible; '
                      'no se pudo conciliar contra los certificados.')
        return []

    agentes = []
    for f in filas[inicio + 1:]:
        if all(c is None or str(c).strip() == '' for c in f):
            break
        etiqueta = _norm(f[idx.get('no.', 0)])
        if etiqueta.startswith('total') or etiqueta == 'ok':
            break
        if not etiqueta:
            continue
        agentes.append({
            'nit': str(f[idx['nit']] or '').strip(),
            'tercero': str(f[idx.get('tercero', 2)] or '').strip(),
            'base': _num(f[idx['base']]),
            'contabilidad': _num(f[idx.get('contabilidad')]) if 'contabilidad' in idx else 0.0,
            'certificado': _num(f[idx.get('certificado')]) if 'certificado' in idx else 0.0,
            'diferencia': _num(f[idx.get('diferencias')]) if 'diferencias' in idx else 0.0,
        })
    return agentes


def _leer_seg_social(ws, avisos):
    """Reporte de PILA con maquetación libre: bloques de encabezado + una fila
    de planilla + una de total, repetidos por administradora."""
    filas = list(ws.iter_rows(values_only=True))
    planillas = []
    idx = None
    for f in filas:
        nn = [_norm(c) for c in f]
        if 'periodo' in nn and 'nombre administradora' in nn:
            idx = _indices(f)
            continue
        if idx is None:
            continue
        periodo = f[idx['periodo']] if idx['periodo'] < len(f) else None
        if periodo is None or _norm(periodo).startswith('total'):
            continue
        admin = f[idx['nombre administradora']] if idx['nombre administradora'] < len(f) else None
        if not admin:
            continue
        planillas.append({
            'periodo': str(periodo).strip(),
            'administradora': str(admin).strip(),
            'fecha_pago': _fecha(f[idx['fecha pago']]) if 'fecha pago' in idx else None,
            'planilla': str(f[idx['planilla numero']] or '').strip() if 'planilla numero' in idx else '',
            'trabajadores': int(_num(f[idx['numero trabajadores']])) if 'numero trabajadores' in idx else 0,
            'dias_mora': int(_num(f[idx['dias en mora']])) if 'dias en mora' in idx else 0,
            'aporte': _num(f[idx['valor sin mora']]) if 'valor sin mora' in idx else 0.0,
            'intereses': _num(f[idx['valor intereses']]) if 'valor intereses' in idx else 0.0,
            'total': _num(f[idx['total pagado']]) if 'total pagado' in idx else 0.0,
        })
    if not planillas:
        avisos.append('No se encontraron planillas de seguridad social: el descuento '
                      'por aportes a pensión queda en cero y hay que capturarlo a mano.')
    return planillas


# ------------------------------------------- modo crudo: solo Rp_Doc / Rp_Docpras

def _derivar_ventas(ws, avisos, conteo, tarifa_iva, tarifa_reteiva):
    """Ventas a partir del reporte crudo de documentos emitidos de la DIAN.

    El reporte solo trae IVA y Total; el desglose se deduce, y da exacto:

        gravado    = IVA ÷ tarifa general
        no gravado = Total − gravado − IVA
        ReteIVA    = IVA × tarifa de retención

    Una factura íntegramente excluida de IVA cae sola en su sitio: IVA 0 →
    gravado 0 y todo el total como no gravado.
    """
    enc = next(ws.iter_rows(max_row=1, values_only=True))
    idx = _indices(enc)
    for c in ('iva', 'total'):
        if c not in idx:
            raise ErrorLectura('El reporte de documentos emitidos no trae la columna %s.'
                               % c.upper())

    # El IVA viene ya redondeado al peso en la factura, así que dividirlo entre
    # la tarifa devuelve la base con un error de hasta medio peso ÷ tarifa. Por
    # debajo de ese umbral el «no gravado» que sobra es ruido del redondeo, no
    # un ingreso excluido: se le devuelve al gravado y la factura queda limpia.
    # Con el umbral fijo de un peso que había antes, una factura íntegramente
    # gravada salía con un «no gravado» negativo de un par de pesos.
    ruido = 0.5 / tarifa_iva + 1.0

    ventas, descartadas = [], 0
    for f in _filas(ws):
        tipo = str(f[idx.get('tipo de documento', 0)] or '')
        conteo['filas'] = conteo.get('filas', 0) + 1
        if not _norm(tipo):
            _contar(conteo, 'pie')
            continue
        entra, signo = clasificar(tipo, 'emitidos')
        if not entra:
            descartadas += 1
            _contar(conteo, 'excluidos')
            continue
        _contar(conteo, 'notas_credito' if signo < 0 else 'ingresos')
        iva = _num(f[idx['iva']])
        total = _num(f[idx['total']])
        gravado = iva / tarifa_iva if iva else 0.0
        no_gravado = total - gravado - iva
        if abs(no_gravado) < ruido:
            gravado, no_gravado = total - iva, 0.0
        ventas.append({
            'tipo': tipo,
            'nota_credito': signo < 0,
            'fecha': _fecha(f[idx['fecha emision']] if 'fecha emision' in idx else None),
            'prefijo': str(f[idx.get('prefijo', 3)] or '').strip(),
            'folio': str(f[idx.get('folio', 2)] or '').strip(),
            'nit': str(f[idx.get('nit receptor', 11)] or '').strip(),
            'tercero': str(f[idx.get('nombre receptor', 12)] or '').strip(),
            'gravado': signo * gravado,
            'no_gravado': signo * no_gravado,
            'iva': signo * iva,
            'reteiva': signo * iva * tarifa_reteiva,
            'total': signo * total,
            'estado': str(f[idx.get('estado', -2)] or '').strip(),
        })
    if descartadas:
        avisos.append('Del reporte de emitidos se excluyeron %d documentos que no son '
                      'ingreso: documentos soporte con no obligados (son compras a '
                      'personas naturales), nómina y acuses.' % descartadas)
    return ventas


def _derivar_compras(ws, avisos, tarifa_iva):
    """Compras a partir del reporte crudo de documentos recibidos."""
    enc = next(ws.iter_rows(max_row=1, values_only=True))
    idx = _indices(enc)
    if 'iva' not in idx or 'total' not in idx:
        raise ErrorLectura('El reporte de documentos recibidos no trae IVA y Total.')

    compras, acuses = [], 0
    for f in _filas(ws):
        tipo = str(f[idx.get('tipo de documento', 0)] or '')
        if not _norm(tipo):
            continue
        entra, signo = clasificar(tipo, 'recibidos')
        if not entra:
            acuses += 1          # acuses de recibo y nómina: no son compras
            continue
        iva = _num(f[idx['iva']])
        compras.append({
            'fecha': _fecha(f[idx['fecha emision']] if 'fecha emision' in idx else None),
            'nit': str(f[idx.get('nit emisor', 9)] or '').strip(),
            'proveedor': str(f[idx.get('nombre emisor', 10)] or '').strip(),
            'prefijo': str(f[idx.get('prefijo', 3)] or '').strip(),
            'folio': str(f[idx.get('folio', 2)] or '').strip(),
            'base': signo * (iva / tarifa_iva if iva else 0.0),
            'iva': signo * iva,
            'total': signo * _num(f[idx['total']]),
        })
    if acuses:
        avisos.append('Del reporte de recibidos se excluyeron %d documentos que no son '
                      'compra: «Application response» (acuses de recibo) y nómina '
                      'electrónica.' % acuses)
    return compras


def _derivar_reteiva(ventas, tarifa_reteiva):
    """Conciliación por agente retenedor, agregando las ventas del bimestre.

    El certificado que expide cada agente no está en ningún archivo de la DIAN:
    la columna queda en cero y editable, y la diferencia se calcula sola cuando
    el contador la diligencie.
    """
    por_nit = {}
    for v in ventas:
        if not v['iva']:
            continue
        clave = v['nit'] or v['tercero']
        ag = por_nit.setdefault(clave, {'nit': v['nit'], 'tercero': v['tercero'],
                                        'base': 0.0})
        ag['base'] += v['iva']
    agentes = []
    for ag in sorted(por_nit.values(), key=lambda a: a['tercero']):
        ag['contabilidad'] = round(ag['base'] * tarifa_reteiva, 2)
        ag['certificado'] = 0.0
        ag['diferencia'] = 0.0
        agentes.append(ag)
    return agentes


def periodos(ventas):
    """Períodos (año, bimestre) presentes en el archivo, del más cargado al menos.

    Es lo que evita tener que acertar el bimestre a mano: el archivo ya dice de
    qué período es. Un consolidado descargado para un bimestre trae un solo
    período; si trae varios, el primero de la lista es el que manda y los demás
    quedan a la vista para procesarlos aparte.
    """
    cajas = {}
    for v in ventas:
        f = v['fecha']
        if not f:
            continue
        clave = (f.year, P.bimestre_de(f.month))
        c = cajas.setdefault(clave, {'ano': clave[0], 'bimestre': clave[1],
                                     'documentos': 0, 'total': 0.0})
        c['documentos'] += 1
        c['total'] += v['total']
    return sorted(cajas.values(), key=lambda c: (-c['documentos'], c['ano'], c['bimestre']))


# ------------------------------------------------------------------- fachada

def leer(ruta, tarifa_iva=0.19, tarifa_reteiva=0.15):
    """Archivo de la DIAN → ventas, compras, reteiva, seg_social y avisos.

    Acepta las dos formas del archivo:

    * **crudo** — la exportación tal como la entrega la DIAN, con solo las hojas
      `Rp_Doc_<fecha>` (emitidos) y `Rp_Docpras` (recibidos). Es el caso normal:
      el motor deriva de ahí el desglose de ingresos, las compras y la
      conciliación de ReteIVA.
    * **trabajado** — un consolidado que ya trae `F.VENTA`, `F.COMPRA`, `RETEIVA`
      y `S.SOCIAL` hechas a mano. Si están, se respetan: el trabajo del contador
      manda sobre lo que el motor deduciría.
    """
    avisos = []
    conteo = {}
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        hv = _hoja(wb, 'F.VENTA')
        hc = _hoja(wb, 'F.COMPRA')
        emitidos = _hoja_prefijo(wb, 'Rp_Doc_')
        recibidos = _hoja(wb, 'Rp_Docpras')

        if hv is not None:
            origen = 'trabajado'
            ventas, tot_ventas = _leer_ventas(hv, avisos, conteo)
            if hc is not None:
                compras, tot_compras = _leer_compras(hc, avisos)
            else:
                compras, tot_compras = [], {}
                avisos.append('El archivo no trae hoja F.COMPRA: el IVA descontable '
                              'queda en cero.')
        elif emitidos is not None:
            origen = 'crudo'
            ventas = _derivar_ventas(emitidos, avisos, conteo, tarifa_iva, tarifa_reteiva)
            tot_ventas = {}
            if recibidos is not None:
                compras = _derivar_compras(recibidos, avisos, tarifa_iva)
            else:
                compras = []
                avisos.append('El archivo no trae el reporte de documentos recibidos '
                              '(Rp_Docpras): el IVA descontable queda en cero y hay que '
                              'descargarlo aparte.')
            tot_compras = {}
        else:
            raise ErrorLectura(
                'El archivo no tiene ni la hoja de documentos emitidos (Rp_Doc_…) ni '
                'una hoja F.VENTA ya trabajada. Hojas encontradas: %s'
                % ', '.join(wb.sheetnames))

        hr = _hoja(wb, 'RETEIVA')
        if hr is not None:
            reteiva = _leer_reteiva(hr, avisos)
        else:
            reteiva = _derivar_reteiva(ventas, tarifa_reteiva)
            if reteiva:
                avisos.append('La conciliación de ReteIVA se armó agregando las facturas '
                              'por agente retenedor. La columna del certificado queda en '
                              'cero y editable: ese dato no está en ningún archivo de la '
                              'DIAN y hay que tomarlo de los certificados.')
        hs = _hoja(wb, 'S.SOCIAL')
        seg = _leer_seg_social(hs, avisos) if hs is not None else []
        if hs is None:
            # La planilla de PILA no sale de la facturación electrónica: es otro
            # sistema. Sin ella el anticipo queda MÁS ALTO de lo que debe, así
            # que esto no es un detalle menor.
            avisos.append('El archivo no trae el reporte de seguridad social (PILA): el '
                          'descuento por aportes a pensión del empleador queda en cero y '
                          'el anticipo sale más alto de lo debido. Añada las planillas en '
                          'las filas amarillas de la hoja 7 y la liquidación se recalcula '
                          'sola.')

        n_emitidos = len(_filas(emitidos)) if emitidos is not None else 0
        n_recibidos = len(_filas(recibidos)) if recibidos is not None else 0
    finally:
        wb.close()

    notas = [v for v in ventas if v.get('nota_credito')]
    if notas:
        avisos.append('Se detectaron %d nota(s) crédito por $%s: se restaron del ingreso '
                      'del período.'
                      % (len(notas),
                         f'{abs(sum(v["total"] for v in notas)):,.2f}'.replace(',', '.')))

    return {
        'origen': origen,
        'ventas': ventas,
        'compras': compras,
        'reteiva': reteiva,
        'seg_social': seg,
        'periodos': periodos(ventas),
        'clasificacion': conteo,
        'totales_declarados': {'ventas': tot_ventas, 'compras': tot_compras},
        'conteo_reportes': {'emitidos': n_emitidos, 'recibidos': n_recibidos},
        'avisos': avisos,
    }
