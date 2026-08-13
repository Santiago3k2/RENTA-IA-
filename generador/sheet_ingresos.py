# -*- coding: utf-8 -*-
"""Hoja 3 · Ingresos — notas 11 y 13, alimentada por datos.py."""
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from build_lib import *

import inflacionario

AUT = 'Generador renta'


def build(wb, A, C, T):
    ws = wb.create_sheet('3. Ingresos')
    ws.sheet_properties.tabColor = BLUE[2:]
    s = Sheet(ws, 'B', 'G', BLUE, BLUE_T)
    s.widths({'A': 2.3, 'B': 31.0, 'C': 41.0, 'D': 16.5,
              'E': 16.5, 'F': 15.0, 'G': 16.5})
    UVT = "'1. Resumen'!$G$9"

    s.band('03  ·  RENGLONES R32 / R33 / R36 / R58 / R74',
           'Ingresos por cédula',
           T['sub_ingresos'], h_desc=44.0)
    s.legend('Las casillas con fondo crema son editables: «Certificado» viene prediligenciado con lo reportado —reemplácelo por el valor del certificado— y las casillas de costos y deducciones quedan abiertas.')

    HEAD = [('B', None, 'Pagador', 'left'), ('C', None, 'Concepto', 'left'),
            ('D', None, 'Reportado', 'right'), ('E', None, 'Certificado', 'right'),
            ('F', None, 'Diferencia', 'right'), ('G', None, 'Ingreso', 'right')]

    def partidas(rows, filas=None):
        """Pinta las partidas de una subcédula. Devuelve (primera, última) fila.

        `filas`, si se pasa, recibe la fila en que quedó cada partida. Lo usa el
        componente inflacionario para sumar exactamente las de rendimientos
        financieros: sin eso habría que pegar un total, y en este libro todo
        total es fórmula.
        """
        r0 = s.r
        if not rows:
            r = s.datarow(h=20.0)
            s.txt(r, 'B', 'Sin registros informados en la exógena', 'F', sz=9.0, i=True, color=MUTED, v='center')
            s.money(r, 'G', 0)
            return r0, r
        for i, (pag, con, val) in enumerate(rows):
            r = s.datarow(zebra=bool(i % 2), h=24.0)
            s.txt(r, 'B', pag, v='center')
            s.txt(r, 'C', con, sz=9.0, color=TEXT2, v='center')
            s.money(r, 'D', val)
            s.mark_input(r, 'E')
            s.money(r, 'E', val)
            s.money(r, 'F', f'=E{r}-D{r}', color=MUTED)
            s.money(r, 'G', f'=D{r}')
            if filas is not None:
                filas.append(r)
        r1 = s.r - 1
        ws.conditional_formatting.add(
            f'F{r0}:F{r1}',
            FormulaRule(formula=[f'F{r0}<>0'], font=Font(bold=True, color=RED)))
        return r0, r1

    def tope60(subcedula, ref):
        """Línea de control del 60%: es un límite, no una partida de la depuración.

        Va como referencia y no como subtotal a propósito. Puesto en negrita al
        pie de la depuración se lee como un costo más —uno que además nadie ha
        certificado—, cuando lo único que dice es hasta dónde podrían llegar los
        costos el día que el contribuyente los soporte.
        """
        r = s.datarow(h=19.0)
        s.txt(r, 'B',
              f'Control · límite máximo de costos y deducciones de {subcedula} — '
              f'60% de sus ingresos. Es un tope, no un valor a declarar.',
              'F', sz=8.6, i=True, color=MUTED, v='center')
        s.money(r, 'G', f'=ROUND(G{ref}*0.6,0)', sz=9.2, b=True, color=GOLD)
        return r

    # ══════════════ RENTAS DE TRABAJO — R32 ══════════════
    s.section('RENTAS DE TRABAJO — R32 (ART. 103 E.T.)')
    s.head(HEAD)
    t0, t1 = partidas(C['rentas_trabajo'])
    if C['rentas_trabajo']:
        ws[f'E{t0}'].comment = Comment(
            'NOTA 11 — «Reportado» es la información del prevalidador DIAN.\n'
            'Esta casilla es editable: escriba el valor del certificado de ingresos\n'
            'y retenciones (formulario 220). La columna «Diferencia» se calcula sola\n'
            'y «Ingreso» conserva el valor reportado.', AUT, 320, 100)
    if T.get('nota_trabajo'):
        s.note(T['nota_trabajo'], h=29.25)
    A['r32'] = s.subtotal('Subtotal rentas de trabajo (R32)', f'=SUM(G{t0}:G{t1})')
    s.gap()

    # ── INCRNGO de las rentas de trabajo (NOTA 13) ──
    s.section('INGRESOS NO CONSTITUTIVOS DE RENTA — R33  ·  RENTAS DE TRABAJO')
    s.head([('B', None, 'Pagador', 'left'), ('C', None, 'Concepto', 'left'),
            ('D', None, 'Reportado', 'right'), ('E', None, 'Certificado', 'right'),
            ('F', None, 'Diferencia', 'right'), ('G', None, 'Valor', 'right')])
    n0, n1 = partidas(C['incrngo_trabajo'])
    s.note('Arts. 55 y 56 E.T. — los aportes obligatorios a pensión y salud del trabajador son ingreso no constitutivo de renta ni ganancia ocasional y no se someten al límite del 40%.')
    A['r33'] = s.subtotal('Subtotal INCRNGO de rentas de trabajo (R33)', f'=SUM(G{n0}:G{n1})')
    s.gap()

    # ── Rentas exentas de las rentas de trabajo (NOTA 13) ──
    s.section('RENTAS EXENTAS DE LAS RENTAS DE TRABAJO — SUJETAS AL LÍMITE DEL ART. 336 NUM. 3')
    s.head([('B', 'C', 'Concepto', 'left'), ('D', None, 'Base de cálculo', 'right'),
            ('E', 'F', 'Regla aplicada', 'left'), ('G', None, 'Valor', 'right')])
    x0 = s.r
    r = s.datarow(h=24.0)
    s.txt(r, 'B', 'R36 · Cesantías e intereses de cesantías exentos — art. 206 num. 4', 'C', v='center')
    idx = C.get('indices_cesantias_exentas', [])
    base_ces = '=' + '+'.join(f'G{t0 + i}' for i in idx) if idx else 0
    s.money(r, 'D', base_ces)
    s.txt(r, 'E', 'Exención plena · muy por debajo de 350 UVT mensuales', 'F', sz=8.0, i=True, color=MUTED, v='center')
    s.money(r, 'G', f'=D{r}')
    A['x_ces'] = r
    r = s.datarow(zebra=True, h=24.0)
    s.txt(r, 'B', 'Renta exenta del 25% — art. 206 num. 10', 'C', v='center')
    excl = C.get('base_25_excluir', 0)   # honorarios R43 sujetos a costos: fuera de la base
    s.money(r, 'D', f'=G{A["r32"]}-G{A["r33"]}-G{A["x_ces"]}' + (f'-{excl}' if excl else ''))
    s.txt(r, 'E', '25% de la base · tope anual de 790 UVT' + (' · excluye honorarios R43 sujetos a costos' if excl else ''),
          'F', sz=8.0, i=True, color=MUTED, v='center')
    s.money(r, 'G', f'=ROUND(MIN(D{r}*0.25,790*{UVT}),0)')
    A['x_25'] = r
    r = s.datarow(h=24.0)
    s.txt(r, 'B', 'Otras rentas exentas de las rentas de trabajo', 'C', v='center')
    s.txt(r, 'E', 'Aportes voluntarios a pensión y cuentas AFC · arts. 126-1 y 126-4', 'F', sz=8.0, i=True, color=MUTED, v='center')
    s.mark_input(r, 'G'); s.money(r, 'G', 0)
    A['x_otras'] = r
    x1 = s.r - 1
    s.note('NOTA 13 — Las rentas exentas de trabajo se detallan aquí, pero se imputan en la hoja «6. Liquidación» dentro del límite conjunto del art. 336 num. 3 (40% de la renta líquida, tope 1.340 UVT).', h=22.0)
    A['exentas_trab'] = s.subtotal('Subtotal rentas exentas de trabajo', f'=SUM(G{x0}:G{x1})')
    s.gap()

    tope_t = tope60('las rentas de trabajo', A['r32'])
    ws[f'G{tope_t}'].comment = Comment(
        'NOTA 13 — 60% de los ingresos de esta subcédula.\n'
        'Es el techo de los costos y deducciones imputables, no un valor\n'
        'que se declare: mientras no haya costos certificados, la depuración\n'
        'de esta subcédula va en cero.', AUT, 320, 90)
    A['tope60_trab'] = tope_t
    s.gap(); s.gap()

    # ══════════════ RENTAS DE CAPITAL — R58 ══════════════
    s.section('RENTAS DE CAPITAL — R58 (ART. 338 E.T.)')
    s.head(HEAD)
    filas_cap = []
    c0, c1 = partidas(C['rentas_capital'], filas_cap)
    if T.get('nota_capital'):
        s.note(T['nota_capital'], h=29.25)
    A['r58'] = s.subtotal('Subtotal rentas de capital (R58)', f'=SUM(G{c0}:G{c1})')
    s.gap()

    s.section('DEPURACIÓN DE LAS RENTAS DE CAPITAL')
    s.head([('B', 'C', 'Concepto', 'left'), ('D', 'F', 'Referencia normativa', 'left'),
            ('G', None, 'Valor', 'right')])
    # ── Componente inflacionario de los rendimientos financieros ──
    # Una sola línea: el porcentaje del decreto aplicado sobre las partidas de
    # R58 que el prevalidador marcó con R59, sumadas dentro de la propia
    # fórmula. La fila de base que había aquí se quitó por redundante —el
    # detalle ya está arriba, partida por partida— pero la casilla queda crema:
    # trae la fórmula prediligenciada y admite que se sobrescriba si el
    # certificado obliga a otra cifra.
    ci = inflacionario.calcular(C)
    r = s.datarow()
    if ci['porcentaje']:
        rf = [filas_cap[i] for i in ci['indices'] if i < len(filas_cap)]
        pct = ci['porcentaje']
        # 0.5543 → «55,43», con coma decimal: el libro es para leerlo en español.
        etiqueta = ('%.2f' % (pct * 100)).replace('.', ',')
        s.txt(r, 'B', 'R59 · Ingresos no constitutivos de renta — componente '
                      'inflacionario (%s%%)' % etiqueta, 'C')
        s.txt(r, 'D', 'Art. 38 E.T. — %s%% de los rendimientos financieros que el prevalidador '
                      'marcó con R59; porcentaje fijado por decreto para el año gravable %s'
                      % (etiqueta, C['ano_gravable']), 'F', sz=8.0, i=True, color=MUTED)
        s.mark_input(r, 'G')
        s.money(r, 'G', ('=ROUND((%s)*%s,0)' % ('+'.join(f'G{x}' for x in rf), pct))
                        if rf else 0)
        ws[f'G{r}'].comment = Comment(
            'Se calcula solo: los rendimientos financieros de la lista de arriba\n'
            'que la DIAN marcó con R59, por %s%%. El porcentaje lo fija el\n'
            'Gobierno por decreto cada año y está en generador\\inflacionario.py.\n'
            'La casilla es editable: si el certificado obliga a otra cifra,\n'
            'escríbala encima y el resto del libro se recalcula.' % etiqueta,
            AUT, 340, 110)
    else:
        # Año sin decreto cargado: casilla abierta y en cero, como antes.
        s.txt(r, 'B', 'R59 · Ingresos no constitutivos de renta — componente inflacionario', 'C')
        s.txt(r, 'D', 'Art. 38 E.T. — porcentaje fijado por decreto reglamentario; se liquida en cero mientras no esté publicado', 'F', sz=8.0, i=True, color=MUTED)
        s.mark_input(r, 'G'); s.money(r, 'G', 0)
    A['incr58'] = r
    r = s.datarow(zebra=True)
    s.txt(r, 'B', 'R67 · Costos y deducciones procedentes', 'C')
    s.txt(r, 'D', 'Art. 339 E.T. — comisión de administración inmobiliaria, predial, seguros, reparaciones y depreciación de los inmuebles arrendados', 'F', sz=8.0, i=True, color=MUTED)
    s.mark_input(r, 'G'); s.money(r, 'G', 0)
    A['costos58'] = r
    ws[f'G{r}'].comment = Comment(
        'NOTA 13 — Casilla editable para los costos y deducciones procedentes\n'
        'de esta subcédula. No deben superar el tope del 60% que aparece abajo.', AUT, 300, 70)
    A['dep_cap'] = s.subtotal('Total depuración de las rentas de capital',
                              f'=G{A["incr58"]}+G{A["costos58"]}')
    s.note('NOTA 13 — El componente inflacionario de los rendimientos financieros (art. 38 E.T.) se calcula solo sobre las partidas que el prevalidador marcó con R59; los costos y deducciones (art. 339 E.T.) siguen abiertos porque no se informan en la exógena y son plenamente procedentes con certificado. El total es lo que se resta del R58 en la hoja 6.'
           if ci['porcentaje'] else
           'NOTA 13 — Las rentas de capital admiten ingresos no constitutivos de renta y costos y deducciones procedentes (art. 339 E.T.). Diligéncielos con los soportes: no se informan en la exógena y son plenamente procedentes con certificado. El total es lo que se resta del R58 en la hoja 6.', h=22.0)
    A['tope60_cap'] = tope60('las rentas de capital', A['r58'])
    s.gap()

    if T.get('callout_reclasificacion'):
        titulo, cuerpo = T['callout_reclasificacion']
        s.callout(titulo, cuerpo, h=48.0)
        s.gap()

    # ══════════════ RENTAS NO LABORALES — R74 ══════════════
    s.section('RENTAS NO LABORALES — R74')
    s.head(HEAD)
    l0, l1 = partidas(C['rentas_no_laborales'])
    A['r74'] = s.subtotal('Subtotal rentas no laborales (R74)', f'=SUM(G{l0}:G{l1})')
    s.gap()

    s.section('DEPURACIÓN DE LAS RENTAS NO LABORALES')
    s.head([('B', 'C', 'Concepto', 'left'), ('D', 'F', 'Referencia normativa', 'left'),
            ('G', None, 'Valor', 'right')])
    r = s.datarow()
    s.txt(r, 'B', 'Ingresos no constitutivos de renta', 'C')
    s.txt(r, 'D', 'Arts. 36 a 57-2 E.T. — no se informan en la exógena para esta subcédula', 'F', sz=8.0, i=True, color=MUTED)
    s.mark_input(r, 'G'); s.money(r, 'G', 0)
    A['incr74'] = r
    r = s.datarow(zebra=True)
    s.txt(r, 'B', 'Costos y deducciones procedentes', 'C')
    s.txt(r, 'D', 'Art. 341 E.T. — costos y gastos con relación de causalidad, necesidad y proporcionalidad', 'F', sz=8.0, i=True, color=MUTED)
    s.mark_input(r, 'G'); s.money(r, 'G', 0)
    A['costos74'] = r
    ws[f'G{r}'].comment = Comment(
        'NOTA 13 — Casilla editable para los costos y deducciones procedentes\n'
        'de esta subcédula. No deben superar el tope del 60% que aparece abajo.', AUT, 300, 70)
    s.note('NOTA 13 — Las dos casillas anteriores quedan en cero mientras no haya soportes: la exógena no informa costos de esta subcédula. La línea de control de abajo solo dice hasta dónde podrían llegar; no es un valor que se declare.', h=22.0)
    A['tope60_nol'] = tope60('las rentas no laborales', A['r74'])
    s.gap(); s.gap()

    # ══════════════ CONSOLIDADO ══════════════
    s.section('CONSOLIDADO DE INGRESOS')
    s.head([('B', 'F', 'Subcédula', 'left'), ('G', None, 'Valor', 'right')])
    for lab, ref in [('R32 · Rentas de trabajo', A['r32']),
                     ('R58 · Rentas de capital', A['r58']),
                     ('R74 · Rentas no laborales', A['r74'])]:
        r = s.datarow(zebra=True)
        s.txt(r, 'B', lab, 'F', b=True)
        s.money(r, 'G', f'=G{ref}')
    A['ing_total'] = s.hero('TOTAL INGRESOS BRUTOS CÉDULA GENERAL',
                            f'=G{A["r32"]}+G{A["r58"]}+G{A["r74"]}')
    r = s.datarow(h=18.0)
    s.txt(r, 'B', 'Referencia · 60% del total de la cédula general', 'F', sz=8.6, i=True, color=MUTED)
    s.money(r, 'G', f'=ROUND(G{A["ing_total"]}*0.6,0)', sz=9.2, b=True, color=GOLD)
    s.note('NOTA 13 — El tope del 60% se calcula sobre los ingresos de cada subcédula por separado; esta línea es solo una referencia agregada.')
    s.gap(); s.gap()

    # ══════════════ LIQUIDACIÓN PRIVADA ══════════════
    # Las tres subcédulas se depuran arriba por separado —esa separación es la
    # regla del libro y no se toca— pero el contribuyente necesita ver en una
    # sola columna qué le descontaron y con qué se queda. Esto es esa columna.
    #
    # Los cinco escalones de abajo del subtotal salen de la hoja 6 POR NOMBRE
    # DEFINIDO: esta hoja se arma antes que aquella, así que no puede citar sus
    # filas. La ventaja de fondo es que son literalmente la misma celda, de
    # modo que las dos hojas no pueden decir cifras distintas.
    s.section('LIQUIDACIÓN PRIVADA')
    s.head([('B', 'F', 'Concepto', 'left'), ('G', None, 'Valor', 'right')])

    def escalon(etiqueta, formula, zebra=False, b=False, color=TEXT):
        f = s.datarow(zebra=zebra, h=19.5)
        s.txt(f, 'B', etiqueta, 'F', sz=9.4, b=b, color=color, v='center')
        s.money(f, 'G', formula, sz=9.6, b=b, color=color)
        return f

    e_ing = escalon('Ingresos brutos de la cédula general', f'=G{A["ing_total"]}', b=True)
    e_inc = escalon('(−) Ingresos no constitutivos de renta  ·  R33 + R59 + no laborales',
                    f'=-(G{A["r33"]}+G{A["incr58"]}+G{A["incr74"]})', zebra=True)
    e_cos = escalon('(−) Costos y deducciones procedentes  ·  arts. 339 y 341 E.T.',
                    f'=-(G{A["costos58"]}+G{A["costos74"]})')
    e_exe = escalon('(−) Rentas exentas y deducciones imputables aceptadas  ·  límite del 40%',
                    '=EXENTAS_ACEPTADAS', zebra=True)
    e_ded = escalon('(−) Deducciones no sujetas al límite  ·  1% factura electrónica y dependientes',
                    '=DEDUC_SIN_LIMITE')
    e_rlg = s.subtotal('(=) Renta líquida gravable de la cédula general',
                       f'=SUM(G{e_ing}:G{e_ded})')
    ws[f'G{e_rlg}'].comment = Comment(
        'Tiene que coincidir con la renta líquida gravable de la hoja 6.\n'
        'Aquí se llega sumando los escalones de arriba y allá por la\n'
        'depuración separada de cada subcédula: son dos caminos al mismo\n'
        'número, y que coincidan es la comprobación.', AUT, 330, 90)
    e_imp = escalon('Impuesto sobre la renta líquida gravable  ·  art. 241 E.T.',
                    '=IMPUESTO_RENTA', zebra=True)
    escalon('(−) Descuentos tributarios', '=-DESCUENTOS_TRIB')
    A['imp_neto_ing'] = s.hero('IMPUESTO NETO DE RENTA', '=IMPUESTO_NETO')
    s.note('NOTA 13 — Este bloque no liquida nada: reúne en una sola columna lo que ya se depuró por separado en cada subcédula y lo que la hoja 6 imputa dentro del límite del 40%. Sirve para ver de un vistazo cuánto de los ingresos brutos quedó gravado y cuánto se descontó legalmente. Las cifras son las mismas celdas de la hoja 6, no una copia.', h=29.25)

    ws.freeze_panes = 'A6'
    return ws
