# -*- coding: utf-8 -*-
"""Hoja 2 · Patrimonio — notas 1 a 10, alimentada por datos.py."""
from openpyxl.comments import Comment
from build_lib import *

AUT = 'Generador renta'


def build(wb, A, C, T):
    ws = wb.create_sheet('2. Patrimonio')
    ws.sheet_properties.tabColor = TEAL[2:]
    s = Sheet(ws, 'B', 'H', TEAL, TEAL_T)
    s.widths({'A': 2.3, 'B': 33.0, 'C': 23.0, 'D': 22.0,
              'E': 16.0, 'F': 16.5, 'G': 16.5, 'H': 17.5})
    ag, ant = C['ano_gravable'], C['ano_anterior']

    s.band('02  ·  RENGLONES R29 / R30 / R31',
           f'Patrimonio a 31 de diciembre de {ag}',
           T['sub_patrimonio'], h_desc=44.0)
    s.legend(f'Las casillas con fondo crema son editables: diligéncielas con los certificados, los extractos y la declaración del AG {ant}. Todos los totales son fórmulas y se recalculan solos.')

    sub = {}

    # ══════════════ A · EFECTIVO  (NOTA 1) ══════════════
    s.section('A · EFECTIVO')
    s.head([('B', 'E', 'Detalle', 'left'), ('F', None, 'Valor real', 'right'),
            ('G', None, 'Ajuste Fiscal', 'right'), ('H', None, 'Valor patrimonial', 'right')])
    ef0 = s.r
    for i, lab in enumerate(C['efectivo']):
        r = s.datarow(zebra=bool(i))
        s.txt(r, 'B', lab, 'E')
        s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(N(F{r})+N(G{r}),0)')
    ef1 = s.r - 1
    ws[f'F{ef0}'].comment = Comment(
        'NOTA 1 — Estas casillas se dejan siempre vacías.\n'
        'Diligencie aquí el efectivo en caja a 31-dic según el arqueo o la\n'
        'certificación del contribuyente: la exógena nunca lo reporta.', AUT, 260, 100)
    s.note('Art. 261 E.T. — el efectivo en caja hace parte del patrimonio bruto pero no viaja por exógena. Las casillas quedan siempre abiertas para diligenciarlas por concepto.')
    sub['A'] = s.subtotal('Subtotal Efectivo', f'=SUM(H{ef0}:H{ef1})')
    s.gap()

    # ══════════════ B · SALDOS EN CUENTAS BANCARIAS  (NOTA 2) ══════════════
    s.section('B · SALDOS EN CUENTAS BANCARIAS')
    s.head([('B', None, 'Entidad y cuenta', 'left'), ('C', None, 'Tipo', 'left'),
            ('D', 'E', 'No. Cuenta', 'left'), ('F', None, 'Valor real', 'right'),
            ('G', None, 'Ajuste Fiscal', 'right'), ('H', None, 'Valor patrimonial', 'right')])
    cb0 = s.r
    for i, (ent, tipo, val) in enumerate(C['cuentas']):
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', ent)
        s.txt(r, 'C', tipo, sz=9.0, color=TEXT2, al='left', v='center')
        s.mark_input(r, 'D', 'E'); s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.put(r, 'D', None, F(9.0, color=TEXT), AL('center', 'center'), '@', 'E')
        s.money(r, 'F', val); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(N(F{r})+N(G{r}),0)')
    cb1 = s.r - 1
    ws[f'D{cb0}'].comment = Comment(
        'NOTA 2 — Espacio en blanco para el número completo de la cuenta\n'
        'tal como aparece en el extracto (la exógena informa el número interno).', AUT, 260, 80)
    if T.get('nota_cuentas'):
        s.note(T['nota_cuentas'], h=29.25)
    sub['B'] = s.subtotal('Subtotal Cuentas Bancarias', f'=SUM(H{cb0}:H{cb1})')
    s.gap()

    # ══════════════ C · CDT  (NOTA 3) ══════════════
    s.section('C · CERTIFICADOS DE DEPÓSITO A TÉRMINO')
    s.head([('B', None, 'Entidad y cuenta', 'left'), ('C', None, 'Calidad del titular', 'left'),
            ('D', 'E', 'Porcentaje de participación', 'center'), ('F', None, 'Valor real', 'right'),
            ('G', None, 'Ajuste Fiscal', 'right'), ('H', None, 'Valor patrimonial', 'right')],
           h=24.0)
    cd0 = s.r
    for i, (ent, cal, pct, val, warn) in enumerate(C['cdt']):
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', ent)
        s.txt(r, 'C', cal, sz=9.0, b=warn, color=RED if warn else TEXT2, al='left', v='center')
        s.mark_input(r, 'D', 'E'); s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.pctc(r, 'D', pct, 'E', color=TEXT)
        s.money(r, 'F', val); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(N(F{r})*N(D{r})+N(G{r}),0)')
    if not C['cdt']:      # sin registros: fila abierta para diligenciar
        r = s.datarow()
        for cl in ('B', 'C', 'F', 'G'):
            s.mark_input(r, cl)
        s.mark_input(r, 'D', 'E')
        s.pctc(r, 'D', None, 'E', color=TEXT)
        s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(N(F{r})*N(D{r})+N(G{r}),0)')
    cd1 = s.r - 1
    if C['cdt']:
        ws[f'D{cd0}'].comment = Comment(
            'NOTA 3 — Cambie el porcentaje por el de propiedad efectiva y el\n'
            'valor patrimonial se recalcula solo:\n'
            'Valor real × Porcentaje + Ajuste fiscal.', AUT, 300, 90)
    if T.get('nota_cdt'):
        s.note(T['nota_cdt'], h=29.25)
    sub['C'] = s.subtotal('Subtotal CDT', f'=SUM(H{cd0}:H{cd1})')
    s.gap()

    # ══════════════ D · INVERSIONES  (NOTA 4) ══════════════
    s.section('D · INVERSIONES, APORTES Y DERECHOS SOCIALES')
    s.head([('B', 'C', 'Entidad', 'left'), ('D', 'E', 'Participación', 'center'),
            ('F', None, 'Costo Fiscal', 'right'), ('G', None, 'Ajuste Fiscal', 'right'),
            ('H', None, 'Valor patrimonial', 'right')])
    iv0 = s.r
    for i, (ent, pct, muisca) in enumerate(C['inversiones']):
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', ent, 'C')
        s.mark_input(r, 'D', 'E'); s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.pctc(r, 'D', pct, 'E', color=TEXT)
        s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=IF(N(F{r})+N(G{r})=0,{muisca},ROUND(N(F{r})+N(G{r}),0))')
    iv1 = s.r - 1
    if C['inversiones']:
        ws[f'F{iv0}'].comment = Comment(
            'NOTA 4 — El costo fiscal se deja en blanco: diligéncielo cuando tenga\n'
            'el certificado de la sociedad. Mientras esté vacío, el valor patrimonial\n'
            'muestra el declarado en el MUISCA; al diligenciarlo prevalece\n'
            'Costo fiscal + Ajuste fiscal.', AUT, 300, 100)
    if T.get('nota_inversiones'):
        s.note(T['nota_inversiones'], h=29.25)
    sub['D'] = s.subtotal('Subtotal Inversiones', f'=SUM(H{iv0}:H{iv1})')
    s.gap()

    # ══════════════ E · CUENTAS POR COBRAR  (NOTA 5) ══════════════
    s.section('E · CUENTAS POR COBRAR')
    s.head([('B', 'C', 'Deudor', 'left'), ('D', 'G', 'Concepto', 'left'),
            ('H', None, 'Valor', 'right')])
    cc0 = s.r
    n = len(C['cuentas_cobrar'])
    for i, (deu, con, val) in enumerate(C['cuentas_cobrar']):
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', deu, 'C')
        s.put(r, 'D', con, F(9.2, True, color=GOLD), AL('left', 'center', True, 1), None, 'G')
        s.money(r, 'H', val)
    for k in range(C.get('cuentas_cobrar_filas_libres', 0)):
        r = s.datarow(zebra=bool((n + k) % 2))
        s.mark_input(r, 'B', 'C'); s.mark_input(r, 'D', 'G'); s.mark_input(r, 'H')
        s.txt(r, 'B', None, 'C')
        s.put(r, 'D', None, F(9.2, True, color=GOLD), AL('left', 'center', True, 1), None, 'G')
        s.money(r, 'H', None)
    cc1 = s.r - 1
    if T.get('nota_cxc'):
        s.note(T['nota_cxc'], h=29.25)
    sub['E'] = s.subtotal('Subtotal Cuentas por Cobrar', f'=SUM(H{cc0}:H{cc1})')
    s.gap()

    # ══════════════ F · CESANTÍAS  (NOTA 6) ══════════════
    s.section('F · CESANTÍAS EN FONDO')
    s.head([('B', 'C', 'Entidad', 'left'), ('D', 'E', 'Fondo', 'left'),
            ('F', None, 'Saldo Año Anterior', 'right'), ('G', None, 'Consignadas', 'right'),
            ('H', None, 'Valor patrimonial', 'right')], h=24.0)
    ce0 = s.r
    for i, (ent, fondo, val) in enumerate(C['cesantias']):
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', ent, 'C')
        s.put(r, 'D', fondo, F(9.2, True, color=GOLD), AL('left', 'center', False, 1), None, 'E')
        s.mark_input(r, 'F')
        s.money(r, 'F', None); s.money(r, 'G', val)
        s.money(r, 'H', f'=ROUND(N(F{r})+N(G{r}),0)')
    if not C['cesantias']:
        r = s.datarow()
        s.mark_input(r, 'B', 'C'); s.mark_input(r, 'D', 'E')
        s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.txt(r, 'B', None, 'C')
        s.put(r, 'D', None, F(9.2, True, color=GOLD), AL('left', 'center', False, 1), None, 'E')
        s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(N(F{r})+N(G{r}),0)')
    ce1 = s.r - 1
    if C['cesantias']:
        ws[f'F{ce0}'].comment = Comment(
            'NOTA 6 — Diligencie el saldo reportado en el año anterior según el\n'
            'certificado del fondo. El valor patrimonial es la suma de los dos\n'
            'conceptos: saldo anterior + consignadas en el período.', AUT, 280, 90)
    if T.get('nota_cesantias'):
        s.note(T['nota_cesantias'], h=29.25)
    sub['F'] = s.subtotal('Subtotal Cesantías', f'=SUM(H{ce0}:H{ce1})')
    s.gap()

    # ══════════════ G · INMUEBLES  (NOTA 7) ══════════════
    s.section('G · BIENES INMUEBLES')
    s.head([('B', None, 'Matrícula', 'center'), ('C', None, 'Participación', 'center'),
            ('D', None, 'Dirección', 'left'), ('E', None, 'Saldo Año Anterior', 'right'),
            ('F', None, 'Ajuste Fiscal', 'right'), ('G', None, 'Avalúo catastral', 'right'),
            ('H', None, 'Valor patrimonial', 'right')], h=26.0)
    im0 = s.r
    n = len(C['inmuebles'])
    for i, (mat, pct, av) in enumerate(C['inmuebles']):
        r = s.datarow(zebra=bool(i % 2))
        s.put(r, 'B', mat, F(9.2, color=TEXT), AL('center', 'center'), '@')
        s.mark_input(r, 'C'); s.mark_input(r, 'D'); s.mark_input(r, 'E'); s.mark_input(r, 'F')
        s.pctc(r, 'C', pct, color=TEXT)
        s.put(r, 'D', None, F(9.0, color=TEXT), AL('left', 'center', True, 1))
        s.money(r, 'E', None); s.money(r, 'F', None); s.money(r, 'G', av)
        s.money(r, 'H', f'=ROUND(MAX(N(E{r})+N(F{r}),N(G{r})*N(C{r})),0)')
    for k in range(C.get('inmuebles_filas_libres', 0)):
        r = s.datarow(zebra=bool((n + k) % 2))
        for cl in ('B', 'C', 'D', 'E', 'F', 'G'):
            s.mark_input(r, cl)
        s.put(r, 'B', None, F(9.2, color=TEXT), AL('center', 'center'), '@')
        s.pctc(r, 'C', 1.0, color=TEXT)
        s.money(r, 'E', None); s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(MAX(N(E{r})+N(F{r}),N(G{r})*N(C{r})),0)')
    im1 = s.r - 1
    if C['inmuebles']:
        ws[f'C{im0}'].comment = Comment(
            'NOTA 7 — El avalúo del concepto 1476 es el del PREDIO COMPLETO.\n'
            'Cuando hay copropiedad, al contribuyente solo le corresponde este\n'
            'porcentaje, y el valor patrimonial lo aplica automáticamente:\n'
            '  MAX( saldo año anterior + ajuste fiscal ;  avalúo × participación )',
            AUT, 330, 100)
    s.note('NOTA 7 — Matrícula, participación, dirección, saldo del año anterior y ajuste fiscal son editables; el avalúo catastral viene de la exógena (concepto 1476) y corresponde al predio COMPLETO. Valor patrimonial = mayor entre (saldo año anterior + ajuste fiscal) y (avalúo × porcentaje de participación), conforme al art. 277 E.T. En predios de un solo dueño la participación es 100% y no cambia nada; en copropiedades —herencias, sucesiones, edificios— es lo que evita declarar el inmueble entero. Las filas libres quedan abiertas para inmuebles en municipios que no reportan.', h=48.0)
    sub['G'] = s.subtotal('Subtotal Inmuebles', f'=SUM(H{im0}:H{im1})')
    s.gap()

    # ══════════════ H · VEHÍCULOS  (NOTA 8) ══════════════
    s.section('H · VEHÍCULOS')
    s.head([('B', None, 'Placa', 'center'), ('C', 'D', 'Reportante', 'left'),
            ('E', None, 'Saldo Año Anterior', 'right'), ('F', None, 'Ajuste Fiscal', 'right'),
            ('G', None, 'Avalúo', 'right'), ('H', None, 'Valor patrimonial', 'right')], h=24.0)
    ve0 = s.r
    for i, (pl, rep, av) in enumerate(C['vehiculos']):
        r = s.datarow(zebra=bool(i % 2))
        s.put(r, 'B', pl, F(9.2, True, color=TEXT), AL('center', 'center'), '@')
        s.txt(r, 'C', rep, 'D', sz=9.0, color=TEXT2, v='center')
        s.mark_input(r, 'E'); s.mark_input(r, 'F')
        s.money(r, 'E', av); s.money(r, 'F', None); s.money(r, 'G', av)
        s.money(r, 'H', f'=ROUND(N(E{r})+N(F{r}),0)')
    if not C['vehiculos']:
        r = s.datarow()
        s.mark_input(r, 'B'); s.mark_input(r, 'C', 'D')
        s.mark_input(r, 'E'); s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.put(r, 'B', None, F(9.2, True, color=TEXT), AL('center', 'center'), '@')
        s.txt(r, 'C', None, 'D')
        s.money(r, 'E', None); s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=ROUND(N(E{r})+N(F{r}),0)')
    ve1 = s.r - 1
    if C['vehiculos']:
        ws[f'E{ve0}'].comment = Comment(
            'NOTA 8 — Valor patrimonial = Saldo año anterior + Ajuste fiscal.\n'
            'Art. 267 E.T.: el vehículo se declara por su costo fiscal, no por el\n'
            'avalúo. La casilla viene prediligenciada con el avalúo del MUISCA\n'
            'para que el libro cuadre; reemplácela por el costo fiscal declarado\n'
            'en el año anterior.', AUT, 320, 110)
    s.note(f'NOTA 8 — Placa y entidad donde está inscrito el vehículo, saldo del año anterior y ajuste fiscal editables, y avalúo reportado por el MUISCA (concepto 1480). Valor patrimonial = saldo del año anterior + ajuste fiscal. El saldo del año anterior viene prediligenciado con el avalúo informado: reemplácelo por el costo fiscal declarado en el AG {ant} (art. 267 E.T.).', h=29.25)
    sub['H'] = s.subtotal('Subtotal Vehículos', f'=SUM(H{ve0}:H{ve1})')
    s.gap()

    # ══════════════ I · PASIVOS  (NOTA 9) ══════════════
    s.section('I · PASIVOS', accent=RED)
    s.head([('B', 'C', 'Acreedor', 'left'), ('D', 'E', 'Concepto', 'left'),
            ('F', None, 'Certificado', 'right'), ('G', None, 'Ajustes Fiscales', 'right'),
            ('H', None, 'Saldo', 'right')], accent=RED)
    pa0 = s.r
    for i, (acr, con, val) in enumerate(C['pasivos']):
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', acr, 'C')
        s.txt(r, 'D', con, 'E', sz=9.0, color=TEXT2, v='center')
        s.mark_input(r, 'F'); s.mark_input(r, 'G')
        s.money(r, 'F', None); s.money(r, 'G', None)
        s.money(r, 'H', f'=IF(N(F{r})+N(G{r})=0,{val},ROUND(N(F{r})+N(G{r}),0))')
    pa1 = s.r - 1
    if C['pasivos']:
        ws[f'F{pa0}'].comment = Comment(
            'NOTA 9 — Espacio editable para la información de lo certificado.\n'
            'Mientras esté vacío, el saldo muestra lo reportado en la exógena;\n'
            'al diligenciarlo prevalece Certificado + Ajustes fiscales.', AUT, 300, 90)
    if T.get('nota_pasivos'):
        s.note(T['nota_pasivos'], h=29.25)
    A['pas_tot'] = s.subtotal('Total deudas (R30)', f'=SUM(H{pa0}:H{pa1})', accent=RED, tint=RED_T)
    s.gap(); s.gap()

    # ══════════════ CONSOLIDADO ══════════════
    s.section('CONSOLIDADO PATRIMONIAL')
    s.head([('B', 'G', 'Concepto', 'left'), ('H', None, 'Valor', 'right')])
    grupos = [('A', 'Efectivo'), ('B', 'Saldos en cuentas bancarias'),
              ('C', 'Certificados de depósito a término'), ('D', 'Inversiones, aportes y derechos sociales'),
              ('E', 'Cuentas por cobrar'), ('F', 'Cesantías en fondo'),
              ('G', 'Bienes inmuebles'), ('H', 'Vehículos')]
    bruto_r = s.datarow()
    s.txt(bruto_r, 'B', 'R29 · Patrimonio bruto', 'G', b=True)
    s.money(bruto_r, 'H', '=SUM(' + ','.join(f'H{sub[k]}' for k, _ in grupos) + ')', b=True)
    A['pat_bruto'] = bruto_r
    for k, nom in grupos:
        r = s.datarow(zebra=True)
        s.txt(r, 'B', f'        {k} · {nom}', 'G', sz=8.9, color=TEXT2)
        s.money(r, 'H', f'=H{sub[k]}')
    r = s.datarow()
    s.txt(r, 'B', 'R30 · Menos: deudas', 'G', b=True)
    s.money(r, 'H', f'=-H{A["pas_tot"]}', b=True)
    A['pat_liq'] = s.hero('R31 · PATRIMONIO LÍQUIDO', f'=H{bruto_r}-H{A["pas_tot"]}')
    s.gap(); s.gap()

    # ══════════════ CONCILIACIÓN PATRIMONIAL  (NOTA 10) ══════════════
    s.hero('CONCILIACIÓN PATRIMONIAL', None, val_col='H', h=26.0)
    s.gap(5.0)
    # La comparación patrimonial del art. 236 E.T. se hace sobre patrimonio
    # LÍQUIDO, no bruto: lo que debe justificarse es el incremento del neto.
    r = s.datarow(h=20.5)
    s.txt(r, 'B', f'Patrimonio Líquido Declarado {ag}', 'G', sz=10.0, v='center')
    s.money(r, 'H', f'=H{A["pat_liq"]}', sz=10.0)
    liq_ag = r
    # El patrimonio líquido del año anterior se toma DIRECTO de la casilla 31
    # del formulario 210 presentado. Antes se reconstruía restándole al bruto
    # unas deudas que había que teclear aparte: dos casillas para llegar a un
    # dato que la declaración del año anterior ya trae calculado y firmado.
    r = s.datarow(zebra=True, h=20.5)
    s.txt(r, 'B', f'Patrimonio Líquido Declarado {ant}  ·  casilla 31 del Formulario 210',
          'G', sz=10.0, b=True, v='center')
    s.mark_input(r, 'H')
    s.money(r, 'H', None, sz=10.0, b=True)
    liq_ant = r
    A['pat_liq_anterior'] = liq_ant
    ws[f'H{liq_ant}'].comment = Comment(
        f'NOTA 10 — Casilla 31 del formulario 210 del año gravable {ant}, tal\n'
        f'como quedó presentado. Es el patrimonio LÍQUIDO, ya con las deudas\n'
        f'restadas: no hay que volver a restarlas aquí.\n'
        f'La exógena informa otra cosa —el patrimonio BRUTO del año anterior,\n'
        f'{fmt_pesos(C["patrimonio_bruto_anterior"])}— que NO es esta casilla y por eso no se\n'
        f'prediligencia: tómela de la declaración, no del reporte.', AUT, 350, 120)
    variacion = s.subtotal(f'(=) Variación patrimonial del año  ·  {ant} → {ag}',
                           f'=H{liq_ag}-H{liq_ant}', lab_col='B', val_col='H')
    s.gap()

    # ── Con qué se justifica esa variación ──────────────────────────────
    # El patrimonio solo puede haber crecido con lo que entró y no se gastó en
    # impuestos: la renta gravable del año, más lo que entró sin ser renta
    # (los INCRNGO), menos el impuesto que se pagó. Lo que no cuadre con eso es
    # lo que la DIAN pregunta.
    #
    # Las tres cifras vienen por NOMBRE DEFINIDO y no por referencia de celda:
    # esta hoja se arma antes que la de liquidación y la de ingresos, así que
    # sus filas todavía no existen. Los nombres se resuelven al abrir el libro
    # (se declaran al final de generar.py) y además sobreviven a que cualquiera
    # de esas hojas cambie de tamaño.
    s.section('JUSTIFICACIÓN DE LA VARIACIÓN PATRIMONIAL')
    s.head([('B', 'G', 'Concepto', 'left'), ('H', None, 'Valor', 'right')])
    r = s.datarow(h=20.5)
    s.txt(r, 'B', f'(+) Renta líquida gravable {ag}', 'G', sz=10.0, v='center')
    s.money(r, 'H', '=RENTA_LIQ_GRAVABLE', sz=10.0)
    j_rlg = r
    r = s.datarow(zebra=True, h=20.5)
    s.txt(r, 'B', '(+) Ingresos no constitutivos de renta  ·  R33 + R59 + no laborales',
          'G', sz=10.0, v='center')
    s.money(r, 'H', '=INCRNGO_TRABAJO+INCRNGO_CAPITAL+INCRNGO_NO_LABORAL', sz=10.0)
    j_incr = r
    ws[f'H{j_incr}'].comment = Comment(
        'NOTA 10 — Entraron al patrimonio pero no pagaron impuesto: aportes\n'
        'obligatorios a salud y pensión (R33), componente inflacionario de los\n'
        'rendimientos financieros (R59) y los no constitutivos de las rentas no\n'
        'laborales. Se toman de la hoja 3 y se actualizan solos.', AUT, 330, 90)
    r = s.datarow(h=20.5)
    s.txt(r, 'B', f'(−) Impuesto de renta pagado durante {ag}', 'G', sz=10.0, v='center')
    s.mark_input(r, 'H')
    s.money(r, 'H', None, sz=10.0)
    j_imp = r
    ws[f'H{j_imp}'].comment = Comment(
        f'NOTA 10 — Lo efectivamente pagado por impuesto de renta durante {ag}:\n'
        f'el saldo a pagar de la declaración del año anterior más los anticipos.\n'
        f'Salió del patrimonio sin ser un gasto de renta, así que justifica que\n'
        f'el patrimonio no haya crecido tanto. Escríbalo en positivo: la fórmula\n'
        f'ya lo resta.', AUT, 340, 100)
    r = s.datarow(zebra=True, h=20.5)
    s.txt(r, 'B', '(±) Ajuste de capital', 'G', sz=10.0, v='center')
    s.mark_input(r, 'H')
    s.money(r, 'H', None, sz=10.0)
    j_aj = r
    A['ajuste_capital'] = j_aj
    ws[f'H{j_aj}'].comment = Comment(
        'NOTA 10 — Casilla abierta para todo lo demás que mueve el patrimonio\n'
        'sin pasar por la renta del año: rentas exentas, ganancias ocasionales,\n'
        'herencias y donaciones recibidas, gastos personales de subsistencia\n'
        '(en negativo), valorizaciones y correcciones de avalúo. Escriba el\n'
        'neto, positivo o negativo, y deje la explicación por escrito: es\n'
        'exactamente lo que se pregunta en un requerimiento.', AUT, 350, 120)
    justificada = s.subtotal('(=) Variación patrimonial justificada',
                             f'=H{j_rlg}+H{j_incr}-N(H{j_imp})+N(H{j_aj})',
                             lab_col='B', val_col='H')
    s.gap()

    r = s.datarow(h=22.0)
    s.txt(r, 'B', 'Diferencia Patrimonial Sin Justificar', 'G', sz=10.5, b=True, color=RED, v='center')
    s.money(r, 'H', f'=H{variacion}-H{justificada}', sz=11.5, b=True, color=RED, fmt=NUM)
    A['pat_dif'] = r
    s.note(f'NOTA 10 — La comparación patrimonial del art. 236 E.T. se hace sobre el patrimonio líquido: enfrenta el reconstruido para {ag} —bruto menos deudas— contra el que quedó declarado en la casilla 31 del formulario 210 de {ant}. Esa casilla se digita de la declaración presentada y no de la exógena, que solo informa el patrimonio bruto. Lo que la variación no explique con la renta gravable, los ingresos no constitutivos y el impuesto pagado queda aquí, en rojo: es renta líquida gravable por comparación patrimonial si no se justifica con la casilla de ajuste de capital.', h=38.0)
    s.gap()
    if T.get('callout_brecha'):
        titulo, cuerpo = T['callout_brecha']
        s.callout(titulo, cuerpo, accent=RED, tint=RED_T, h=48.0)

    ws.freeze_panes = 'A6'
    return ws
