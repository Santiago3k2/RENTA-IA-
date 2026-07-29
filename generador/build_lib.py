# -*- coding: utf-8 -*-
"""Sistema de diseño compartido para la reconstrucción del libro."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ---------- paleta (extraída del libro original) ----------
INK      = 'FF12262B'
TEXT     = 'FF12262B'
TEXT2    = 'FF3D5459'
MUTED    = 'FF6A8085'
ONDARK   = 'FFC3D6D3'
ONDARK2  = 'FF8FA9A6'
WHITE    = 'FFFFFFFF'
ZEBRA    = 'FFF2F6F5'
BORD     = 'FFB3C3C0'
BORD_IN  = 'FFD5DFDD'

TEAL, TEAL_T, TEAL_ACC = 'FF0D6E64', 'FFE3EFED', 'FF48B9AA'
BLUE, BLUE_T           = 'FF1D5A8C', 'FFE7EFF6'
PURPLE, PURPLE_T       = 'FF74386A', 'FFF2EAF0'
GOLD, GOLD_T           = 'FF8D6209', 'FFF6EFE0'
RED, RED_T             = 'FFA32C33', 'FFF8EAEA'
GREEN                  = 'FF1C6F43'
GREY                   = 'FF6A8085'

# celdas editables
IN_FILL  = 'FFFDF6E7'
IN_BORD  = 'FFD9BE86'

# ---------- formatos numéricos ----------
NUM   = r'#,##0;[Red]\(#,##0\)'
ACC   = r'_-"$"\ * #,##0_-;[Red]\-"$"\ * #,##0_-;_-"$"\ * "-"_-;_-@_-'
MONEY = r'"$"\ #,##0;[Red]\("$"\ #,##0\)'
PCT   = r'0.00%'
UVTF  = r'#,##0.00'

SUI, GEO = 'Segoe UI', 'Georgia'

# ---------- identificación que se repite en todas las hojas ----------
IDENT_LINE = ''


def set_identidad(nombre, identificacion, ano_gravable):
    """Fija la franja de identidad ANTES de construir cualquier hoja."""
    global IDENT_LINE
    IDENT_LINE = (f'CONTRIBUYENTE:  {nombre}          ·          '
                  f'IDENTIFICACIÓN:  {identificacion}          ·          '
                  f'AÑO GRAVABLE:  {ano_gravable}')


def fnum(v):
    """1234567 -> '1.234.567' (separador de miles colombiano)."""
    return f'{int(round(v)):,}'.replace(',', '.')


def fmt_pesos(v):
    return '$ ' + fnum(v)


def S(c):  # side
    return Side(style='thin', color=c)


def F(sz, b=False, i=False, color=TEXT, name=SUI):
    return Font(name=name, size=sz, bold=b, italic=i, color=color)


def AL(h='left', v='center', wrap=False, ind=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=ind)


def FILL(c):
    return PatternFill('solid', fgColor=c)


class Sheet:
    """Envoltura con el lenguaje visual del libro."""

    def __init__(self, ws, first='B', last='H', accent=TEAL, tint=TEAL_T):
        self.ws = ws
        self.f = column_index_from_string(first)
        self.l = column_index_from_string(last)
        self.accent = accent
        self.tint = tint
        self.r = 1
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        ws.page_setup.orientation = 'landscape'
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    # ---------- utilidades ----------
    def col(self, letter):
        return column_index_from_string(letter)

    def span(self, r, c1, c2):
        if c1 != c2:
            self.ws.merge_cells(start_row=r, start_column=self.col(c1),
                                end_row=r, end_column=self.col(c2))

    def frame(self, r, fill=WHITE, bottom=BORD_IN, top=None, outer=BORD):
        """Pinta el marco de una fila completa de tabla."""
        for ci in range(self.f, self.l + 1):
            c = self.ws.cell(row=r, column=ci)
            c.fill = FILL(fill)
            c.border = Border(
                left=S(outer) if ci == self.f else None,
                right=S(outer) if ci == self.l else None,
                top=S(top) if top else None,
                bottom=S(bottom) if bottom else None,
            )

    def put(self, r, colletter, value, font=None, align=None, fmt=None, span_to=None):
        c = self.ws.cell(row=r, column=self.col(colletter))
        c.value = value
        if font:
            c.font = font
        if align:
            c.alignment = align
        if fmt:
            c.number_format = fmt
        if span_to:
            self.span(r, colletter, span_to)
        return c

    def mark_input(self, r, colletter, span_to=None):
        c1 = self.col(colletter)
        c2 = self.col(span_to) if span_to else c1
        for ci in range(c1, c2 + 1):
            c = self.ws.cell(row=r, column=ci)
            c.fill = FILL(IN_FILL)
            c.border = Border(
                left=S(IN_BORD) if ci == c1 else S(IN_BORD),
                right=S(IN_BORD) if ci == c2 else S(IN_BORD),
                top=S(IN_BORD), bottom=S(IN_BORD))

    # ---------- bloques ----------
    def band(self, eyebrow, title, desc, h_desc=35.55):
        ws, r = self.ws, self.r
        ws.row_dimensions[1].height = 6.0
        for rr, (val, fnt, al, hh) in enumerate([
            (eyebrow, F(7.8, True, color=ONDARK), AL('left', 'center'), 16.95),
            (title,   F(18, True, color=WHITE, name=GEO), AL('left', 'center'), 30.0),
            (desc,    F(8.8, color=ONDARK2), AL('left', 'top', True), h_desc),
        ], start=2):
            for ci in range(self.f, self.l + 1):
                ws.cell(row=rr, column=ci).fill = FILL(INK)
            c = ws.cell(row=rr, column=self.f)
            c.value, c.font, c.alignment = val, fnt, al
            ws.row_dimensions[rr].height = hh
            self.span(rr, get_column_letter(self.f), get_column_letter(self.l))
        ident_strip(self, 5)
        ws.row_dimensions[6].height = 7.0
        self.r = 7
        return self

    def legend(self, text):
        """Franja de muestra: su propio fondo es el de las casillas editables."""
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = 22.0
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = FILL(IN_FILL)
            c.border = Border(top=S(IN_BORD), bottom=S(IN_BORD),
                              left=S(IN_BORD) if ci == self.f else None,
                              right=S(IN_BORD) if ci == self.l else None)
        c = ws.cell(row=r, column=self.f)
        c.value = '▨   ' + text
        c.font = F(8.2, False, True, 'FF8A6D22')
        c.alignment = AL('left', 'center', True, 1)
        self.span(r, get_column_letter(self.f), get_column_letter(self.l))
        self.r = r + 2
        ws.row_dimensions[r + 1].height = 7.0
        return self

    def section(self, title, accent=None):
        acc = accent or self.accent
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = 18.0
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r, column=ci)
            c.border = Border(bottom=S(acc))
        c = ws.cell(row=r, column=self.f)
        c.value = '  ' + title
        c.font = F(8.8, True, color=acc)
        c.alignment = AL('left')
        self.r = r + 1
        return self

    def head(self, cols, accent=None, h=19.95):
        """cols = [(colletter, span_to|None, label, align)]"""
        acc = accent or self.accent
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = h
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = FILL(acc)
            c.border = Border(top=S(BORD),
                              left=S(BORD) if ci == self.f else None,
                              right=S(BORD) if ci == self.l else None)
        for cl, sp, label, al in cols:
            c = ws.cell(row=r, column=self.col(cl))
            c.value = label
            c.font = F(8.3, True, color=WHITE)
            c.alignment = AL(al, 'center', True, 1)
            if sp:
                self.span(r, cl, sp)
        self.r = r + 1
        return self

    def datarow(self, zebra=False, h=19.35):
        r = self.r
        self.ws.row_dimensions[r].height = h
        self.frame(r, ZEBRA if zebra else WHITE)
        self.r = r + 1
        return r

    def txt(self, r, cl, val, sp=None, sz=9.2, b=False, i=False, color=TEXT, al='left', v='top'):
        return self.put(r, cl, val, F(sz, b, i, color), AL(al, v, True, 1), None, sp)

    def money(self, r, cl, val, sp=None, sz=9.2, b=False, color=TEXT, fmt=ACC):
        return self.put(r, cl, val, F(sz, b, color=color), AL('right', 'center', False, 1), fmt, sp)

    def pctc(self, r, cl, val, sp=None, color=TEXT2):
        return self.put(r, cl, val, F(9.2, color=color), AL('center', 'center'), PCT, sp)

    def note(self, text, h=17.55, tint=None):
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = h
        self.frame(r, tint or WHITE)
        c = ws.cell(row=r, column=self.f)
        c.value = text
        c.font = F(8.0, False, True, MUTED)
        c.alignment = AL('left', 'top', True, 2)
        self.span(r, get_column_letter(self.f), get_column_letter(self.l))
        self.r = r + 1
        return r

    def subtotal(self, label, formula, accent=None, tint=None, lab_col=None, val_col=None,
                 sz=10.0, h=19.05):
        acc = accent or self.accent
        tnt = tint or self.tint
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = h
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = FILL(tnt)
            c.border = Border(top=S(acc), bottom=S(BORD),
                              left=S(BORD) if ci == self.f else None,
                              right=S(BORD) if ci == self.l else None)
        lc = lab_col or get_column_letter(self.f)
        vc = val_col or get_column_letter(self.l)
        c = ws.cell(row=r, column=self.col(lc))
        c.value, c.font, c.alignment = label, F(9.6, True, color=TEXT), AL('left', 'center', False, 1)
        self.span(r, lc, get_column_letter(self.col(vc) - 1))
        v = ws.cell(row=r, column=self.col(vc))
        v.value, v.font = formula, F(sz, True, color=acc)
        v.alignment, v.number_format = AL('right', 'center', False, 1), ACC
        self.r = r + 1
        return r

    def hero(self, label, formula, accent=None, val_col=None, h=24.0):
        acc = accent or self.accent
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = h
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = FILL(acc)
            c.border = Border(left=S(BORD) if ci == self.f else None,
                              right=S(BORD) if ci == self.l else None)
        c = ws.cell(row=r, column=self.f)
        c.value, c.font = label, F(11.5, True, color=WHITE, name=GEO)
        c.alignment = AL('left', 'center', False, 1)
        vc = val_col or get_column_letter(self.l)
        self.span(r, get_column_letter(self.f), get_column_letter(self.col(vc) - 1))
        v = ws.cell(row=r, column=self.col(vc))
        v.value, v.font = formula, F(12.0, True, color=WHITE)
        v.alignment, v.number_format = AL('right', 'center', False, 1), MONEY
        self.r = r + 1
        return r

    def callout(self, title, body, accent=None, tint=None, h=35.25):
        acc = accent or self.accent
        tnt = tint or self.tint
        ws, r = self.ws, self.r
        ws.row_dimensions[r].height = 16.05
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = FILL(tnt)
            c.border = Border(top=S(acc),
                              left=S(acc) if ci == self.f else None,
                              right=S(acc) if ci == self.l else None)
        c = ws.cell(row=r, column=self.f)
        c.value, c.font = title, F(8.4, True, color=acc)
        c.alignment = AL('left', 'center', False, 1)
        self.span(r, get_column_letter(self.f), get_column_letter(self.l))
        r2 = r + 1
        ws.row_dimensions[r2].height = h
        for ci in range(self.f, self.l + 1):
            c = ws.cell(row=r2, column=ci)
            c.fill = FILL(tnt)
            c.border = Border(bottom=S(acc),
                              left=S(acc) if ci == self.f else None,
                              right=S(acc) if ci == self.l else None)
        c = ws.cell(row=r2, column=self.f)
        c.value, c.font = body, F(8.7, color=TEXT2)
        c.alignment = AL('left', 'top', True, 1)
        self.span(r2, get_column_letter(self.f), get_column_letter(self.l))
        self.r = r2 + 1
        return r

    def gap(self, h=7.0):
        self.ws.row_dimensions[self.r].height = h
        self.r += 1
        return self

    def widths(self, mapping):
        for k, v in mapping.items():
            self.ws.column_dimensions[k].width = v
        return self


def ident_strip(sheet_or_ws, row, first=None, last=None, accent=None):
    """Franja fija con contribuyente, identificación y año gravable."""
    if isinstance(sheet_or_ws, Sheet):
        ws, f, l = sheet_or_ws.ws, sheet_or_ws.f, sheet_or_ws.l
        acc = accent or sheet_or_ws.accent
    else:
        ws = sheet_or_ws
        f = column_index_from_string(first)
        l = column_index_from_string(last)
        acc = accent or TEAL
    ws.row_dimensions[row].height = 17.4
    for ci in range(f, l + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = FILL(ZEBRA)
        c.border = Border(top=S(BORD), bottom=S(BORD),
                          left=Side(style='medium', color=acc) if ci == f else None,
                          right=S(BORD) if ci == l else None)
    c = ws.cell(row=row, column=f)
    c.value = IDENT_LINE
    c.font = F(8.2, True, color=TEXT2)
    c.alignment = AL('left', 'center', False, 1)
    if l > f:
        ws.merge_cells(start_row=row, start_column=f, end_row=row, end_column=l)
    return row
