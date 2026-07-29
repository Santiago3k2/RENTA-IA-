# -*- coding: utf-8 -*-
"""Hoja 4 · Retenciones — nota 12, alimentada por datos.py."""
from openpyxl.comments import Comment
from build_lib import *

AUT = 'Generador renta'


def build(wb, A, C, T):
    ws = wb.create_sheet('4. Retenciones')
    ws.sheet_properties.tabColor = PURPLE[2:]
    s = Sheet(ws, 'B', 'G', PURPLE, PURPLE_T)
    s.widths({'A': 2.3, 'B': 27.0, 'C': 34.0, 'D': 16.0,
              'E': 17.0, 'F': 10.0, 'G': 17.0})
    ag, ant = C['ano_gravable'], C['ano_anterior']

    s.band('04  ·  RENGLÓN R132',
           'Retenciones en la fuente',
           T['sub_retenciones'], h_desc=38.0)
    s.legend('Las casillas con fondo crema son editables: diligencie la base y el porcentaje del certificado. Base × % debe reconciliar con la retención informada por el agente retenedor.')

    s.section('RETENCIONES INFORMADAS EN EXÓGENA')
    s.head([('B', None, 'Agente retenedor', 'left'), ('C', None, 'Concepto / cuenta', 'left'),
            ('D', None, 'Retención', 'right'), ('E', None, 'Base', 'right'),
            ('F', None, '%', 'center'), ('G', None, 'Certificado', 'right')])

    sub_rows = []
    grupos = C['retenciones']
    for gi, (concepto, filas) in enumerate(grupos):
        g0 = s.r
        for i, (ag_ret, con, val) in enumerate(filas):
            r = s.datarow(zebra=bool(i % 2))
            s.txt(r, 'B', ag_ret, v='center')
            s.txt(r, 'C', con, sz=9.0, color=TEXT2, v='center')
            s.money(r, 'D', val)
            s.mark_input(r, 'E'); s.mark_input(r, 'F'); s.mark_input(r, 'G')
            s.money(r, 'E', None)
            s.put(r, 'F', None, F(9.2, color=TEXT), AL('center', 'center'), PCT)
            s.money(r, 'G', val)
        g1 = s.r - 1
        rs_ = s.subtotal(f'Suma del concepto · {concepto}',
                         f'=SUM(D{g0}:D{g1})', lab_col='B', val_col='D', sz=9.4, h=18.0)
        s.money(rs_, 'E', f'=SUM(E{g0}:E{g1})', sz=9.4, b=True, color=PURPLE)
        s.money(rs_, 'G', f'=SUM(G{g0}:G{g1})', sz=9.4, b=True, color=PURPLE)
        sub_rows.append(rs_)
        if gi == 0 and filas:
            ws[f'E{g0}'].comment = Comment(
                'NOTA 12 — Casillas editables para la base y el porcentaje del\n'
                'certificado de retención. Base × % debe reconciliar con la\n'
                'retención informada en el prevalidador DIAN.', AUT, 300, 80)
        if gi < len(grupos) - 1:
            s.gap(4.0)

    if T.get('nota_retencion'):
        s.note(T['nota_retencion'])
    suma_d = '=' + '+'.join(f'D{x}' for x in sub_rows) if sub_rows else 0
    tot = s.hero(f'TOTAL RETENCIONES AG {ag} (R132)', suma_d, val_col='D')
    A['ret_total'] = tot
    if sub_rows:
        s.money(tot, 'E', '=' + '+'.join(f'E{x}' for x in sub_rows), sz=10.5, b=True, color=WHITE, fmt=ACC)
        s.money(tot, 'G', '=' + '+'.join(f'G{x}' for x in sub_rows), sz=10.5, b=True, color=WHITE, fmt=ACC)
    s.gap()

    # ══════════════ RETENCIONES ESPERADAS Y NO INFORMADAS ══════════════
    if C.get('retenciones_esperadas'):
        s.section('RETENCIONES ESPERADAS Y NO INFORMADAS')
        s.head([('B', None, 'Origen', 'left'), ('C', None, 'Base informada', 'right'),
                ('D', None, 'Tarifa', 'center'), ('E', 'F', 'Retención estimada', 'right'),
                ('G', None, 'Estado', 'center')])
        for i, (org, base, tar, est, estado) in enumerate(C['retenciones_esperadas']):
            r = s.datarow(zebra=bool(i % 2))
            s.txt(r, 'B', org, v='center')
            if base is None:
                s.put(r, 'C', 'sin ingreso', F(9.0, False, True, MUTED), AL('right', 'center', False, 1))
            else:
                s.money(r, 'C', base)
            s.put(r, 'D', tar, F(9.0, color=TEXT2), AL('center', 'center'))
            if est is None:
                s.put(r, 'E', 'por confirmar', F(9.0, False, True, MUTED), AL('right', 'center', False, 1), None, 'F')
            else:
                s.money(r, 'E', est, 'F')
            s.put(r, 'G', estado, F(7.8, True, color=RED), AL('center', 'center'))
        if T.get('nota_esperadas'):
            s.note(T['nota_esperadas'], h=29.25)
        s.gap()

    # ══════════════ OTROS CRÉDITOS ══════════════
    s.section('OTROS CRÉDITOS CONTRA EL IMPUESTO')
    s.head([('B', 'F', 'Concepto', 'left'), ('G', None, 'Valor', 'right')])
    r = s.datarow()
    s.txt(r, 'B', f'R131 · Saldo a favor del AG {ant} sin solicitud de devolución ni compensación', 'F')
    s.mark_input(r, 'G'); s.money(r, 'G', C['saldo_favor_anterior'])
    A['saldo_favor'] = r
    r = s.datarow(zebra=True)
    s.txt(r, 'B', f'Anticipo del AG {ag} liquidado en la declaración del AG {ant}', 'F')
    s.mark_input(r, 'G'); s.money(r, 'G', C['anticipo_previo'])
    A['ant_prev'] = r
    ws[f'G{r}'].comment = Comment(
        f'No aparece en la exógena: se obtiene de la declaración del AG {ant}.\n'
        'Cada peso liquidado allí reduce en igual medida el saldo a pagar\n'
        'de la hoja «5. Anticipo».', AUT, 300, 80)
    s.note('El anticipo liquidado en la declaración anterior no aparece en la exógena y solo se obtiene de esa declaración. Debe restarse en la liquidación final.')

    ws.freeze_panes = 'A6'
    return ws
