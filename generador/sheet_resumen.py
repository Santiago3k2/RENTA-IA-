# -*- coding: utf-8 -*-
"""Hoja 1 · Resumen — portada, KPIs, índice, obligación, perfil y plazo.

Se puebla al FINAL, cuando ya existen las anclas (A) de las demás hojas.
La celda G9 (UVT) es fija: todo el libro la referencia como '1. Resumen'!$G$9.
"""
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from build_lib import *

import plazos

AUT = 'Generador renta'

TEXTO_DOC_TRABAJO = ('Liquidación propuesta elaborada sobre información exógena; no es un formulario oficial '
                     'de la DIAN ni una declaración presentada. La exógena no reemplaza la realidad económica '
                     'del contribuyente: antes de diligenciar el formulario deben incorporarse los soportes '
                     'propios —las casillas de fondo crema de cada hoja están para eso— y resolverse las '
                     'partidas de la hoja «7. Alertas».')

PERFIL = [
    ('¿Es casado o vive en unión permanente?', 'Nombre e identificación del cónyuge',
     'El cónyuge en situación de dependencia cuenta como dependiente — art. 387 par. 2 E.T.'),
    ('¿Tiene hijos? ¿Cuántos?', 'Número de hijos', None),
    ('¿Tiene crédito hipotecario de vivienda?', 'Entidad e intereses pagados en el año',
     'Los intereses son deducibles hasta 1.200 UVT al año — art. 119 E.T.'),
    ('¿Tiene crédito educativo (ICETEX u otro)?', 'Entidad e intereses pagados en el año',
     'Los intereses de préstamos del ICETEX son deducibles hasta 100 UVT — art. 119 E.T.'),
    ('¿Tiene residencia fiscal en Colombia?', 'Días de permanencia en el país durante el año',
     'Define si tributa sobre renta de fuente mundial y presenta el formulario 210 — art. 10 E.T.'),
]


def populate(wb, A, C, T):
    ws = wb['1. Resumen']
    ws.sheet_properties.tabColor = TEAL[2:]
    s = Sheet(ws, 'B', 'I', TEAL, TEAL_T)
    s.widths({'A': 2.3, 'B': 20.0, 'C': 18.0, 'D': 14.0, 'E': 16.0,
              'F': 13.0, 'G': 13.0, 'H': 15.0, 'I': 18.0})
    ag, ant, sig = C['ano_gravable'], C['ano_anterior'], C['ano_siguiente']
    dig = C['ultimos_digitos']
    venc = plazos.vencimiento(ag, dig)      # None si no hay calendario para ese año
    PAT, ING = "'2. Patrimonio'", "'3. Ingresos'"
    RET, LIQ = "'4. Retenciones'", "'6. Liquidación'"
    ANT, CONS = "'5. Anticipo'", "'8. Consignaciones'"

    # ── banda de portada (4 filas, sin franja de identidad: aquí va el cuadro) ──
    ws.row_dimensions[1].height = 6.0
    filas = [
        (2, f'FORMULARIO 210  ·  PERSONAS NATURALES RESIDENTES  ·  AÑO GRAVABLE {ag}',
         F(7.8, True, color=ONDARK), AL('left', 'center'), 16.95),
        (3, 'Declaración de renta y complementarios',
         F(18, True, color=WHITE, name=GEO), AL('left', 'center'), 30.0),
        (4, C['nombre_titulo'], F(14, True, color=TEAL_ACC, name=GEO), AL('left', 'center'), 24.0),
        (5, T['sub_resumen'], F(8.8, color=ONDARK2), AL('left', 'top', True), 35.55),
    ]
    for rr, val, fnt, al, hh in filas:
        for ci in range(2, 10):
            ws.cell(row=rr, column=ci).fill = FILL(INK)
        c = ws.cell(row=rr, column=2)
        c.value, c.font, c.alignment = val, fnt, al
        ws.row_dimensions[rr].height = hh
        s.span(rr, 'B', 'I')
    ws.row_dimensions[6].height = 7.0
    ws.row_dimensions[7].height = 7.0

    # ── cuadro de identificación (la UVT queda SIEMPRE en G9) ──
    pares = [(2, 3, 'CONTRIBUYENTE', C['nombre']),
             (4, 5, 'IDENTIFICACIÓN', C['identificacion']),
             (6, 6, 'AÑO GRAVABLE', ag),
             (7, 7, f'UVT {ag}', C['uvt']),
             (8, 8, 'ÚLT. 2 DÍGITOS', dig),
             (9, 9, 'FUENTE', C['fuente'])]
    inicios = {p[0] for p in pares}
    for rr, hh in ((8, 13.05), (9, 25.95)):
        ws.row_dimensions[rr].height = hh
        for ci in range(2, 10):
            c = ws.cell(row=rr, column=ci)
            c.fill = FILL(ZEBRA)
            c.border = Border(top=S(BORD) if rr == 8 else None,
                              bottom=S(BORD) if rr == 9 else None,
                              left=S(BORD) if ci in inicios else None,
                              right=S(BORD) if ci == 9 else None)
    for c1, c2, lab, val in pares:
        c = ws.cell(row=8, column=c1)
        c.value, c.font = lab, F(7.2, True, color=MUTED)
        c.alignment = AL('left', 'center', False, 1)
        v = ws.cell(row=9, column=c1)
        v.value, v.font = val, F(9.4, True, color=TEXT)
        v.alignment = AL('left', 'top', True, 1)
        if c2 > c1:
            ws.merge_cells(start_row=8, start_column=c1, end_row=8, end_column=c2)
            ws.merge_cells(start_row=9, start_column=c1, end_row=9, end_column=c2)
    uvt_cell = ws['G9']
    uvt_cell.number_format = MONEY
    uvt_cell.fill = FILL(IN_FILL)
    uvt_cell.border = Border(left=S(IN_BORD), right=S(IN_BORD), top=S(IN_BORD), bottom=S(IN_BORD))
    uvt_cell.comment = Comment(
        'UVT del año gravable. Todas las hojas del libro toman el valor de\n'
        'esta celda: si cambia, el libro completo se recalcula.', AUT, 280, 60)
    ws.row_dimensions[10].height = 7.0
    ws.row_dimensions[11].height = 7.0

    # ── KPIs ──
    tarjetas = [
        (2, 3, TEAL, 'PATRIMONIO LÍQUIDO', f'={PAT}!H{A["pat_liq"]}',
         f'="Bruto "&TEXT({PAT}!H{A["pat_bruto"]},"#.##0")&" − deudas "&TEXT({PAT}!H{A["pas_tot"]},"#.##0")'),
        (4, 5, BLUE, 'INGRESOS BRUTOS', f'={ING}!G{A["ing_total"]}', T['kpi_ingresos_foot']),
        (6, 7, PURPLE, 'IMPUESTO A CARGO', f'={LIQ}!D{A["imp_cargo"]}', T['kpi_impuesto_foot']),
        (8, 9, GOLD, f'ANTICIPO AG {sig}', f'={ANT}!I{A["ant_row"]}',
         f'Método 2 · sujeto al impuesto neto AG {ant}'),
    ]
    for rr, hh in ((12, 13.05), (13, 24.0), (14, 24.0)):
        ws.row_dimensions[rr].height = hh
        for c1, c2, acc, _, _, _ in tarjetas:
            for ci in range(c1, c2 + 1):
                c = ws.cell(row=rr, column=ci)
                c.fill = FILL(WHITE)
                c.border = Border(top=S(BORD) if rr == 12 else None,
                                  bottom=S(BORD) if rr == 14 else None,
                                  left=S(BORD) if ci == c1 else None,
                                  right=S(BORD) if ci == c2 else None)
    for c1, c2, acc, lab, val, foot in tarjetas:
        c = ws.cell(row=12, column=c1)
        c.value, c.font, c.alignment = lab, F(7.4, True, color=acc), AL('left', 'center', False, 1)
        v = ws.cell(row=13, column=c1)
        v.value, v.font = val, F(14.5, True, color=acc)
        v.alignment, v.number_format = AL('left', 'center', False, 1), MONEY
        p = ws.cell(row=14, column=c1)
        p.value, p.font, p.alignment = foot, F(7.6, color=MUTED), AL('left', 'top', True, 1)
        for rr in (12, 13, 14):
            if c2 > c1:
                ws.merge_cells(start_row=rr, start_column=c1, end_row=rr, end_column=c2)
    ws.row_dimensions[15].height = 7.0

    # ── índice del libro ──
    s.r = 16
    s.section('CONTENIDO DEL LIBRO')
    s.head([('B', 'C', 'Hoja', 'left'), ('D', 'G', 'Qué contiene', 'left'),
            ('H', 'I', 'Cifra clave', 'right')])
    n_altas = sum(1 for a in C['alertas'] if a[1] == 'ALTA')
    indice = [
        ('2 · Patrimonio', f'Ocho grupos de activos, pasivos y conciliación patrimonial a 31-dic-{ag}',
         f'={PAT}!H{A["pat_liq"]}', TEAL, 'num'),
        ('3 · Ingresos', 'Rentas por subcédula, reportado frente a certificado y tope de costos del 60%',
         f'={ING}!G{A["ing_total"]}', BLUE, 'num'),
        ('4 · Retenciones', 'Certificados agrupados por concepto y créditos contra el impuesto',
         f'={RET}!D{A["ret_total"]}', PURPLE, 'num'),
        ('5 · Anticipo', 'Métodos del art. 807 E.T., sensibilidad y saldo a pagar',
         f'={ANT}!I{A["ant_row"]}', GOLD, 'num'),
        ('6 · Liquidación', 'Depuración separada por cédula y tabla del art. 241 E.T.',
         f'={LIQ}!D{A["imp_cargo"]}', TEAL, 'num'),
        ('7 · Alertas', f'{len(C["alertas"])} partidas por resolver antes de presentar',
         f'{n_altas} de sev. alta', RED, 'alert'),
        ('8 · Consignaciones', 'Movimientos bancarios, consumos con tarjeta y compras',
         f'={CONS}!G{A["mov_tope4"]}', GOLD, 'num'),
        ('9 · Detalle exógena', f'Los {C["registros"]} registros fuente con la clasificación asignada a cada uno',
         f'{C["registros"]} registros', GREY, 'text'),
    ]
    for lab, desc, val, acc, kind in indice:
        r = s.datarow(h=24.5)
        s.put(r, 'B', lab, F(9.2, True, color=acc), AL('left', 'center', True, 1), None, 'C')
        s.put(r, 'D', desc, F(9.2, color=TEXT2), AL('left', 'center', True, 1), None, 'G')
        if kind == 'num':
            s.put(r, 'H', val, F(9.2, color=TEXT), AL('right', 'center', False, 1), NUM, 'I')
        elif kind == 'alert':
            s.put(r, 'H', val, F(8.3, True, color=RED), AL('right', 'center', False, 1), None, 'I')
        else:
            s.put(r, 'H', val, F(8.3, True, color=MUTED), AL('right', 'center', False, 1), None, 'I')
    s.gap()

    # ── obligación de declarar ──
    s.section(f'OBLIGACIÓN DE DECLARAR  ·  TOPES CON UVT {ag} = {fmt_pesos(C["uvt"])}')
    s.head([('B', 'C', 'Causal', 'left'), ('D', None, 'Tope UVT', 'center'),
            ('E', None, 'Tope en pesos', 'right'), ('F', None, 'Valor del contribuyente', 'right'),
            ('G', None, 'Resultado', 'center'), ('H', 'I', 'Referencia', 'left')])
    causales = [
        (f'Patrimonio bruto a 31-dic-{ant}', 4500, f'={PAT}!H{A["pat_2024"]}', 'Art. 592 E.T.'),
        ('Ingresos brutos del período', 1400, f'={ING}!G{A["ing_total"]}', 'Art. 592 E.T.'),
        ('Consignaciones, depósitos e inversiones', 1400, f'={CONS}!G{A["mov_tope4"]}', 'Decreto de plazos'),
        ('Consumos con tarjeta de crédito', 1400, f'={CONS}!E{A["consumos"]}', 'Decreto de plazos'),
        ('Compras y consumos totales', 1400, f'={CONS}!E{A["compras"]}', 'Decreto de plazos'),
    ]
    t0 = s.r
    for causal, uvts, ref, norma in causales:
        r = s.datarow()
        s.put(r, 'B', causal, F(9.2, color=TEXT), AL('left', 'center', True, 1), None, 'C')
        s.put(r, 'D', uvts, F(9.2, color=TEXT2), AL('center', 'center'), NUM)
        s.put(r, 'E', f'=ROUND(D{r}*$G$9,0)', F(9.2, color=TEXT2), AL('right', 'center', False, 1), NUM)
        s.put(r, 'F', ref, F(9.2, color=TEXT), AL('right', 'center', False, 1), NUM)
        s.put(r, 'G', f'=IF(N(F{r})>=N(E{r}),"SUPERA","No supera")',
              F(8.3, True, color=MUTED), AL('center', 'center'))
        s.put(r, 'H', norma, F(9.2, color=TEXT2), AL('left', 'center', True, 1), None, 'I')
    t1 = s.r - 1
    ws.conditional_formatting.add(
        f'G{t0}:G{t1}',
        FormulaRule(formula=[f'N(F{t0})>=N(E{t0})'], font=Font(bold=True, color=RED)))
    r = s.r
    ws.row_dimensions[r].height = 19.05
    for ci in range(2, 10):
        c = ws.cell(row=r, column=ci)
        c.fill = FILL(TEAL_T)
        c.border = Border(top=S(TEAL), bottom=S(BORD),
                          left=S(BORD) if ci == 2 else None,
                          right=S(BORD) if ci == 9 else None)
    s.put(r, 'B',
          f'=IF(COUNTIF(G{t0}:G{t1},"SUPERA")>0,"Resultado: OBLIGADO A DECLARAR por "&COUNTIF(G{t0}:G{t1},"SUPERA")&" causales","Resultado: no obligado a declarar por estas causales")',
          F(9.6, True, color=TEXT), AL('left', 'center', False, 1), None, 'G')
    s.put(r, 'H', f'=COUNTIF(G{t0}:G{t1},"SUPERA")&" de {len(causales)}"',
          F(9.2, True, color=TEXT), AL('right', 'center', False, 1), None, 'I')
    s.r = r + 1
    s.gap()

    # ── validación cruzada y advertencia ──
    if T.get('validacion_cruzada'):
        s.callout('VALIDACIÓN CRUZADA DE LA CLASIFICACIÓN', T['validacion_cruzada'], h=35.25)
        s.gap()
    s.callout('DOCUMENTO DE TRABAJO', TEXTO_DOC_TRABAJO, accent=GOLD, tint=GOLD_T, h=35.25)
    s.gap()

    # ── perfil del contribuyente ──
    s.section('PERFIL DEL CONTRIBUYENTE  ·  MARQUE LA CASILLA Y COMPLETE EL DATO')
    hr = s.r
    ws.row_dimensions[hr].height = 19.95
    for ci in range(2, 10):
        c = ws.cell(row=hr, column=ci)
        c.fill = FILL(TEAL)
        c.border = Border(top=S(BORD), left=S(BORD) if ci == 2 else None,
                          right=S(BORD) if ci == 9 else None)
    s.put(hr, 'B', 'Concepto', F(8.3, True, color=WHITE), AL('left', 'center', True, 1), None, 'D')
    s.put(hr, 'E', 'Marque', F(8.3, True, color=WHITE), AL('center', 'center', True))
    s.put(hr, 'F', 'Detalle o cantidad', F(8.3, True, color=WHITE), AL('left', 'center', True, 1), None, 'G')
    s.put(hr, 'H', 'Dónde impacta', F(8.3, True, color=WHITE), AL('left', 'center', True, 1), None, 'I')
    s.r = hr + 1

    chk = DataValidation(type='list', formula1='"✔,OK,—"', allow_blank=True, showDropDown=False)
    chk.prompt = 'Marque con ✔ u OK si la respuesta es afirmativa.'
    chk.promptTitle = 'Perfil del contribuyente'
    ws.add_data_validation(chk)
    fila_hijos = None
    for i, (concepto, detalle, efecto) in enumerate(PERFIL):
        r = s.datarow(zebra=bool(i % 2), h=27.0)
        s.put(r, 'B', concepto, F(9.2, color=TEXT), AL('left', 'center', True, 1), None, 'D')
        box = ws.cell(row=r, column=5)
        box.fill = FILL(IN_FILL)
        box.border = Border(left=Side('medium', color=IN_BORD), right=Side('medium', color=IN_BORD),
                            top=Side('medium', color=IN_BORD), bottom=Side('medium', color=IN_BORD))
        box.font = F(13.0, True, color=TEAL)
        box.alignment = AL('center', 'center')
        chk.add(box)
        for ci in (6, 7):
            cc = ws.cell(row=r, column=ci)
            cc.fill = FILL(IN_FILL)
            cc.border = Border(left=S(IN_BORD), right=S(IN_BORD), top=S(IN_BORD), bottom=S(IN_BORD))
        s.put(r, 'F', None, F(9.0, color=TEXT), AL('left', 'center', True, 1), None, 'G')
        if efecto is None:
            fila_hijos = r
            efecto = (f'="Deducción potencial: "&TEXT(MIN(N(F{r}),4)*72*$G$9,"$ #.##0")'
                      f'&" — 72 UVT c/u, máximo cuatro. Va en la hoja «6. Liquidación»."')
        s.put(r, 'H', efecto, F(7.9, False, True, MUTED), AL('left', 'center', True, 1), None, 'I')
    if fila_hijos:
        ws.cell(row=fila_hijos, column=6).number_format = '0'
    r = s.datarow(h=26.0)
    s.put(r, 'B',
          'Ninguno de estos datos llega por exógena: son la única fuente el contribuyente y sus soportes. Marque la casilla, escriba el detalle y cargue el valor en la hoja indicada — la liquidación no los toma de forma automática.',
          F(8.0, False, True, MUTED), AL('left', 'center', True, 2), None, 'I')
    s.gap(10.0)

    # ── plazo de presentación ──
    r = s.r
    bloques = [
        (f'PLAZO PARA PRESENTAR LA DECLARACIÓN DE RENTA  ·  AÑO GRAVABLE {ag}  ·  ÚLTIMOS DOS DÍGITOS DEL DOCUMENTO: {dig}',
         F(8.0, True, color=ONDARK), AL('left', 'center'), 20.0),
        (plazos.texto_grande(venc) if venc else T['plazo_grande'],
         F(24.0, True, color=TEAL_ACC, name=GEO), AL('left', 'center'), 42.0),
    ]
    for k, (val, fnt, al, hh) in enumerate(bloques):
        rr = r + k
        for ci in range(2, 10):
            ws.cell(row=rr, column=ci).fill = FILL(INK)
        c = ws.cell(row=rr, column=2)
        c.value, c.font, c.alignment = val, fnt, al
        ws.row_dimensions[rr].height = hh
        s.span(rr, 'B', 'I')
    rf = r + 2
    for ci in range(2, 10):
        ws.cell(row=rf, column=ci).fill = FILL(INK)
    ws.row_dimensions[rf].height = 30.0
    s.put(rf, 'B', 'Fecha exacta de vencimiento:', F(11.0, True, color=WHITE, name=GEO),
          AL('left', 'center', False, 1), None, 'D')
    fx = ws.cell(row=rf, column=5)
    fx.fill = FILL(IN_FILL)
    fx.border = Border(left=Side('medium', color=TEAL_ACC), right=Side('medium', color=TEAL_ACC),
                       top=Side('medium', color=TEAL_ACC), bottom=Side('medium', color=TEAL_ACC))
    fx.font = F(12.0, True, color=TEXT)
    fx.alignment = AL('center', 'center')
    fx.number_format = 'dd/mm/yyyy'
    if venc:
        fx.value = venc
        fx.comment = Comment(
            f'Fecha tomada del calendario oficial de plazos del año gravable {ag}\n'
            f'para los últimos dos dígitos {dig} del documento. Es editable: si el\n'
            f'Gobierno modifica el decreto, corrija aquí y el conteo se ajusta solo.', AUT, 320, 70)
        nota = (f'=IF(E{rf}="","",'
                f'IF(TODAY()>E{rf},"Venció hace "&TEXT(TODAY()-E{rf},"0")&" día(s): ya corre la sanción por extemporaneidad.",'
                f'IF(TODAY()=E{rf},"VENCE HOY.",'
                f'"Faltan "&TEXT(E{rf}-TODAY(),"0")&" día(s) calendario. Declaración y pago.")))')
    else:
        fx.comment = Comment(
            f'Escriba aquí la fecha exacta de vencimiento que fije el decreto\n'
            f'de plazos para los últimos dos dígitos {dig} del año gravable {ag}.', AUT, 300, 60)
        nota = f'Tómela del decreto de plazos publicado por el Gobierno para el año gravable {ag}.'
    s.put(rf, 'F', nota, F(9.0, False, True, ONDARK2), AL('left', 'center', True, 1), None, 'I')
    rn = rf + 1
    for ci in range(2, 10):
        ws.cell(row=rn, column=ci).fill = FILL(INK)
    ws.row_dimensions[rn].height = 42.0
    s.put(rn, 'B',
          (f'Es la fecha que el calendario oficial del año gravable {ag} asigna a los documentos terminados en {dig}; en esa misma fecha vencen la presentación y el pago, no solo la presentación. Confírmela en el portal de la DIAN antes de programar el pago: presentar fuera de término genera sanción por extemporaneidad del 5% del impuesto por cada mes o fracción (art. 641 E.T.) más intereses de mora, y ninguna de las dos se puede evitar después.'
           if venc else
           f'La ventana ordinaria de las personas naturales residentes va de agosto a octubre; el día exacto depende de los dos últimos dígitos del documento —en este caso {dig}— y lo fija cada año el decreto de plazos. Confírmelo en el portal de la DIAN antes de programar el pago: presentar fuera de término genera sanción por extemporaneidad del 5% del impuesto por cada mes o fracción (art. 641 E.T.) más intereses de mora, y ninguna de las dos se puede evitar después.'),
          F(8.7, color=ONDARK2), AL('left', 'top', True, 1), None, 'I')

    ws.freeze_panes = 'A6'
    ws.print_area = f'A1:I{rn}'
    return ws
