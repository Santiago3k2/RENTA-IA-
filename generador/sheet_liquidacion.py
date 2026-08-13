# -*- coding: utf-8 -*-
"""Hoja 6 · Liquidación — nota 16 (depuración separada por cédula)."""
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from build_lib import *

import inflacionario

AUT = 'Generador renta'
ING = "'3. Ingresos'"
CON = "'8. Consignaciones'"
UVT = "'1. Resumen'!$G$9"


def build(wb, A, C, T):
    ws = wb.create_sheet('6. Liquidación')
    ws.sheet_properties.tabColor = TEAL[2:]
    s = Sheet(ws, 'B', 'E', TEAL, TEAL_T)
    s.widths({'A': 2.3, 'B': 11.5, 'C': 56.0, 'D': 18.0, 'E': 58.0})
    ag = C['ano_gravable']
    uvt = C['uvt']

    s.band('06  ·  DEPURACIÓN SEPARADA POR CÉDULA',
           'Liquidación privada del impuesto',
           'Cada renta se depura por separado —bloques A, B y C— para que se vean con claridad sus ingresos no constitutivos, sus costos, sus deducciones y sus rentas exentas. El proceso solo se unifica en el bloque F, cuando se llega al impuesto de renta. La primera columna indica el renglón del Formulario 210 cuando el reporte de la DIAN lo identifica; el guion señala partidas cuya casilla depende de la versión del formulario del año.',
           h_desc=46.0)
    s.legend('Las casillas con fondo crema son editables. Todo lo demás son fórmulas enlazadas con las hojas 3 · Ingresos y 8 · Consignaciones: al cambiar un dato allí, esta liquidación se recalcula sola.')

    def blockhead(title, accent=TEAL):
        s.section(title, accent)
        s.head([('B', None, 'Renglón', 'center'), ('C', None, 'Concepto', 'left'),
                ('D', None, 'Valor', 'right'), ('E', None, 'Nota', 'left')], accent=accent)

    def line(reng, concepto, valor, nota=None, ind=3, b=False, zebra=False, h=24.5,
             color=TEXT, fmt=ACC, inp=False):
        r = s.datarow(zebra=zebra, h=h)
        s.put(r, 'B', reng, F(7.8, True, color=MUTED), AL('center', 'center'))
        s.put(r, 'C', concepto, F(9.0 if not b else 9.4, b, color=color),
              AL('left', 'center', True, ind))
        if inp:
            s.mark_input(r, 'D')
        s.put(r, 'D', valor, F(9.0 if not b else 9.6, b, color=color),
              AL('right', 'center', False, 1), fmt)
        if nota:
            s.put(r, 'E', nota, F(7.9, False, True, MUTED), AL('left', 'center', True, 1))
        return r

    def total(concepto, formula, accent=TEAL, tint=TEAL_T, nota=None, sz=10.2, h=24.5,
              fmt=ACC, reng='—'):
        r = s.r
        s.ws.row_dimensions[r].height = h
        for ci in range(s.f, s.l + 1):
            c = s.ws.cell(row=r, column=ci)
            c.fill = FILL(tint)
            c.border = Border(bottom=S(accent),
                              left=S(BORD) if ci == s.f else None,
                              right=S(BORD) if ci == s.l else None)
        s.put(r, 'B', reng, F(7.8, True, color=MUTED), AL('center', 'center'))
        s.put(r, 'C', concepto, F(sz, True, color=TEXT), AL('left', 'center', True, 1))
        s.put(r, 'D', formula, F(sz, True, color=TEXT), AL('right', 'center', False, 1), fmt)
        if nota:
            s.put(r, 'E', nota, F(7.9, False, True, MUTED), AL('left', 'center', True, 1))
        s.r = r + 1
        return r

    # ══════════════ A · RENTAS DE TRABAJO ══════════════
    blockhead('BLOQUE A  ·  DEPURACIÓN DE LAS RENTAS DE TRABAJO', BLUE)
    a1 = line('R32', 'Ingresos brutos por rentas de trabajo', f"={ING}!G{A['r32']}", ind=1, b=True)
    a2 = line('R33', '(−) Ingresos no constitutivos de renta — aportes obligatorios a salud y pensión',
              f"=-{ING}!G{A['r33']}", 'Arts. 55 y 56 E.T. — no se someten al límite del 40%.', zebra=True)
    a3 = total('Renta líquida de las rentas de trabajo', f'=D{a1}+D{a2}', BLUE, BLUE_T)
    line('—', 'Rentas exentas y deducciones imputables — sujetas al límite del art. 336 num. 3',
         None, 'Se detallan aquí y se imputan en el bloque D.', ind=1, b=True)
    ax1 = line('R36', 'Cesantías e intereses de cesantías exentos — art. 206 num. 4',
               f"={ING}!G{A['x_ces']}", T.get('nota_r36', ''), zebra=True, h=28.95)
    ax2 = line('—', 'Renta exenta del 25% — art. 206 num. 10',
               f"={ING}!G{A['x_25']}",
               '25% × (rentas de trabajo − INCRNGO − cesantías exentas), con tope anual de 790 UVT. No aplica a honorarios que detraen costos.',
               h=28.95)
    ax3 = line('—', 'Otras rentas exentas de las rentas de trabajo',
               f"={ING}!G{A['x_otras']}",
               'Aportes voluntarios a pensión y cuentas AFC — arts. 126-1 y 126-4 E.T. Se diligencian en la hoja 3.',
               zebra=True, h=28.95)
    a4 = total('Total rentas exentas y deducciones de las rentas de trabajo',
               f'=SUM(D{ax1}:D{ax3})', BLUE, BLUE_T, 'Se imputa en el bloque D, dentro del límite conjunto.', sz=9.8)
    s.gap()

    # ══════════════ B · RENTAS DE CAPITAL ══════════════
    blockhead('BLOQUE B  ·  DEPURACIÓN DE LAS RENTAS DE CAPITAL', GOLD)
    b1 = line('R58', 'Ingresos brutos por rentas de capital', f"={ING}!G{A['r58']}", ind=1, b=True)
    pct_ci = inflacionario.porcentaje(C.get('ano_gravable'))
    b2 = line('R59', '(−) Ingresos no constitutivos de renta — componente inflacionario',
              f"=-{ING}!G{A['incr58']}",
              ('Art. 38 E.T. — %s%% de los rendimientos financieros del año gravable %s, '
               'porcentaje fijado por decreto. La base se detalla en la hoja 3.'
               % (('%.2f' % (pct_ci * 100)).replace('.', ','), C.get('ano_gravable'))
               ) if pct_ci else
              'Art. 38 E.T. — porcentaje fijado por decreto reglamentario. Se liquida en cero por no estar publicado; ver la sensibilidad al final de la hoja.',
              zebra=True, h=28.95)
    b3 = line('R67', '(−) Costos y deducciones procedentes', f"=-{ING}!G{A['costos58']}",
              'Art. 339 E.T. — comisión de administración inmobiliaria, predial, seguros, reparaciones y depreciación de los inmuebles arrendados. No se informan en exógena y son plenamente procedentes con soporte.',
              h=34.35)
    line('—', 'Control · tope del 60% de los ingresos de esta subcédula',
         f"={ING}!G{A['tope60_cap']}", 'Los costos y deducciones imputables no pueden superar este valor.',
         zebra=True, color=GOLD, h=19.05)
    b4 = total('Renta líquida de las rentas de capital', f'=D{b1}+D{b2}+D{b3}', GOLD, GOLD_T)
    s.gap()

    # ══════════════ C · RENTAS NO LABORALES ══════════════
    blockhead('BLOQUE C  ·  DEPURACIÓN DE LAS RENTAS NO LABORALES', PURPLE)
    c1 = line('R74', 'Ingresos brutos por rentas no laborales', f"={ING}!G{A['r74']}", ind=1, b=True)
    c2 = line('—', '(−) Ingresos no constitutivos de renta', f"=-{ING}!G{A['incr74']}",
              'Arts. 36 a 57-2 E.T. — no se informan en la exógena para esta subcédula.', zebra=True)
    c3 = line('—', '(−) Costos y deducciones procedentes', f"=-{ING}!G{A['costos74']}",
              'Art. 341 E.T. — con relación de causalidad, necesidad y proporcionalidad.')
    line('—', 'Control · tope del 60% de los ingresos de esta subcédula',
         f"={ING}!G{A['tope60_nol']}", 'Los costos y deducciones imputables no pueden superar este valor.',
         zebra=True, color=GOLD)
    c4 = total('Renta líquida de las rentas no laborales', f'=D{c1}+D{c2}+D{c3}', PURPLE, PURPLE_T)
    s.gap()

    # ══════════════ D · LÍMITE DEL 40% ══════════════
    blockhead('BLOQUE D  ·  LÍMITE CONJUNTO DEL ART. 336 NUM. 3', TEAL)
    d0 = line('—', 'Renta líquida de la cédula general  (bloques A + B + C)',
              f'=D{a3}+D{b4}+D{c4}', 'Suma de las tres subcédulas, antes de rentas exentas y deducciones.',
              ind=1, b=True)
    d1 = line('—', 'Cupo máximo: 40% de los ingresos menos los INCRNGO, con tope de 1.340 UVT',
              f"=ROUND(MIN(0.4*({ING}!G{A['ing_total']}-{ING}!G{A['r33']}-{ING}!G{A['incr58']}-{ING}!G{A['incr74']}),1340*{UVT}),0)",
              'Art. 336 num. 3 E.T. — el cupo se calcula sobre la cédula general completa, no por subcédula.',
              zebra=True, h=28.95)
    d2 = line('—', 'Rentas exentas y deducciones solicitadas', f'=D{a4}',
              'Proviene del bloque A. Si se solicitan rentas exentas en otras subcédulas, súmelas aquí.')
    d3 = line('—', 'Exceso rechazado por superar el límite', f'=MAX(0,D{d2}-D{d1})',
              'Si es mayor que cero, la porción que excede el cupo no es deducible.', zebra=True, color=RED)
    d4 = line('—', 'Cupo disponible sin usar', f'=MAX(0,D{d1}-D{d2})',
              'Espacio que aún cabe dentro del límite: intereses de crédito de vivienda, medicina prepagada, aportes voluntarios o cuentas AFC.',
              h=28.95, color=GREEN)
    d5 = total('(−) Rentas exentas y deducciones imputables aceptadas',
               f'=-MIN(D{d2},D{d1})', TEAL, TEAL_T)
    A['exentas_aceptadas'] = d5
    s.gap()

    # ══════════════ E · DEDUCCIONES NO SUJETAS AL LÍMITE ══════════════
    blockhead('BLOQUE E  ·  DEDUCCIONES NO SUJETAS AL LÍMITE DEL 40%', TEAL)
    e1 = line('—', '1% de compras con factura electrónica — art. 336 par. 5',
              f"=ROUND(MIN({CON}!E{A['base_fe']}*0.01,240*{UVT}),0)",
              'Base tomada de la hoja 8: monto de facturación electrónica susceptible del beneficio. Tope de 240 UVT. Requiere pago por medios electrónicos y factura validada por la DIAN.',
              ind=1, h=34.35)
    e2 = line('—', 'Deducción por dependientes — art. 336 par. 2', 0,
              '72 UVT por dependiente, hasta cuatro. No se informa en exógena: confirmar con el contribuyente. Cada dependiente reduce el impuesto en ≈ $681.000.',
              zebra=True, h=28.95, ind=1, inp=True)
    A['dependientes'] = e2
    ws[f'D{e2}'].comment = Comment(
        'Casilla editable. Escriba el valor total de la deducción por\n'
        'dependientes: 72 UVT por dependiente (máximo cuatro).\n'
        'No está sujeta al límite del 40%.', AUT, 300, 80)
    e3 = total('(−) Total deducciones no sujetas al límite', f'=-SUM(D{e1}:D{e2})', TEAL, TEAL_T)
    A['ded_sin_limite'] = e3
    s.gap()

    # ══════════════ F · UNIFICACIÓN ══════════════
    blockhead('BLOQUE F  ·  UNIFICACIÓN — IMPUESTO SOBRE LA RENTA', TEAL)
    f1 = line('—', 'Renta líquida de la cédula general', f'=D{d0}', ind=1, b=True)
    f2 = line('—', '(−) Rentas exentas y deducciones imputables aceptadas', f'=D{d5}', zebra=True)
    f3 = line('—', '(−) Deducciones no sujetas al límite del 40%', f'=D{e3}')
    f4 = total('Renta líquida gravable de la cédula general', f'=D{f1}+D{f2}+D{f3}', TEAL, TEAL_T)
    A['rlg'] = f4
    f5 = line('—', 'Renta presuntiva', 0,
              'Art. 188 E.T. — tarifa del 0% a partir del año gravable 2021. No compite con la renta líquida ordinaria.',
              zebra=True, h=28.95)
    f6 = line('—', 'Base gravable en UVT', f'=D{f4}/{UVT}',
              f'Renta líquida gravable dividida por la UVT del año gravable {ag}.', fmt=UVTF)
    A['uvt_row'] = f6
    f7 = line('—', 'Impuesto sobre la renta líquida gravable — art. 241 E.T.',
              f'=ROUND(IF(D{f6}<=1090,0,IF(D{f6}<=1700,(D{f6}-1090)*0.19,'
              f'IF(D{f6}<=4100,(D{f6}-1700)*0.28+116,IF(D{f6}<=8670,(D{f6}-4100)*0.33+788,'
              f'IF(D{f6}<=18970,(D{f6}-8670)*0.35+2296,IF(D{f6}<=31000,(D{f6}-18970)*0.37+5901,'
              f'(D{f6}-31000)*0.39+10352))))))*{UVT}/1000,0)*1000',
              'Tarifa marginal aplicada según la tabla del final de la hoja; el resultado se aproxima al múltiplo de mil más cercano (art. 577 E.T.).',
              zebra=True, h=28.95)
    A['imp_241'] = f7
    f8 = line('—', '(−) Descuentos tributarios', 0,
              'Donaciones (art. 257), impuestos pagados en el exterior (art. 254) u otros descuentos. No se informan en la exógena.',
              h=28.95, inp=True)
    A['descuentos'] = f8
    f9 = total('Impuesto neto de renta', f'=D{f7}-D{f8}', TEAL, TEAL_T)
    A['imp_neto'] = f9
    f10 = line('—', '(+) Impuesto de ganancias ocasionales', 0,
               'No se reportan enajenaciones de activos fijos, herencias, donaciones, loterías ni rifas.',
               zebra=True, inp=True)
    r = s.r
    s.ws.row_dimensions[r].height = 25.95
    for ci in range(s.f, s.l + 1):
        c = s.ws.cell(row=r, column=ci)
        c.fill = FILL(TEAL)
        c.border = Border(bottom=S(BORD),
                          left=S(BORD) if ci == s.f else None,
                          right=S(BORD) if ci == s.l else None)
    s.put(r, 'B', '—', F(7.8, True, color=WHITE), AL('center', 'center'))
    s.put(r, 'C', 'TOTAL IMPUESTO A CARGO', F(11.5, True, color=WHITE, name=GEO), AL('left', 'center', False, 1))
    s.put(r, 'D', f'=D{f9}+D{f10}', F(12.0, True, color=WHITE), AL('right', 'center', False, 1), MONEY)
    A['imp_cargo'] = r
    s.r = r + 1
    s.gap(); s.gap()

    # ══════════════ SENSIBILIDAD ══════════════
    if T.get('sensibilidad'):
        s.callout('SENSIBILIDAD DEL COMPONENTE INFLACIONARIO', T['sensibilidad'], h=42.0)
        s.gap()

    # ══════════════ TABLA ART. 241 ══════════════
    s.section(f'TABLA DEL ARTÍCULO 241 E.T. — AÑO GRAVABLE {ag}')
    s.head([('B', 'C', 'Rango en UVT', 'left'), ('D', None, 'Tarifa marginal', 'center'),
            ('E', None, 'Rango en pesos y fórmula del impuesto', 'left')])
    brackets = [
        (0, 1090, '0%', 'impuesto cero'),
        (1090, 1700, '19%', '(Base UVT − 1.090) × 19%'),
        (1700, 4100, '28%', '(Base UVT − 1.700) × 28% + 116 UVT'),
        (4100, 8670, '33%', '(Base UVT − 4.100) × 33% + 788 UVT'),
        (8670, 18970, '35%', '(Base UVT − 8.670) × 35% + 2.296 UVT'),
        (18970, 31000, '37%', '(Base UVT − 18.970) × 37% + 5.901 UVT'),
        (31000, None, '39%', '(Base UVT − 31.000) × 39% + 10.352 UVT'),
    ]
    uvtc = f'$D${f6}'
    for i, (lo, hi, tar, formula) in enumerate(brackets):
        if lo == 0:
            rango = f'>0 a {fnum(hi)}'
            pesos = f'0 a {fnum(hi * uvt)}  ·  {formula}'
            cond = f'{uvtc}<={hi}'
        elif hi is None:
            rango = f'>{fnum(lo)} en adelante'
            pesos = f'más de {fnum(lo * uvt)}  ·  {formula}'
            cond = f'{uvtc}>{lo}'
        else:
            rango = f'>{fnum(lo)} a {fnum(hi)}'
            pesos = f'{fnum(lo * uvt + 1)} a {fnum(hi * uvt)}  ·  {formula}'
            cond = f'AND({uvtc}>{lo},{uvtc}<={hi})'
        r = s.datarow(zebra=bool(i % 2))
        s.txt(r, 'B', rango, 'C', v='center')
        s.put(r, 'D', tar, F(9.2, color=TEXT2), AL('center', 'center'))
        s.txt(r, 'E', pesos, sz=8.8, color=TEXT2, v='center')
        ws.conditional_formatting.add(
            f'B{r}:E{r}',
            FormulaRule(formula=[cond], fill=FILL(TEAL_T), font=Font(bold=True, color=TEAL)))
    s.note('La fila resaltada es el rango aplicable al contribuyente y se marca sola a partir de la base gravable en UVT calculada en el bloque F.')

    ws.freeze_panes = 'A6'
    return ws
