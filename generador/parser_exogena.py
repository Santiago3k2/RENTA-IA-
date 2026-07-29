# -*- coding: utf-8 -*-
"""Lee el reporte de exógena del prevalidador MUISCA y lo vuelve estructura.

No interpreta nada tributario: solo extrae identificación, topes precalculados
y los registros con su metadata. La interpretación va en clasificador.py.
"""
import datetime
import re
import unicodedata

import openpyxl

# Concepto Códigos Tributaria → tipo de cuenta legible
TIPO_CUENTA = {
    '1': 'Cuenta de ahorros',
    '2': 'Cuenta corriente',
    '3': 'Cuenta de ahorros',
    '5': 'Depósito electrónico',
    '6': 'Cartera colectiva (FIC)',
}


def sinacento(t):
    t = unicodedata.normalize('NFD', str(t or ''))
    return ''.join(c for c in t if unicodedata.category(c) != 'Mn')


def norm(t):
    """Minúsculas, sin acentos, espacios colapsados — para comparar textos."""
    return re.sub(r'\s+', ' ', sinacento(t).lower()).strip()


def _meta(info):
    """'Clave: valor | Clave2: valor2' → dict con claves normalizadas."""
    out = {}
    for parte in str(info or '').split('|'):
        if ':' not in parte:
            continue
        k, v = parte.split(':', 1)
        out[norm(k)] = v.strip()
    return out


def _num(v):
    """Valores monetarios: llegan como número, o en formato colombiano."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = re.sub(r'[^\d,.\-]', '', str(v)).replace('.', '').replace(',', '.')
    try:
        return float(t)
    except ValueError:
        return None


def _num_en(v):
    """Metadata de la columna «Información Adicional»: viene en formato inglés
    (punto decimal) y a veces en notación científica — «4.17», «1.3624193955E10».
    Usar _num aquí convertiría 4.17 en 417."""
    if v is None or v == '':
        return None
    try:
        return float(str(v).strip().replace(',', ''))
    except ValueError:
        return None


class Registro:
    __slots__ = ('nit', 'reportante', 'detalle', 'valor', 'uso', 'info', 'meta',
                 'fila', 'cuenta', 'tipo_cuenta', 'matricula', 'placa',
                 'participacion', 'secundario')

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k))

    def __repr__(self):
        return f'<{self.detalle[:38]}… {self.valor:,.0f}>'


def leer(path):
    """Devuelve dict con identificación, topes y registros."""
    if str(path).lower().endswith('.xls'):
        raise ValueError(
            'El archivo está en formato .xls antiguo. Ábralo en Excel y guárdelo '
            'como .xlsx (Libro de Excel) antes de subirlo.')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    # ── cabecera: se busca la fila que tiene 'Detalle' en la columna E ──
    hdr = None
    for r in range(1, 40):
        if norm(ws.cell(r, 5).value) == 'detalle':
            hdr = r
            break
    if hdr is None:
        raise ValueError('No parece un reporte de exógena de la DIAN: no se '
                         'encontró la fila de encabezados con la columna «Detalle».')

    ident = {'identificacion': None, 'nombre': None, 'ano': None,
             'tipo_doc': 'C.C.', 'fecha_corte': None, 'fecha_reporte': None}
    for r in range(1, hdr):
        etiqueta = norm(ws.cell(r, 1).value)
        val = ws.cell(r, 3).value
        if etiqueta.startswith('identificacion:'):
            ident['identificacion'] = str(val).strip() if val else None
        elif etiqueta.startswith('nombres'):
            ident['nombre'] = str(val).strip() if val else None
        elif etiqueta.startswith('ano al que se refiere'):
            ident['ano'] = str(val).strip() if val else None
        elif etiqueta.startswith('tipo de documento'):
            ident['tipo_doc'] = str(val).strip() if val else 'C.C.'
        elif etiqueta.startswith('fecha corte'):
            ident['fecha_corte'] = val
        for c in (7, 8):
            if norm(ws.cell(r, c).value).startswith('fecha') and 'reporte' in norm(ws.cell(r, c).value):
                ident['fecha_reporte'] = ws.cell(r, c + 1).value

    # ── topes precalculados por la DIAN ──
    topes = {}
    LLAVES = {'tope 1': 'ingresos', 'tope 2': 'patrimonio', 'tope 3': 'consumos',
              'tope 4': 'movimientos', 'tope 5': 'compras'}
    registros = []
    for r in range(hdr + 1, ws.max_row + 1):
        det = ws.cell(r, 5).value
        if det is None or str(det).strip() == '':
            continue
        d = norm(det)
        marcado = False
        for pref, llave in LLAVES.items():
            if d.startswith(pref) and _num(ws.cell(r, 6).value) is not None:
                topes[llave] = int(_num(ws.cell(r, 6).value))
                marcado = True
                break
        if marcado:
            continue
        valor = _num(ws.cell(r, 6).value)
        if valor is None:
            continue
        info = ws.cell(r, 8).value
        m = _meta(info)
        cuenta = m.get('numero de cuenta / documento') or m.get('numero de cuenta')
        tipo = None
        cct = m.get('concepto codigos tributaria', '')
        mm = re.search(r'\*(\d+)\*', cct)
        if mm:
            tipo = TIPO_CUENTA.get(mm.group(1))
        if not tipo and cct:
            tipo = re.sub(r'^\*\d+\*\s*', '', cct).strip().capitalize() or None
        # Porcentaje de participación. «Posición decimal» indica cuántos decimales
        # trae implícitos el número: se acepta solo cuando el resultado es un
        # porcentaje plausible (0 < x <= 100). En los demás casos se deja None y
        # el valor reportado se toma como propio del contribuyente.
        part = None
        pv = _num_en(m.get('porcentaje de participacion'))
        pd = _num_en(m.get('porcentaje de participacion posicion decimal'))
        if pv is not None:
            crudo = pv / (10 ** int(pd)) if pd else pv
            if 0 < crudo <= 100:
                part = round(crudo / 100.0, 8)
        registros.append(Registro(
            nit=str(ws.cell(r, 1).value or '').strip(),
            reportante=str(ws.cell(r, 2).value or '').strip(),
            detalle=str(det).strip(),
            valor=valor,
            uso=str(ws.cell(r, 7).value or '').strip(),
            info=str(info or '').strip(),
            meta=m, fila=r, cuenta=cuenta, tipo_cuenta=tipo,
            matricula=m.get('matricula'), placa=m.get('placa'),
            participacion=part,
            secundario=('secundario' in norm(det) or 'secundario' in norm(info)),
        ))

    if not registros:
        raise ValueError('El reporte se leyó pero no contiene registros de terceros.')

    def _fecha(v):
        if isinstance(v, datetime.datetime):
            return v.date()
        if isinstance(v, datetime.date):
            return v
        return None

    ident['fecha_corte'] = _fecha(ident['fecha_corte'])
    ident['fecha_reporte'] = _fecha(ident['fecha_reporte'])
    return {'ident': ident, 'topes': topes, 'registros': registros}
