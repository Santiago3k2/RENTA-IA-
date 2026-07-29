# -*- coding: utf-8 -*-
"""Hoja 5 · Anticipo — art. 807 E.T., alimentada por datos.py."""
from openpyxl.comments import Comment
from build_lib import *

AUT = 'Generador renta'


def build(wb, A, C, T):
    ws = wb.create_sheet('5. Anticipo')
    ws.sheet_properties.tabColor = GOLD[2:]
    s = Sheet(ws, 'B', 'J', GOLD, GOLD_T)
    s.widths({'A': 2.3, 'B': 24.0, 'C': 13.0, 'D': 13.0, 'E': 14.0,
              'F': 3.0, 'G': 24.0, 'H': 13.0, 'I': 13.0, 'J': 14.0})
    ag, ant, sig = C['ano_gravable'], C['ano_anterior'], C['ano_siguiente']
    pct = C.get('porcentaje_anticipo', 0.75)
    pct_lab = f'{int(round(pct * 100))}%'
    imp = f"'6. Liquidación'!D{A['imp_neto']}"
    cargo = f"'6. Liquidación'!D{A['imp_cargo']}"
    rtot = f"'4. Retenciones'!D{A['ret_total']}"
    sf = f"'4. Retenciones'!G{A['saldo_favor']}"
    ap = f"'4. Retenciones'!G{A['ant_prev']}"

    s.band('05  ·  ARTÍCULO 807 E.T.',
           f'Liquidación del anticipo para el año gravable {sig}',
           T['sub_anticipo'], h_desc=35.55)
    s.gap()

    # ══════════════ LOS DOS MÉTODOS (sin rótulo, a pedido del usuario) ══════════════
    r8 = s.r
    e9 = r8 + 1
    t0 = r8 + 10          # primera fila de la tabla de sensibilidad
    cajas = [(2, 5, 'E', GOLD), (7, 10, 'J', GREEN)]
    ws.row_dimensions[r8].height = 19.95
    for c1, c2, _, fill in cajas:
        for ci in range(c1, c2 + 1):
            cc = ws.cell(row=r8, column=ci)
            cc.fill = FILL(fill)
            cc.border = Border(top=S(fill),
                               left=S(fill) if ci == c1 else None,
                               right=S(fill) if ci == c2 else None)
    s.put(r8, 'B', 'MÉTODO 1 · IMPUESTO DEL AÑO', F(9.0, True, color=WHITE),
          AL('left', 'center', False, 1), None, 'E')
    s.put(r8, 'G', 'MÉTODO 2 · PROMEDIO DE DOS AÑOS — RECOMENDADO', F(9.0, True, color=WHITE),
          AL('left', 'center', False, 1), None, 'J')

    izq = [(f'Impuesto neto de renta AG {ag}', f'={imp}'),
           (f'× {pct_lab}', f'=ROUND(E{e9}*{pct},0)'),
           (f'(−) Retenciones en la fuente AG {ag}', f'=-{rtot}')]
    der = [(f'Impuesto neto de renta AG {ag}', f'={imp}', 'n'),
           (f'Impuesto neto de renta AG {ant}  (X)  — escenario base', f'=B{t0}', 'x'),
           (f'((Impuesto AG {ag} + X) ÷ 2) × {pct_lab}  −  retenciones', 'ver tabla', 't')]
    for i in range(3):
        r = r8 + 1 + i
        ws.row_dimensions[r].height = 19.0
        for c1, c2, vc, fill in cajas:
            for ci in range(c1, c2 + 1):
                cc = ws.cell(row=r, column=ci)
                cc.fill = FILL(WHITE)
                cc.border = Border(bottom=S(BORD_IN),
                                   left=S(fill) if ci == c1 else None,
                                   right=S(fill) if ci == c2 else None)
        lab, val = izq[i]
        s.put(r, 'B', lab, F(8.8, color=TEXT2), AL('left', 'center', True, 1), None, 'D')
        s.put(r, 'E', val, F(9.0, color=TEXT), AL('right', 'center', False, 1), NUM)
        lab, val, kind = der[i]
        s.put(r, 'G', lab, F(8.8, color=TEXT2), AL('left', 'center', True, 1), None, 'I')
        if kind == 't':
            s.put(r, 'J', val, F(8.8, False, True, MUTED), AL('right', 'center', False, 1))
        else:
            s.put(r, 'J', val, F(9.0, kind == 'x', color=RED if kind == 'x' else TEXT),
                  AL('right', 'center', False, 1), NUM)
    rh = r8 + 4
    ws.row_dimensions[rh].height = 22.05
    for c1, c2, vc, fill in cajas:
        for ci in range(c1, c2 + 1):
            cc = ws.cell(row=rh, column=ci)
            cc.fill = FILL(fill)
            cc.border = Border(bottom=S(fill),
                               left=S(fill) if ci == c1 else None,
                               right=S(fill) if ci == c2 else None)
    s.put(rh, 'B', 'ANTICIPO MÉTODO 1', F(9.5, True, color=WHITE, name=GEO),
          AL('left', 'center', False, 1), None, 'D')
    s.put(rh, 'E', f'=ROUND(MAX(0,E{e9 + 1}+E{e9 + 2})/1000,0)*1000',
          F(11.0, True, color=WHITE), AL('right', 'center', False, 1), MONEY)
    s.put(rh, 'G', 'ANTICIPO MÉTODO 2  (si X = 0)', F(9.5, True, color=WHITE, name=GEO),
          AL('left', 'center', False, 1), None, 'I')
    s.put(rh, 'J', f'=C{t0}', F(11.0, True, color=WHITE), AL('right', 'center', False, 1), MONEY)
    met1 = rh

    s.r = rh + 1
    s.gap(5.0)
    r = s.r
    ws.row_dimensions[r].height = 27.0
    s.put(r, 'B', T['nota_x_anticipo'], F(8.2, False, True, MUTED), AL('left', 'top', True, 1), None, 'J')
    s.r = r + 1
    s.gap()

    # ══════════════ TABLA DE SENSIBILIDAD ══════════════
    s.section(f'ANTICIPO Y SALDO A PAGAR SEGÚN EL IMPUESTO NETO DEL AG {ant}')
    hr = s.r
    ws.row_dimensions[hr].height = 19.95
    for ci in range(2, 10):
        cc = ws.cell(row=hr, column=ci)
        cc.fill = FILL(GOLD)
        cc.border = Border(top=S(BORD), left=S(BORD) if ci == 2 else None,
                           right=S(BORD) if ci == 9 else None)
    s.put(hr, 'B', f'Impuesto neto AG {ant} (X)', F(8.3, True, color=WHITE), AL('right', 'center', True, 1))
    s.put(hr, 'C', 'Anticipo método 2', F(8.3, True, color=WHITE), AL('center', 'center', True), None, 'D')
    s.put(hr, 'E', 'Anticipo método 1', F(8.3, True, color=WHITE), AL('center', 'center', True), None, 'F')
    s.put(hr, 'G', 'Saldo a pagar', F(8.3, True, color=WHITE), AL('center', 'center', True))
    s.put(hr, 'H', 'Método a elegir', F(8.3, True, color=WHITE), AL('center', 'center', True), None, 'I')
    s.r = hr + 1
    assert s.r == t0, f'la tabla debía empezar en {t0} y empezó en {s.r}'

    filas_x = [0, 1000000, 2000000, 3000000, 4000000, None]
    for i, x in enumerate(filas_x):
        r = s.r
        ws.row_dimensions[r].height = 16.05
        for ci in range(2, 10):
            cc = ws.cell(row=r, column=ci)
            cc.fill = FILL(ZEBRA if i % 2 else WHITE)
            cc.border = Border(bottom=S(BORD_IN), left=S(BORD) if ci == 2 else None,
                               right=S(BORD) if ci == 9 else None)
        if x is None:
            s.put(r, 'B', f'=TEXT($E${e9},"#.##0")&" o más"', F(9.2, color=TEXT),
                  AL('right', 'center', False, 1))
            xref = f'$E${e9}'
        else:
            s.mark_input(r, 'B')
            s.put(r, 'B', x if x else C.get('impuesto_neto_anterior', 0),
                  F(9.2, color=TEXT), AL('right', 'center', False, 1), NUM)
            xref = f'B{r}'
        s.put(r, 'C', f'=ROUND(MAX(0,($E${e9}+{xref})/2*{pct}-{rtot})/1000,0)*1000',
              F(9.2, True, color=GREEN), AL('right', 'center', False, 1), NUM, 'D')
        s.put(r, 'E', f'=$E${met1}', F(9.2, color=TEXT2), AL('right', 'center', False, 1), NUM, 'F')
        s.put(r, 'G', f'=ROUND(($E${e9}+MIN(C{r},E{r})-{sf}-{rtot}-{ap})/1000,0)*1000',
              F(9.2, color=TEXT), AL('right', 'center', False, 1), NUM)
        s.put(r, 'H', f'=IF(C{r}<E{r},"MÉTODO 2",IF(C{r}=E{r},"INDIFERENTE","MÉTODO 1"))',
              F(8.3, True, color=GREEN), AL('center', 'center'), None, 'I')
        s.r = r + 1
    ws[f'B{t0}'].comment = Comment(
        f'Casilla editable. Escriba el impuesto neto de renta del AG {ant}\n'
        'tomado de esa declaración: el escenario base y el saldo a pagar\n'
        'de esta hoja se recalculan con este valor.', AUT, 300, 80)
    if T.get('nota_tabla_anticipo'):
        s.note(T['nota_tabla_anticipo'], h=29.25)
    s.gap(); s.gap()

    # ══════════════ DETERMINACIÓN DEL SALDO ══════════════
    s.section('DETERMINACIÓN DEL SALDO — ESCENARIO BASE (MÉTODO 2, X = 0)')
    s.head([('B', 'H', 'Concepto', 'left'), ('I', 'J', 'Valor', 'right')])
    r1 = s.datarow()
    s.txt(r1, 'B', 'Total impuesto a cargo', 'H', v='center')
    s.money(r1, 'I', f'={cargo}', 'J')
    r2 = s.datarow(zebra=True)
    s.txt(r2, 'B', f'(+) Anticipo por el año gravable {sig}  ·  método 2, X = 0', 'H', v='center')
    s.money(r2, 'I', f'=C{t0}', 'J')
    A['ant_row'] = r2
    r3 = s.subtotal('Subtotal antes de créditos', f'=I{r1}+I{r2}', lab_col='B', val_col='I', h=19.05)
    s.span(r3, 'I', 'J')
    r4 = s.datarow()
    s.txt(r4, 'B', f'R131 · (−) Saldo a favor del AG {ant} sin devolución ni compensación', 'H', v='center')
    s.money(r4, 'I', f'=-{sf}', 'J')
    r5 = s.datarow(zebra=True)
    s.txt(r5, 'B', f'R132 · (−) Retenciones en la fuente del AG {ag}', 'H', v='center')
    s.money(r5, 'I', f'=-{rtot}', 'J')
    r6 = s.datarow()
    s.txt(r6, 'B', f'(−) Anticipo del AG {ag} liquidado en la declaración anterior  ·  hoja 4', 'H', v='center')
    s.put(r6, 'I', f'=-{ap}', F(9.2, color=TEXT), AL('right', 'center', False, 1), NUM, 'J')
    if T.get('nota_saldo_anticipo'):
        s.note(T['nota_saldo_anticipo'])
    A['saldo_row'] = s.hero('SALDO A PAGAR', f'=ROUND(SUM(I{r3}:I{r6})/1000,0)*1000', val_col='I', h=24.0)
    s.span(A['saldo_row'], 'I', 'J')
    s.gap(); s.gap()

    if T.get('callout_rango'):
        titulo, cuerpo = T['callout_rango']
        s.callout(titulo, cuerpo, h=47.85)

    ws.freeze_panes = 'A6'
    return ws
